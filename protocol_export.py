"""Génération de la feuille de protocole Excel (feuille de paillasse).

Pour chaque réactif (Solution A–F, Échantillon, Eau de contrôle) et chaque tube,
le tableau comporte :
  - la valeur de volume (µL)
  - une case "Coordinateur" (à cocher lors de l'annonce de la quantité)
  - une case "Opérateur"    (à cocher lors de l'ajout effectif)

La mise en forme suit le modèle CitizenSers :
  en-têtes pivotés à 90°, couleurs par rôle (volume / coordinateur / opérateur),
  colonne "Réactifs" verticale orange, cases à cocher ☐.
"""

from __future__ import annotations

import copy
import os
import sys

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import metadata_model as mm


def _resource_dir() -> str:
    """Retourne le dossier contenant les ressources embarquées.

    - En mode PyInstaller (onefile/onedir) : sys._MEIPASS (dossier temporaire
      où PyInstaller extrait les fichiers data).
    - En mode normal (script Python) : dossier du fichier source.
    """
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        return sys._MEIPASS
    return os.path.dirname(os.path.abspath(__file__))


# Chemin du modèle de mise en page (feuille de protocole)
_TEMPLATE_PATH = os.path.join(_resource_dir(), "modele_tableau.xlsx")

# ── Palette ───────────────────────────────────────────────────────────────────
_TUBE_HDR    = PatternFill("solid", fgColor="2E75B6")  # bleu foncé  — volume
_COORD_HDR   = PatternFill("solid", fgColor="5B9BD5")  # bleu moyen  — coordinateur
_OPER_HDR    = PatternFill("solid", fgColor="9DC3E6")  # bleu clair  — opérateur
_REACT_SIDE  = PatternFill("solid", fgColor="ED7D31")  # orange      — label Réactifs
_VERIF_HDR   = PatternFill("solid", fgColor="70AD47")  # vert        — vérification
_VERIF_CELL  = PatternFill("solid", fgColor="E2EFDA")  # vert clair
_META_HDR    = PatternFill("solid", fgColor="D6DCE4")  # gris        — nom/conc/unité
_ROW_A       = PatternFill("solid", fgColor="DEEAF1")  # bleu très clair (lignes paires)
_ROW_B       = PatternFill("solid", fgColor="FFFFFF")  # blanc       (lignes impaires)

# ── Polices ───────────────────────────────────────────────────────────────────
_WHITE_BOLD  = Font(color="FFFFFF", bold=True, size=9)
_HDR_FONT    = Font(color="FFFFFF", bold=True, size=8)
_DARK        = Font(color="000000", size=9)
_BOLD        = Font(bold=True, size=9)
_CHECK_FONT  = Font(size=13)

# ── Bordures ──────────────────────────────────────────────────────────────────
_T  = Side(style="thin",   color="BFBFBF")
_M  = Side(style="medium", color="808080")
_THIN_BORDER = Border(left=_T, right=_T, top=_T, bottom=_T)
_MED_BORDER  = Border(left=_M, right=_M, top=_M, bottom=_M)

# ── Alignements ───────────────────────────────────────────────────────────────
_ROT    = Alignment(text_rotation=90, horizontal="center", vertical="bottom",  wrap_text=True)
_VROT   = Alignment(text_rotation=90, horizontal="center", vertical="center",  wrap_text=True)
_CTR    = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# ── Symbole case à cocher ─────────────────────────────────────────────────────
_BOX = "☐"


# ── Helpers ───────────────────────────────────────────────────────────────────

def _s(ws, row: int, col: int, value=None, *, fill=None, font=None,
       align=None, border=None):
    """Écriture rapide d'une cellule avec styles optionnels."""
    c = ws.cell(row=row, column=col)
    if value is not None:
        c.value = value
    if fill   is not None: c.fill      = fill
    if font   is not None: c.font      = font
    if align  is not None: c.alignment = align
    if border is not None: c.border    = border
    return c


def _col_letter(n: int) -> str:
    return get_column_letter(n)


def _apply_template_settings(ws, tmpl_path: str,
                              last_col: int, last_row: int,
                              col_first_tube: int = 7) -> bool:
    """Copie la mise en page de la feuille 'Protocole' du modèle vers *ws*.

    Copie intégralement : orientation, papier, échelle, marges, fitToPage,
    largeur par défaut des colonnes, largeurs explicites, et toutes les colonnes
    tubes (vol + coord + opér).  Zone d'impression recalculée dynamiquement.
    """
    try:
        tmpl_wb = openpyxl.load_workbook(tmpl_path)
    except Exception:
        return False

    tmpl_ws = (tmpl_wb["Protocole"] if "Protocole" in tmpl_wb.sheetnames
               else tmpl_wb.active)
    tps = tmpl_ws.page_setup
    tpm = tmpl_ws.page_margins
    tsf = tmpl_ws.sheet_format

    # ── Page setup complet ────────────────────────────────────────────────────
    if tps.orientation:   ws.page_setup.orientation = tps.orientation
    if tps.paperSize:     ws.page_setup.paperSize   = tps.paperSize
    if tps.scale:         ws.page_setup.scale        = tps.scale
    ws.page_setup.fitToWidth  = tps.fitToWidth
    ws.page_setup.fitToHeight = tps.fitToHeight

    # ── fitToPage (sheet_properties) ─────────────────────────────────────────
    try:
        from openpyxl.worksheet.properties import PageSetupProperties
        tpp = (tmpl_ws.sheet_properties.pageSetUpPr
               if tmpl_ws.sheet_properties else None)
        if tpp and tpp.fitToPage:
            if ws.sheet_properties.pageSetUpPr is None:
                ws.sheet_properties.pageSetUpPr = PageSetupProperties()
            ws.sheet_properties.pageSetUpPr.fitToPage = True
    except Exception:
        pass

    # ── Marges (attribut par attribut pour éviter le partage de référence) ───
    for attr in ("left", "right", "top", "bottom", "header", "footer"):
        val = getattr(tpm, attr, None)
        if val is not None:
            setattr(ws.page_margins, attr, val)

    # ── Largeur par défaut de colonne ─────────────────────────────────────────
    if tsf.defaultColWidth:
        ws.sheet_format.defaultColWidth = tsf.defaultColWidth

    # ── Largeurs explicites (colonnes fixes du modèle) ────────────────────────
    for col_letter, dim in tmpl_ws.column_dimensions.items():
        if dim.width:
            ws.column_dimensions[col_letter].width = dim.width

    # ── Colonnes tubes : appliquer la largeur de la colonne G à tous les vol,
    #    et une largeur réduite aux sous-colonnes coord/opér ──────────────────
    g_dim = tmpl_ws.column_dimensions.get("G")
    tube_vol_w = g_dim.width if (g_dim and g_dim.width) else 4.66
    for col_idx in range(col_first_tube, last_col + 1):
        rel = (col_idx - col_first_tube) % 3   # 0=vol, 1=coord, 2=opér
        w   = tube_vol_w if rel == 0 else 3.5
        cl  = get_column_letter(col_idx)
        ws.column_dimensions[cl].width = w

    # ── Zone d'impression dynamique ───────────────────────────────────────────
    ws.print_area = f"$A$1:${get_column_letter(last_col)}${last_row}"

    return True


def _autofit_columns(ws, min_width: int = 3, max_width: int = 30, padding: int = 1) -> None:
    """Ajuste chaque colonne à la largeur minimale nécessaire pour son contenu."""
    for col_cells in ws.columns:
        max_len = min_width
        for cell in col_cells:
            if cell.value is None:
                continue
            lines = str(cell.value).split("\n")
            cell_max = max(len(ln) for ln in lines) if lines else 0
            max_len = max(max_len, cell_max)
        col_letter = col_cells[0].column_letter
        ws.column_dimensions[col_letter].width = min(max_width, max_len + padding)


# ── Fonction principale ───────────────────────────────────────────────────────

_CHECK_ON  = "✓"   # case cochée dans l'export


def _write_df_to_sheet(wb, sheet_name: str, df) -> None:
    """Écrit un DataFrame dans une feuille openpyxl (données brutes, sans style)."""
    import pandas as pd
    ws = wb.create_sheet(sheet_name)
    for c_idx, col in enumerate(df.columns, 1):
        ws.cell(1, c_idx).value = str(col)
    for r_idx, row in enumerate(df.itertuples(index=False), 2):
        for c_idx, val in enumerate(row, 1):
            # Convertir NaN/NaT en chaîne vide
            if val is None or (isinstance(val, float) and val != val):
                val = ""
            ws.cell(r_idx, c_idx).value = val


def export_protocol_excel(df_comp, meta: dict, path: str,
                           states: dict | None = None,
                           df_map=None) -> None:
    """Génère le fichier Excel complet (protocole + métadonnées).

    Le fichier produit contient 4 feuilles :
      - Protocole      : feuille de paillasse stylée (cases à cocher)
      - Volumes        : df_comp brut (rechargeable comme métadonnées)
      - Correspondance : df_map brut (rechargeable comme métadonnées)
      - EtatProtocole  : états des cases (rechargeable par l'application)

    Parameters
    ----------
    df_comp : pd.DataFrame
        Tableau des volumes (colonnes : Réactif, Concentration, Unité, Tube 1, …)
    meta : dict
        Informations d'en-tête : nom_manip, date, lieu, coordinateur, operateur
    path : str
        Chemin de sortie (.xlsx)
    states : dict | None
        États des cases à cocher issus de ProtocolDialog.get_states().
        Clés : ("verif", r_idx) | ("coord", r_idx, t_idx) | ("oper", r_idx, t_idx).
    df_map : pd.DataFrame | None
        Tableau de correspondance spectres ↔ tubes (optionnel).
    """
    tube_cols = mm.get_tube_columns(df_comp)
    n_tubes   = len(tube_cols)
    states    = states or {}

    def _box(key) -> str:
        return _CHECK_ON if states.get(key, False) else _BOX

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Protocole"

    # ── Indices de colonnes ───────────────────────────────────────────────────
    COL_LABEL      = 1   # "Réactifs" — label vertical orange (fusionné)
    COL_REACT      = 2   # nom du réactif
    COL_CONC       = 3   # concentration
    COL_UNIT       = 4   # unité concentration
    COL_VERIF      = 5   # vérification solution (☐)
    COL_VOL_UNIT   = 6   # unité volume (µL ou gouttes)
    COL_FIRST_TUBE = 7   # début des colonnes tubes

    # ── Indices de lignes ─────────────────────────────────────────────────────
    ROW_META1  = 1   # nom manip / date / lieu
    ROW_META2  = 2   # coordinateur / opérateur
    ROW_SEP    = 3   # ligne vide
    ROW_HDR    = 4   # en-têtes du tableau (pivotés)
    ROW_DATA   = 5   # première ligne de données

    n_data = len(df_comp)
    last_data_row = ROW_DATA + n_data - 1

    # ── En-tête d'information manip ───────────────────────────────────────────
    info_line1 = [
        ("Nom de la manip :", meta.get("nom_manip", "")),
        ("Date :",            meta.get("date", "")),
        ("Lieu :",            meta.get("lieu", "")),
    ]
    info_line2 = [
        ("Coordinateur :", meta.get("coordinateur", "")),
        ("Opérateur :",   meta.get("operateur", "")),
    ]
    for i, (label, val) in enumerate(info_line1):
        ws.cell(ROW_META1, COL_REACT + i * 4).value = label
        ws.cell(ROW_META1, COL_REACT + i * 4).font  = _BOLD
        ws.cell(ROW_META1, COL_REACT + i * 4 + 1).value = val
        ws.cell(ROW_META1, COL_REACT + i * 4 + 1).font  = _DARK
    for i, (label, val) in enumerate(info_line2):
        ws.cell(ROW_META2, COL_REACT + i * 4).value = label
        ws.cell(ROW_META2, COL_REACT + i * 4).font  = _BOLD
        ws.cell(ROW_META2, COL_REACT + i * 4 + 1).value = val
        ws.cell(ROW_META2, COL_REACT + i * 4 + 1).font  = _DARK

    # ── Colonne "Réactifs" — fusionnée verticalement sur toute la zone données ─
    ws.merge_cells(
        start_row=ROW_HDR, start_column=COL_LABEL,
        end_row=last_data_row, end_column=COL_LABEL,
    )
    c = ws.cell(ROW_HDR, COL_LABEL)
    c.value     = "Réactifs"
    c.fill      = _REACT_SIDE
    c.font      = _WHITE_BOLD
    c.alignment = _VROT
    c.border    = _MED_BORDER

    # ── En-têtes des colonnes fixes ───────────────────────────────────────────
    for col, label, fill in [
        (COL_REACT,    "Réactif",               _META_HDR),
        (COL_CONC,     "Concentration",          _META_HDR),
        (COL_UNIT,     "Unité",                  _META_HDR),
        (COL_VERIF,    "Vérification\nsolution", _VERIF_HDR),
        (COL_VOL_UNIT, "Unité\nvolume",          _META_HDR),
    ]:
        font = _WHITE_BOLD if fill is _VERIF_HDR else _BOLD
        _s(ws, ROW_HDR, col, label, fill=fill, font=font,
           align=_ROT, border=_THIN_BORDER)

    # ── En-têtes des colonnes tubes ───────────────────────────────────────────
    sub_defs = [
        ("Volume (µL)",  _TUBE_HDR,  _HDR_FONT),
        ("Coordinateur", _COORD_HDR, _HDR_FONT),
        ("Opérateur",    _OPER_HDR,  _HDR_FONT),
    ]
    for t_idx, tube_name in enumerate(tube_cols):
        base = COL_FIRST_TUBE + t_idx * 3
        for sub, (sub_label, fill, font) in enumerate(sub_defs):
            label = tube_name if sub == 0 else sub_label
            _s(ws, ROW_HDR, base + sub, label,
               fill=fill, font=font, align=_ROT, border=_THIN_BORDER)

    # ── Données ───────────────────────────────────────────────────────────────
    for r_idx, (_, row) in enumerate(df_comp.iterrows()):
        dr   = ROW_DATA + r_idx
        rfil = _ROW_A if r_idx % 2 == 0 else _ROW_B

        # Réactif
        _s(ws, dr, COL_REACT, str(row.get("Réactif", "")),
           fill=rfil, font=_DARK, align=_LEFT, border=_THIN_BORDER)

        # Concentration (max 1 décimale)
        raw_conc = row.get("Concentration", "")
        try:
            cv = round(float(raw_conc), 1)
            conc_val = int(cv) if cv == int(cv) else cv
        except (ValueError, TypeError):
            conc_val = str(raw_conc) if raw_conc not in (None, "") else ""
        _s(ws, dr, COL_CONC, conc_val,
           fill=rfil, font=_DARK, align=_CTR, border=_THIN_BORDER)

        # Unité
        _s(ws, dr, COL_UNIT, str(row.get("Unité", "")),
           fill=rfil, font=_DARK, align=_CTR, border=_THIN_BORDER)

        # Vérification solution
        _s(ws, dr, COL_VERIF, _box(("verif", r_idx)),
           fill=_VERIF_CELL, font=_CHECK_FONT, align=_CTR, border=_THIN_BORDER)

        # Mode compte-goutte : pas > 0 → valeurs stockées en n_gouttes
        try:
            pas = float(row.get("Pas (µL)", 0) or 0)
        except (ValueError, TypeError):
            pas = 0.0

        # Colonne Unité volume : "gouttes" ou "µL" une seule fois par ligne
        _s(ws, dr, COL_VOL_UNIT, "gouttes" if pas > 0 else "µL",
           fill=rfil, font=_DARK, align=_CTR, border=_THIN_BORDER)

        # Volumes + cases à cocher par tube
        for t_idx, tube_name in enumerate(tube_cols):
            base = COL_FIRST_TUBE + t_idx * 3
            raw_vol = row.get(tube_name, "")
            try:
                v = float(raw_vol)
                if pas > 0:
                    vol_val = int(round(v))
                else:
                    vol_val = int(v) if v == int(v) else v
            except (ValueError, TypeError):
                vol_val = str(raw_vol) if raw_vol not in (None, "") else ""

            _s(ws, dr, base,     vol_val,                            fill=rfil,      font=_DARK,       align=_CTR, border=_THIN_BORDER)
            _s(ws, dr, base + 1, _box(("coord", r_idx, t_idx)),     fill=rfil,      font=_CHECK_FONT, align=_CTR, border=_THIN_BORDER)
            _s(ws, dr, base + 2, _box(("oper",  r_idx, t_idx)),     fill=rfil,      font=_CHECK_FONT, align=_CTR, border=_THIN_BORDER)

    # ── Dimensions des colonnes et lignes ─────────────────────────────────────
    ws.row_dimensions[ROW_HDR].height = 90
    for r in range(ROW_DATA, last_data_row + 1):
        ws.row_dimensions[r].height = 20

    # Appliquer la mise en page du modèle (largeurs, impression, marges…)
    last_col = COL_FIRST_TUBE + n_tubes * 3 - 1
    used_template = _apply_template_settings(
        ws, _TEMPLATE_PATH, last_col=last_col, last_row=last_data_row
    )
    # Fallback si le modèle est introuvable : ajustement automatique au contenu
    if not used_template:
        _autofit_columns(ws, min_width=3, max_width=28, padding=1)

    # Figer les volets : colonne Réactif visible en scrollant à droite
    ws.freeze_panes = ws.cell(ROW_DATA, COL_FIRST_TUBE)  # figer jusqu'à la colonne des tubes

    # ── Feuilles de données (rechargement comme métadonnées) ──────────────────
    _write_df_to_sheet(wb, "Volumes", df_comp)
    if df_map is not None:
        _write_df_to_sheet(wb, "Correspondance", df_map)
    # EtatProtocole
    if states:
        import pandas as pd
        rows = []
        for key, checked in states.items():
            if key[0] == "verif":
                rows.append({"type": "verif", "r_idx": key[1], "t_idx": -1, "checked": int(checked)})
            else:
                rows.append({"type": key[0], "r_idx": key[1], "t_idx": key[2], "checked": int(checked)})
        _write_df_to_sheet(wb, "EtatProtocole", pd.DataFrame(rows))

    wb.save(path)
