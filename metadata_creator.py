

# metadata_creator.py

from __future__ import annotations

import pandas as pd
import numpy as np
import re
import json
import time
import urllib.parse
import urllib.request
from scipy.stats import norm


from PySide6.QtWidgets import (
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QLabel,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QMessageBox,
    QDialog,
    QDialogButtonBox,
    QHeaderView,
    QLineEdit,
    QDateEdit,
    QDateTimeEdit,
    QTimeEdit,
    QFileDialog,
    QComboBox,
    QSpinBox,
    QDoubleSpinBox,
    QFormLayout,
    QCheckBox,
    QGroupBox,
    QAbstractItemView,
    QInputDialog,
    QMenu,
    QTextBrowser,
    QTextEdit,
)

from PySide6.QtCore import QObject, Qt, QDate, QDateTime, QEvent, QTime, QUrl, Slot, Signal
from PySide6.QtGui import QColor, QBrush, QDesktopServices

import metadata_model as mm

try:
    from PySide6.QtWebChannel import QWebChannel
    from PySide6.QtWebEngineWidgets import QWebEngineView
except Exception:
    QWebChannel = None
    QWebEngineView = None


# Affichage numérique: précision et suppression de l'écriture scientifique.
NUM_DISPLAY_PRECISION = 12
DEFAULT_VTOT_UL = 3090.0
SUM_TARGET_TOL = 1e-3
FIELD_EMPTY_STYLE = "background-color: rgba(217, 83, 79, 77);"
FIELD_FILLED_STYLE = "background-color: rgba(92, 184, 92, 77);"
OPTIONAL_DATE_SENTINEL = QDate(1900, 1, 1)
OPTIONAL_TIME_SENTINEL = QTime(0, 0)
SHEET_MEASURES = "Mesures"
SHEET_METADATA_INDEX = "Index-métadonnées"
SHEET_TITRATION_VOLUMES = "Titration-volumes"
SHEET_TITRATION_CORRESPONDENCE = "Titration-correspondance"
SHEET_TITRATION_PROTOCOL_STATE = "Titration-EtatProtocole"
SHEET_TITRATION_VOLUMES_ALIASES = (SHEET_TITRATION_VOLUMES, "Volumes")
SHEET_TITRATION_CORRESPONDENCE_ALIASES = (SHEET_TITRATION_CORRESPONDENCE, "Correspondance")
SHEET_TITRATION_PROTOCOL_STATE_ALIASES = (SHEET_TITRATION_PROTOCOL_STATE, "EtatProtocole")
MOLAR_UNIT_FACTORS = {
    "M": 1.0,
    "mM": 1e-3,
    "µM": 1e-6,
    "nM": 1e-9,
}


# Affichage numérique: précision et suppression de l'écriture scientifique.
NUM_DISPLAY_PRECISION = 12


def _field_has_value(widget) -> bool:
    if isinstance(widget, QLineEdit):
        return bool(widget.text().strip())
    if isinstance(widget, QTextEdit):
        return bool(widget.toPlainText().strip())
    if isinstance(widget, QComboBox):
        return str(widget.currentText()).strip().lower() not in {"", "choisir..."}
    if isinstance(widget, (QSpinBox, QDoubleSpinBox)):
        if widget.specialValueText() and widget.value() == widget.minimum():
            return False
        return True
    if isinstance(widget, QDateEdit):
        if widget.specialValueText() and widget.date() == widget.minimumDate():
            return False
        return True
    if isinstance(widget, QTimeEdit):
        if widget.specialValueText() and widget.time() == widget.minimumTime():
            return False
        return True
    if isinstance(widget, QDateTimeEdit):
        if widget.specialValueText() and widget.dateTime() == widget.minimumDateTime():
            return False
        return True
    return True


def _apply_field_fill_style(widget) -> None:
    if widget is None:
        return
    widget.setStyleSheet(FIELD_FILLED_STYLE if _field_has_value(widget) else FIELD_EMPTY_STYLE)


def _install_field_fill_style(widget) -> None:
    if widget is None:
        return
    _apply_field_fill_style(widget)
    if widget.property("_fill_state_style_installed"):
        return
    widget.setProperty("_fill_state_style_installed", True)
    if isinstance(widget, QLineEdit):
        widget.textChanged.connect(lambda *_args, w=widget: _apply_field_fill_style(w))
    elif isinstance(widget, QTextEdit):
        widget.textChanged.connect(lambda w=widget: _apply_field_fill_style(w))
    elif isinstance(widget, QComboBox):
        widget.currentTextChanged.connect(lambda *_args, w=widget: _apply_field_fill_style(w))
    elif isinstance(widget, (QSpinBox, QDoubleSpinBox)):
        widget.valueChanged.connect(lambda *_args, w=widget: _apply_field_fill_style(w))
    elif isinstance(widget, QTimeEdit):
        widget.timeChanged.connect(lambda *_args, w=widget: _apply_field_fill_style(w))
    elif isinstance(widget, QDateEdit):
        widget.dateChanged.connect(lambda *_args, w=widget: _apply_field_fill_style(w))
    elif isinstance(widget, QDateTimeEdit):
        widget.dateTimeChanged.connect(lambda *_args, w=widget: _apply_field_fill_style(w))


def _first_existing_sheet(sheet_names: list[str], candidates: tuple[str, ...]) -> str | None:
    for name in candidates:
        if name in sheet_names:
            return name
    return None


class TableEditorDialog(QDialog):
    """Boîte de dialogue générique pour éditer un tableau type Excel.

    Le contenu est édité dans un QTableWidget mais peut être converti
    proprement en DataFrame pandas via `to_dataframe()`.
    """

    def __init__(
        self,
        title: str,
        columns: list[str],
        initial_df: pd.DataFrame | None = None,
        parent=None,
        reset_df: pd.DataFrame | None = None,
        reset_button_text: str = "Reset",
        sum_row: bool = False,
        sum_row_label: str = "Total",
        sum_target: float | None = None,
        show_rename_col: bool = True,
    ) -> None:
        super().__init__(parent)
        self.setWindowTitle(title)

        self._columns = list(columns)
        # self.conc_user_overridden = False
        self._reset_df = reset_df.copy() if reset_df is not None else None
        self._reset_button_text = reset_button_text
        self._sum_row_enabled = bool(sum_row)
        self._sum_row_label = str(sum_row_label)
        self._sum_target = float(sum_target) if sum_target is not None else None
        self._sum_row_idx: int | None = None
        self._sum_updating = False
        self._active_tube_col: int | None = None

        layout = QVBoxLayout(self)
        layout.addWidget(QLabel(title))

        # --- Table d'édition ---
        self.table = QTableWidget(self)
        self.table.setColumnCount(len(self._columns))
        self.table.setHorizontalHeaderLabels(self._columns)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.table.itemChanged.connect(self._on_item_changed)

        # Remplissage initial : soit à partir du DataFrame, soit quelques lignes vides
        if initial_df is not None and not initial_df.empty:
            self._load_from_dataframe(initial_df)
        else:
            self.table.setRowCount(8)  # quelques lignes vides par défaut
            if self._sum_row_enabled:
                self._append_sum_row()

        layout.addWidget(self.table, 1)

        # --- Boutons de modification du tableau ---
        btns_edit = QHBoxLayout()
        self.btn_add_row = QPushButton("+ Ligne", self)
        self.btn_del_row = QPushButton("− Ligne", self)
        self.btn_add_col = QPushButton("Ajouter un tube", self)
        self.btn_del_col = QPushButton("Supprimer le tube sélectionné", self)
        self.btn_dup_col = QPushButton("Dupliquer le tube sélectionné", self)
        self.btn_rename_col = QPushButton("Renommer colonne…", self)
        self.btn_del_col.setToolTip("Cliquez d'abord sur l'en-tête du tube (ligne de titres) pour le sélectionner, puis cliquez ce bouton.")
        self.btn_dup_col.setToolTip("Cliquez d'abord sur l'en-tête du tube (ligne de titres) pour le sélectionner, puis cliquez ce bouton.")
        btns_edit.addWidget(self.btn_add_row)
        btns_edit.addWidget(self.btn_del_row)
        btns_edit.addSpacing(20)
        btns_edit.addWidget(self.btn_add_col)
        btns_edit.addWidget(self.btn_del_col)
        btns_edit.addWidget(self.btn_dup_col)
        self.btn_rename_col.setVisible(bool(show_rename_col))
        if show_rename_col:
            btns_edit.addWidget(self.btn_rename_col)

        btns_edit.addSpacing(20)
        self.btn_reset = QPushButton(self._reset_button_text, self)
        self.btn_reset.setEnabled(self._reset_df is not None)
        btns_edit.addWidget(self.btn_reset)

        btns_edit.addStretch(1)
        layout.addLayout(btns_edit)

        # --- Option saisie manuelle (désactive la sync A↔B) ---
        mode_row = QHBoxLayout()
        self.chk_manual_mode = QCheckBox("Saisie manuelle (désactiver la synchronisation Solution A2 ↔ Solution B)", self)
        self.chk_manual_mode.setToolTip(
            "Coché : vous pouvez saisir librement les volumes de Solution A2 et Solution B "
            "sans qu'ils se recalculent mutuellement.\n"
            "Décoché (défaut) : modifier A2 ajuste B automatiquement pour conserver la somme constante."
        )
        mode_row.addWidget(self.chk_manual_mode)
        mode_row.addStretch(1)
        layout.addLayout(mode_row)

        self.btn_add_row.clicked.connect(self._add_row)
        self.btn_del_row.clicked.connect(self._del_row)
        self.btn_add_col.clicked.connect(self._add_col)
        self.btn_del_col.clicked.connect(self._del_col)
        self.btn_dup_col.clicked.connect(self._duplicate_col)
        self.btn_rename_col.clicked.connect(self._rename_col)
        self.btn_reset.clicked.connect(self._reset_table)

        # Clic simple sur un en-tête → sélectionner la colonne
        self.table.horizontalHeader().sectionClicked.connect(self._on_header_clicked)
        # Double-clic sur un en-tête → renommer
        self.table.horizontalHeader().sectionDoubleClicked.connect(self._rename_col)
        # Clic droit sur un en-tête → menu contextuel
        self.table.horizontalHeader().setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.horizontalHeader().customContextMenuRequested.connect(self._on_header_context_menu)

        # --- Boutons OK / Annuler ---
        btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, self)
        btn_box.accepted.connect(self._accept_if_not_empty)
        btn_box.rejected.connect(self.reject)
        layout.addWidget(btn_box)

        self.resize(800, 500)

    # ------------------------------------------------------------------
    # Utilitaires internes
    # ------------------------------------------------------------------
    def _load_from_dataframe(self, df: pd.DataFrame) -> None:
        """Charge un DataFrame existant dans le QTableWidget."""
        df = df.copy()
        # Réordonner / compléter selon les colonnes attendues
        for col in self._columns:
            if col not in df.columns:
                df[col] = pd.NA
        df = df[self._columns]

        self.table.setRowCount(len(df) + (1 if self._sum_row_enabled else 0))
        self.table.setColumnCount(len(self._columns))
        self.table.setHorizontalHeaderLabels(self._columns)

        for i in range(len(df)):
            for j, col in enumerate(self._columns):
                val = df.iloc[i, j]
                if pd.isna(val):
                    txt = ""
                else:
                    # Affichage lisible : pas d'écriture scientifique, précision configurable.
                    if isinstance(val, (float, int)):
                        txt = self._format_number(float(val))
                    else:
                        txt = str(val)
                item = QTableWidgetItem(txt)
                self.table.setItem(i, j, item)
        self.table.resizeColumnsToContents()
        if self._sum_row_enabled:
            self._sum_row_idx = len(df)
            self._ensure_sum_row_items()
            self._set_sum_row_label()
            self._refresh_sum_row()

    def _format_number(self, val: float) -> str:
        try:
            return np.format_float_positional(
                float(val),
                precision=NUM_DISPLAY_PRECISION,
                unique=False,
                trim="-",
            )
        except Exception:
            return str(val)

    def _parse_float(self, txt: str | None) -> float | None:
        return mm.to_float(txt)

    def _get_tube_column_indices(self) -> list[int]:
        idxs: list[int] = []
        for j in range(self.table.columnCount()):
            header = self.table.horizontalHeaderItem(j)
            name = header.text().strip() if header is not None else ""
            if name.lower().startswith("tube "):
                idxs.append(j)
        return idxs

    def _ensure_sum_row_items(self) -> None:
        if self._sum_row_idx is None:
            return
        for j in range(self.table.columnCount()):
            item = self.table.item(self._sum_row_idx, j)
            if item is None:
                item = QTableWidgetItem("")
                self.table.setItem(self._sum_row_idx, j, item)
            item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)

    def _set_sum_row_label(self) -> None:
        if self._sum_row_idx is None or self.table.columnCount() == 0:
            return
        label_col = 0
        for j in range(self.table.columnCount()):
            header = self.table.horizontalHeaderItem(j)
            name = header.text().strip().lower() if header is not None else ""
            if name in ("réactif", "reactif"):
                label_col = j
                break
        item = self.table.item(self._sum_row_idx, label_col)
        if item is None:
            item = QTableWidgetItem("")
            self.table.setItem(self._sum_row_idx, label_col, item)
        item.setText(self._sum_row_label)
        item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)

    def _append_sum_row(self) -> None:
        self._sum_row_idx = self.table.rowCount()
        self.table.insertRow(self._sum_row_idx)
        self._ensure_sum_row_items()
        self._set_sum_row_label()
        self._refresh_sum_row()

    def _refresh_sum_row(self) -> None:
        if not self._sum_row_enabled or self._sum_row_idx is None:
            return
        if self._sum_updating:
            return
        self._sum_updating = True
        try:
            self.table.blockSignals(True)
            self._ensure_sum_row_items()
            self._set_sum_row_label()
            tube_cols = self._get_tube_column_indices()
            # Indice de la colonne "Pas (µL)" pour les lignes compte-goutte
            pas_col_idx: int | None = None
            for j in range(self.table.columnCount()):
                hdr = self.table.horizontalHeaderItem(j)
                if hdr is not None and hdr.text().strip() == "Pas (µL)":
                    pas_col_idx = j
                    break
            # Nettoie les fonds
            for j in range(self.table.columnCount()):
                item = self.table.item(self._sum_row_idx, j)
                if item is not None:
                    item.setBackground(QBrush())
            for j in tube_cols:
                total = 0.0
                for r in range(self._sum_row_idx):
                    item = self.table.item(r, j)
                    val = self._parse_float(item.text() if item is not None else "")
                    if val is None:
                        continue
                    # Si ligne compte-goutte, multiplier les n_gouttes par le pas
                    if pas_col_idx is not None:
                        pas_item = self.table.item(r, pas_col_idx)
                        pas = self._parse_float(pas_item.text() if pas_item is not None else "") or 0.0
                        if pas > 0:
                            val = val * pas
                    total += val
                item = self.table.item(self._sum_row_idx, j)
                if item is None:
                    item = QTableWidgetItem("")
                    self.table.setItem(self._sum_row_idx, j, item)
                item.setText(self._format_number(total))
                item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                if self._sum_target is not None and abs(total - self._sum_target) <= SUM_TARGET_TOL:
                    item.setBackground(QBrush(QColor(198, 239, 206)))
        finally:
            self.table.blockSignals(False)
            self._sum_updating = False

    def _auto_resize_columns(self) -> None:
        """Ajuste automatiquement la largeur des colonnes au contenu."""
        self.table.resizeColumnsToContents()

    def _on_header_clicked(self, col: int) -> None:
        """Sélectionne visuellement la colonne cliquée et la mémorise."""
        self._active_tube_col = col
        self.table.selectColumn(col)

    def _on_header_context_menu(self, pos) -> None:
        """Menu contextuel sur un en-tête de colonne Tube : dupliquer / supprimer."""
        col = self.table.horizontalHeader().logicalIndexAt(pos)
        if col < 0:
            return
        header = self.table.horizontalHeaderItem(col)
        col_name = header.text().strip() if header is not None else ""
        if not col_name.lower().startswith("tube "):
            return
        self._active_tube_col = col
        self.table.selectColumn(col)
        menu = QMenu(self)
        act_dup = menu.addAction(f"Dupliquer « {col_name} »")
        act_del = menu.addAction(f"Supprimer « {col_name} »")
        action = menu.exec(self.table.horizontalHeader().mapToGlobal(pos))
        if action == act_dup:
            self._duplicate_col()
        elif action == act_del:
            self._del_col()

    def _on_item_changed(self, item: QTableWidgetItem) -> None:
        self._auto_resize_columns()
        if not self._sum_row_enabled or self._sum_row_idx is None:
            return
        if self._sum_updating:
            return
        if item.row() == self._sum_row_idx:
            return
        self._refresh_sum_row()

    def _add_row(self) -> None:
        if self._sum_row_enabled and self._sum_row_idx is not None:
            self.table.insertRow(self._sum_row_idx)
            self._sum_row_idx += 1
        else:
            self.table.insertRow(self.table.rowCount())
        if self._sum_row_enabled:
            self._refresh_sum_row()

    def _del_row(self) -> None:
        selected_rows = sorted({idx.row() for idx in self.table.selectedIndexes()}, reverse=True)
        if not selected_rows:
            row = self.table.currentRow()
            if row < 0:
                return
            selected_rows = [row]

        for row in selected_rows:
            if self._sum_row_enabled and self._sum_row_idx is not None:
                if row == self._sum_row_idx:
                    continue
                if row < self._sum_row_idx:
                    self._sum_row_idx -= 1
            self.table.removeRow(row)
        if self._sum_row_enabled:
            self._refresh_sum_row()

    def _add_col(self) -> None:
        # Trouver les colonnes "Tube N" existantes et le dernier numéro
        tube_indices = self._get_tube_column_indices()
        if tube_indices:
            last_tube_col = tube_indices[-1]
            last_header = self.table.horizontalHeaderItem(last_tube_col)
            last_name = last_header.text().strip() if last_header else ""
            # Extraire le numéro du dernier tube (ex: "Tube 5" → 5)
            import re as _re
            m = _re.search(r"\d+$", last_name)
            next_num = int(m.group()) + 1 if m else len(tube_indices) + 1
            new_name = f"Tube {next_num}"
        else:
            last_tube_col = None
            new_name = "Tube 1"

        # Insérer la nouvelle colonne
        col_count = self.table.columnCount()
        self.table.insertColumn(col_count)
        self._columns.append(new_name)
        self.table.setHorizontalHeaderLabels(self._columns)

        # Pré-remplir depuis le dernier tube (hors ligne de somme)
        # Les signaux sont bloqués pour éviter que sync_A_B se déclenche
        # au milieu de la copie avec un target_sum incomplet/erroné.
        if last_tube_col is not None:
            n_data_rows = self._sum_row_idx if (self._sum_row_enabled and self._sum_row_idx is not None) else self.table.rowCount()
            self.table.blockSignals(True)
            try:
                for r in range(n_data_rows):
                    src = self.table.item(r, last_tube_col)
                    txt = src.text() if src is not None else ""
                    self.table.setItem(r, col_count, QTableWidgetItem(txt))
            finally:
                self.table.blockSignals(False)

        if self._sum_row_enabled:
            self._ensure_sum_row_items()
            self._refresh_sum_row()

    def _resolve_tube_col(self, action_label: str) -> int | None:
        """Retourne l'indice de la colonne Tube active, ou None avec un message d'erreur."""
        col = self._active_tube_col
        if col is None or col < 0 or col >= self.table.columnCount():
            QMessageBox.information(
                self, action_label,
                "Cliquez d'abord sur l'en-tête d'un tube (la ligne de titres) pour le sélectionner."
            )
            return None
        header = self.table.horizontalHeaderItem(col)
        col_name = header.text().strip() if header is not None else ""
        if not col_name.lower().startswith("tube "):
            QMessageBox.information(
                self, action_label,
                f"La colonne « {col_name} » n'est pas une colonne Tube.\n"
                "Cliquez sur l'en-tête d'un tube pour le sélectionner."
            )
            return None
        return col

    def _del_col(self) -> None:
        col = self._resolve_tube_col("Supprimer un tube")
        if col is None:
            return
        if 0 <= col < len(self._columns):
            self._columns.pop(col)
        self.table.removeColumn(col)
        self.table.setHorizontalHeaderLabels(self._columns)
        self._active_tube_col = None
        if self._sum_row_enabled:
            self._refresh_sum_row()

    def _duplicate_col(self) -> None:
        """Duplique le tube de la colonne actuellement sélectionnée."""
        col = self._resolve_tube_col("Dupliquer un tube")
        if col is None:
            return

        # Numéro du nouveau tube = max existant + 1
        tube_indices = self._get_tube_column_indices()
        last_header = self.table.horizontalHeaderItem(tube_indices[-1])
        last_name = last_header.text().strip() if last_header else ""
        import re as _re
        m = _re.search(r"\d+$", last_name)
        next_num = int(m.group()) + 1 if m else len(tube_indices) + 1
        new_name = f"Tube {next_num}"

        # Insérer juste après la colonne sélectionnée
        new_col = col + 1
        self.table.insertColumn(new_col)
        self._columns.insert(new_col, new_name)
        self.table.setHorizontalHeaderLabels(self._columns)

        # Copier les valeurs (hors ligne de somme)
        n_data_rows = (
            self._sum_row_idx
            if (self._sum_row_enabled and self._sum_row_idx is not None)
            else self.table.rowCount()
        )
        self.table.blockSignals(True)
        try:
            for r in range(n_data_rows):
                src = self.table.item(r, col)
                txt = src.text() if src is not None else ""
                self.table.setItem(r, new_col, QTableWidgetItem(txt))
        finally:
            self.table.blockSignals(False)

        if self._sum_row_enabled:
            self._ensure_sum_row_items()
            self._refresh_sum_row()

    def _rename_col(self, col: int = -1) -> None:
        """Renomme la colonne `col` (ou la colonne actuellement sélectionnée si col=-1)."""
        if col < 0:
            col = self.table.currentColumn()
        if col < 0 or col >= self.table.columnCount():
            QMessageBox.information(self, "Renommer", "Cliquez d'abord sur une colonne pour la sélectionner.")
            return
        header = self.table.horizontalHeaderItem(col)
        current_name = header.text() if header is not None else ""
        new_name, ok = QInputDialog.getText(
            self, "Renommer la colonne", f"Nouveau nom pour « {current_name} » :", text=current_name
        )
        if not ok or not new_name.strip():
            return
        new_name = new_name.strip()
        if 0 <= col < len(self._columns):
            self._columns[col] = new_name
        self.table.setHorizontalHeaderLabels(self._columns)

    def _reset_table(self) -> None:
        """Restaure le tableau par défaut (si fourni)."""
        if self._reset_df is None:
            return
        self._load_from_dataframe(self._reset_df)

    def _accept_if_not_empty(self) -> None:
        df = self.to_dataframe()
        if df.empty:
            QMessageBox.warning(self, "Tableau vide", "Le tableau est vide. Ajoutez au moins une ligne avant de valider.")
            return
        self.accept()

    # ------------------------------------------------------------------
    # API publique
    # ------------------------------------------------------------------
    def to_dataframe(self) -> pd.DataFrame:
        """Convertit le contenu de la table en DataFrame pandas.

        Les lignes complètement vides sont supprimées.
        """
        rows = self.table.rowCount()
        cols = self.table.columnCount()
        data = {self.table.horizontalHeaderItem(j).text() or f"col{j}": [] for j in range(cols)}

        for i in range(rows):
            row_vals = []
            for j in range(cols):
                item = self.table.item(i, j)
                txt = item.text().strip() if item is not None else ""
                row_vals.append(txt if txt != "" else pd.NA)
            # Ajoute la ligne (on filtrera les lignes vides ensuite)
            for j, key in enumerate(data.keys()):
                data[key].append(row_vals[j])

        df = pd.DataFrame(data)
        # Suppression des lignes entièrement vides
        df = df.dropna(how="all")
        if self._sum_row_enabled and self._sum_row_idx is not None and self._sum_row_idx in df.index:
            df = df.drop(index=self._sum_row_idx)
        # Normalise quelques colonnes si présentes
        if "Tube" in df.columns:
            df["Tube"] = df["Tube"].astype(str).str.strip()
            df = df[df["Tube"] != ""]
        if "Spectrum name" in df.columns:
            df["Spectrum name"] = df["Spectrum name"].astype(str).str.strip()
        return df.reset_index(drop=True)


class MolesDialog(QDialog):
    """Dialogue d'édition des quantités de matière (nmol)."""

    def __init__(self, entries: list[dict], parent=None) -> None:
        super().__init__(parent)
        self.setWindowTitle("Quantités de matière")
        self._entries = entries
        self._spins: dict[str, QDoubleSpinBox] = {}

        layout = QVBoxLayout(self)
        layout.addWidget(
            QLabel(
                "Éditez les quantités de matière (nmol) calculées à partir des "
                "concentrations et volumes actuels."
            )
        )

        form = QFormLayout()
        for entry in self._entries:
            key = str(entry.get("key", ""))
            label = str(entry.get("label", ""))
            value = entry.get("value", None)
            editable = bool(entry.get("editable", True))
            note = str(entry.get("note", "")).strip()

            row_label = f"{label} (nmol)"
            if editable and value is not None:
                spin = QDoubleSpinBox(self)
                spin.setDecimals(6)
                spin.setRange(0.0, 1e12)
                spin.setValue(float(value))
                if note:
                    spin.setToolTip(note)
                form.addRow(row_label, spin)
                self._spins[key] = spin
            else:
                txt = "N/A"
                if value is not None:
                    txt = str(value)
                le = QLineEdit(txt, self)
                le.setReadOnly(True)
                if note:
                    le.setToolTip(note)
                form.addRow(row_label, le)

        layout.addLayout(form)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, self)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

        self.resize(520, 320)

    def values(self) -> dict[str, float]:
        return {k: float(spin.value()) for k, spin in self._spins.items()}


class AmmoniumTestDialog(QDialog):
    """Dialogue de saisie des résultats des tests ammonium et du pH."""

    def __init__(self, values: dict | None = None, parent=None) -> None:
        super().__init__(parent)
        self.setWindowTitle("Tests ammonium")
        values = values or {}

        layout = QVBoxLayout(self)
        form = QFormLayout()

        self.edit_operator = QLineEdit(str(values.get("operator", "") or ""), self)
        self.edit_test_1 = QLineEdit(str(values.get("test_1", "") or ""), self)
        self.edit_test_2 = QLineEdit(str(values.get("test_2", "") or ""), self)
        self.edit_ph = QLineEdit(str(values.get("ph", "") or ""), self)

        self.edit_operator.setPlaceholderText("Prénom puis nom")
        self.edit_test_1.setPlaceholderText("Test 1 grossier")
        self.edit_test_2.setPlaceholderText("Test 2 précis")
        self.edit_ph.setPlaceholderText("pH")
        for edit in (self.edit_operator, self.edit_test_1, self.edit_test_2, self.edit_ph):
            _install_field_fill_style(edit)

        form.addRow("Test réalisé par :", self.edit_operator)
        form.addRow("Valeur du test numéro 1 (grossier) :", self.edit_test_1)
        form.addRow("Valeur du test numéro 2 (précis) :", self.edit_test_2)
        form.addRow("Valeur du pH :", self.edit_ph)
        layout.addLayout(form)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, self)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def values(self) -> dict[str, str]:
        return {
            "operator": self.edit_operator.text().strip(),
            "test_1": self.edit_test_1.text().strip(),
            "test_2": self.edit_test_2.text().strip(),
            "ph": self.edit_ph.text().strip(),
        }


class DebitFluxDialog(QDialog):
    """Dialogue de saisie du tableau débit / flux pour les prélèvements en eau douce."""

    def __init__(self, columns: list[str], values: dict[str, str] | None = None, parent=None) -> None:
        super().__init__(parent)
        self.setWindowTitle("Débit / flux")
        self.resize(520, 360)
        self._columns = list(columns)
        values = values or {}

        layout = QVBoxLayout(self)
        title = QLabel("Débit / flux", self)
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title.setStyleSheet("font-weight: 800; font-size: 16px;")
        layout.addWidget(title)

        form = QFormLayout()
        self._edits: dict[str, QLineEdit] = {}
        for column in self._columns:
            edit = QLineEdit(str(values.get(column, "") or ""), self)
            edit.setPlaceholderText("valeur à renseigner")
            edit.setMinimumWidth(260)
            _install_field_fill_style(edit)
            form.addRow(f"{column} :", edit)
            self._edits[column] = edit
        layout.addLayout(form)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, self)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def values(self) -> dict[str, str]:
        return {column: edit.text().strip() for column, edit in self._edits.items()}


class _MapPickerBridge(QObject):
    def __init__(self, dialog: "MapPickerDialog") -> None:
        super().__init__(dialog)
        self._dialog = dialog

    @Slot(float, float)
    def setCoordinates(self, lat: float, lon: float) -> None:
        self._dialog.set_coordinates(float(lat), float(lon))


class MapPickerDialog(QDialog):
    """Fenêtre interne permettant de choisir un point sur une carte OpenStreetMap."""

    def __init__(self, lat: float | None = None, lon: float | None = None, parent=None) -> None:
        super().__init__(parent)
        self.setWindowTitle("Pointer le prélèvement sur la carte")
        self.resize(850, 620)
        self._lat = lat
        self._lon = lon

        layout = QVBoxLayout(self)
        self.lbl_coords = QLabel(self._coords_text(), self)
        layout.addWidget(self.lbl_coords)

        self.view = QWebEngineView(self)
        self._channel = QWebChannel(self)
        self._bridge = _MapPickerBridge(self)
        self._channel.registerObject("coordBridge", self._bridge)
        self.view.page().setWebChannel(self._channel)
        self.view.setHtml(self._map_html(), QUrl("https://www.openstreetmap.org/"))
        layout.addWidget(self.view, 1)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, self)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def _coords_text(self) -> str:
        if self._lat is None or self._lon is None:
            return "Cliquez sur la carte pour choisir le point de prélèvement."
        return f"Point sélectionné : latitude {self._lat:.6f}, longitude {self._lon:.6f}"

    def set_coordinates(self, lat: float, lon: float) -> None:
        self._lat = lat
        self._lon = lon
        self.lbl_coords.setText(self._coords_text())

    def has_coordinates(self) -> bool:
        return self._lat is not None and self._lon is not None

    def coordinates(self) -> tuple[float, float]:
        return float(self._lat), float(self._lon)

    def _map_html(self) -> str:
        center_lat = self._lat if self._lat is not None else 46.6
        center_lon = self._lon if self._lon is not None else 2.4
        has_point = "true" if self.has_coordinates() else "false"
        html = """
<!doctype html>
<html>
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css">
  <script src="qrc:///qtwebchannel/qwebchannel.js"></script>
  <script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
  <style>
    html, body, #map { height: 100%; width: 100%; margin: 0; padding: 0; }
    body { background: #202020; color: #f5f5f5; font-family: -apple-system, BlinkMacSystemFont, sans-serif; }
    #hint {
      position: absolute; z-index: 1000; top: 10px; left: 50px;
      background: rgba(20, 20, 20, 0.82); color: white;
      padding: 6px 10px; border-radius: 4px; font-weight: 600;
    }
  </style>
</head>
<body>
  <div id="map"></div>
  <div id="hint">Cliquez sur la carte pour placer le prélèvement.</div>
  <script>
    const startLat = __CENTER_LAT__;
    const startLon = __CENTER_LON__;
    const hasPoint = __HAS_POINT__;
    let coordBridge = null;
    let marker = null;

    new QWebChannel(qt.webChannelTransport, function(channel) {
      coordBridge = channel.objects.coordBridge;
    });

    const map = L.map('map').setView([startLat, startLon], hasPoint ? 16 : 6);
    L.tileLayer('https://tile.openstreetmap.org/{z}/{x}/{y}.png', {
      maxZoom: 19,
      attribution: '&copy; OpenStreetMap'
    }).addTo(map);

    function setPoint(lat, lon) {
      if (marker === null) {
        marker = L.marker([lat, lon], { draggable: true }).addTo(map);
        marker.on('dragend', function(event) {
          const p = event.target.getLatLng();
          setPoint(p.lat, p.lng);
        });
      } else {
        marker.setLatLng([lat, lon]);
      }
      document.getElementById('hint').textContent =
        'Latitude ' + lat.toFixed(6) + ' · Longitude ' + lon.toFixed(6);
      if (coordBridge) {
        coordBridge.setCoordinates(lat, lon);
      }
    }

    if (hasPoint) {
      setPoint(startLat, startLon);
    }

    map.on('click', function(event) {
      setPoint(event.latlng.lat, event.latlng.lng);
    });
  </script>
</body>
</html>
"""
        return (
            html.replace("__CENTER_LAT__", f"{center_lat:.6f}")
            .replace("__CENTER_LON__", f"{center_lon:.6f}")
            .replace("__HAS_POINT__", has_point)
        )


class GaussianVolumesDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Générer un tableau de volume")
        self.resize(520, 420)

        layout = QVBoxLayout(self)
        layout.addWidget(QLabel(
            "Génère automatiquement les volumes de titrant (Solution B) autour de l'équivalence.\n"
            "Solution A2 (tampon NH3+, 7 mM) et Solution B (titrant) varient de tube en tube."
        ))

        # Case à cocher mode compte‑goutte
        self.chk_dropper = QCheckBox("Utilisation d'un compte‑goutte (pas = 40 µL, marge = 0)", self)
        self.chk_dropper.setChecked(False)
        layout.addWidget(self.chk_dropper)

        box = QGroupBox("Paramètres")
        form = QFormLayout(box)

        self.spin_n = QSpinBox(self); self.spin_n.setRange(2, 200); self.spin_n.setValue(11)
        self.spin_C0 = QDoubleSpinBox(self); self.spin_C0.setDecimals(3); self.spin_C0.setRange(0.0, 1e9); self.spin_C0.setValue(1000.0)
        self.spin_Vech = QDoubleSpinBox(self); self.spin_Vech.setDecimals(1); self.spin_Vech.setRange(0.0, 1e9); self.spin_Vech.setValue(1000.0)
        self.spin_Ctit = QDoubleSpinBox(self); self.spin_Ctit.setDecimals(3); self.spin_Ctit.setRange(0.0, 1e9); self.spin_Ctit.setValue(5.0)
        self.spin_Vtot = QDoubleSpinBox(self); self.spin_Vtot.setDecimals(1); self.spin_Vtot.setRange(0.0, 1e9); self.spin_Vtot.setValue(DEFAULT_VTOT_UL)

        form.addRow("Nombre de tubes", self.spin_n)
        form.addRow("C0 Cu (nM)", self.spin_C0)
        form.addRow("V échantillon (µL)", self.spin_Vech)
        form.addRow("C titrant (µM)", self.spin_Ctit)
        form.addRow("V total par tube (µL)", self.spin_Vtot)

        layout.addWidget(box)

        box2 = QGroupBox("Solutions fixes")
        form2 = QFormLayout(box2)

        unit_choices = ["M", "mM", "µM", "nM", "% en masse"]

        self.spin_VA1 = QDoubleSpinBox(self); self.spin_VA1.setDecimals(1); self.spin_VA1.setRange(0.0, 1e9); self.spin_VA1.setValue(100.0)
        self.spin_VC = QDoubleSpinBox(self); self.spin_VC.setDecimals(1); self.spin_VC.setRange(0.0, 1e9); self.spin_VC.setValue(30.0)
        self.spin_VD = QDoubleSpinBox(self); self.spin_VD.setDecimals(1); self.spin_VD.setRange(0.0, 1e9); self.spin_VD.setValue(750.0)
        self.spin_VE = QDoubleSpinBox(self); self.spin_VE.setDecimals(1); self.spin_VE.setRange(0.0, 1e9); self.spin_VE.setValue(750.0)
        self.spin_VF = QDoubleSpinBox(self); self.spin_VF.setDecimals(1); self.spin_VF.setRange(0.0, 1e9); self.spin_VF.setValue(60.0)

        self.spin_conc_A1 = QDoubleSpinBox(self); self.spin_conc_A1.setDecimals(6); self.spin_conc_A1.setRange(0.0, 1e12); self.spin_conc_A1.setValue(70.0)
        self.spin_conc_C = QDoubleSpinBox(self); self.spin_conc_C.setDecimals(6); self.spin_conc_C.setRange(0.0, 1e12); self.spin_conc_C.setValue(10.0)
        self.spin_conc_D = QDoubleSpinBox(self); self.spin_conc_D.setDecimals(6); self.spin_conc_D.setRange(0.0, 1e12); self.spin_conc_D.setValue(0.5)
        self.spin_conc_E = QDoubleSpinBox(self); self.spin_conc_E.setDecimals(6); self.spin_conc_E.setRange(0.0, 1e12); self.spin_conc_E.setValue(1.0)
        self.spin_conc_F = QDoubleSpinBox(self); self.spin_conc_F.setDecimals(6); self.spin_conc_F.setRange(0.0, 1e12); self.spin_conc_F.setValue(1.0)

        self.combo_unit_A1 = QComboBox(self); self.combo_unit_A1.addItems(unit_choices); self.combo_unit_A1.setCurrentText("mM")
        self.combo_unit_C = QComboBox(self); self.combo_unit_C.addItems(unit_choices); self.combo_unit_C.setCurrentText("µM")
        self.combo_unit_D = QComboBox(self); self.combo_unit_D.addItems(unit_choices); self.combo_unit_D.setCurrentText("% en masse")
        self.combo_unit_E = QComboBox(self); self.combo_unit_E.addItems(unit_choices); self.combo_unit_E.setCurrentText("mM")
        self.combo_unit_F = QComboBox(self); self.combo_unit_F.addItems(unit_choices); self.combo_unit_F.setCurrentText("mM")

        def _row(vol_spin: QDoubleSpinBox, conc_spin: QDoubleSpinBox, unit_combo: QComboBox,
                 vol_label: QLabel | None = None) -> QWidget:
            w = QWidget(self)
            h = QHBoxLayout(w)
            h.setContentsMargins(0, 0, 0, 0)
            lbl = vol_label if vol_label is not None else QLabel("V (µL)", w)
            h.addWidget(lbl)
            h.addWidget(vol_spin)
            h.addSpacing(12)
            h.addWidget(QLabel("C", w))
            h.addWidget(conc_spin)
            h.addWidget(unit_combo)
            h.addStretch(1)
            return w

        self._lbl_VA1 = QLabel("V (µL)", self)
        self._lbl_VC = QLabel("V (µL)", self)
        self._lbl_VF = QLabel("V (µL)", self)

        form2.addRow("Solution A1 (NH3+, 70 mM)", _row(self.spin_VA1, self.spin_conc_A1, self.combo_unit_A1, self._lbl_VA1))
        form2.addRow("Solution C", _row(self.spin_VC, self.spin_conc_C, self.combo_unit_C, self._lbl_VC))
        form2.addRow("Solution D", _row(self.spin_VD, self.spin_conc_D, self.combo_unit_D))
        form2.addRow("Solution E", _row(self.spin_VE, self.spin_conc_E, self.combo_unit_E))
        form2.addRow("Solution F", _row(self.spin_VF, self.spin_conc_F, self.combo_unit_F, self._lbl_VF))

        # Quantités de matière de référence (en nmol) pour le couplage C ↔ V
        # Solution F : 6e-8 mol = 60 nmol  |  Solution C : 3e-10 mol = 0.3 nmol
        # Solution A1 : 70 mM × 100 µL = 7000 nmol
        self._n_nmol_A1 = 7000.0
        self._n_nmol_F = 60.0
        self._n_nmol_C = 0.3
        self._updating_fixed = False  # garde contre les mises à jour récursives

        # ── Couplage C ↔ V pour Solution C et Solution F ─────────────────────
        # Règle : n_nmol est l'ancre.  Si C change → V = n/C.  Si V change → C = n/V.
        # Si l'unité change → C est recalculée (volume inchangé).

        def _vol_from_n_conc(n_nmol: float, conc: float, unit: str) -> float | None:
            """Retourne le volume (µL) pour n_nmol nmol à la concentration donnée."""
            factor = MOLAR_UNIT_FACTORS.get(unit)
            if factor is None or conc <= 0:
                return None
            c_m = conc * factor
            return n_nmol / (c_m * 1e3)  # n[nmol] / (c[M] * 1e3) = µL

        def _conc_from_n_vol(n_nmol: float, vol_uL: float, unit: str) -> float | None:
            """Retourne la concentration pour n_nmol nmol dans vol_uL µL."""
            factor = MOLAR_UNIT_FACTORS.get(unit)
            if factor is None or vol_uL <= 0:
                return None
            c_m = n_nmol / (vol_uL * 1e3)  # nmol / (µL * 1e3) = M
            return c_m / factor

        def _make_sync(vol_spin: QDoubleSpinBox, conc_spin: QDoubleSpinBox,
                       unit_combo: QComboBox, get_n, get_step=None):
            """Retourne les trois slots de synchronisation pour une solution.

            get_step : callable retournant le pas en µL/goutte (ou None si pas en mode goutte).
            Quand get_step() > 0, vol_spin contient un nombre de gouttes (entier).
            """

            def _to_uL(val):
                """Convertit la valeur affichée en µL."""
                step = get_step() if get_step else 0.0
                return val * step if step > 0 else val

            def on_conc_changed(val):
                if self._updating_fixed:
                    return
                v_uL = _vol_from_n_conc(get_n(), val, str(unit_combo.currentText()).strip())
                if v_uL is None or v_uL < 0:
                    return
                step = get_step() if get_step else 0.0
                displayed = round(v_uL / step) if step > 0 else round(v_uL, 1)
                self._updating_fixed = True
                try:
                    vol_spin.setValue(displayed)
                finally:
                    self._updating_fixed = False

            def on_vol_changed(val):
                if self._updating_fixed:
                    return
                vol_uL = _to_uL(val)
                c = _conc_from_n_vol(get_n(), vol_uL, str(unit_combo.currentText()).strip())
                if c is None or c < 0:
                    return
                self._updating_fixed = True
                try:
                    conc_spin.setValue(c)
                finally:
                    self._updating_fixed = False

            def on_unit_changed(_text):
                if self._updating_fixed:
                    return
                vol_uL = _to_uL(float(vol_spin.value()))
                c = _conc_from_n_vol(get_n(), vol_uL, str(unit_combo.currentText()).strip())
                if c is None or c < 0:
                    return
                self._updating_fixed = True
                try:
                    conc_spin.setValue(c)
                finally:
                    self._updating_fixed = False

            return on_conc_changed, on_vol_changed, on_unit_changed

        def _get_dropper_step():
            """Retourne le pas en µL/goutte si mode compte-goutte actif, sinon 0."""
            return float(self.spin_step.value()) if self.chk_dropper.isChecked() else 0.0

        _sa1_conc, _sa1_vol, _sa1_unit = _make_sync(
            self.spin_VA1, self.spin_conc_A1, self.combo_unit_A1,
            lambda: self._n_nmol_A1, _get_dropper_step)
        self.spin_conc_A1.valueChanged.connect(_sa1_conc)
        self.spin_VA1.valueChanged.connect(_sa1_vol)
        self.combo_unit_A1.currentTextChanged.connect(_sa1_unit)

        _sc_conc, _sc_vol, _sc_unit = _make_sync(
            self.spin_VC, self.spin_conc_C, self.combo_unit_C,
            lambda: self._n_nmol_C, _get_dropper_step)
        self.spin_conc_C.valueChanged.connect(_sc_conc)
        self.spin_VC.valueChanged.connect(_sc_vol)
        self.combo_unit_C.currentTextChanged.connect(_sc_unit)

        _sf_conc, _sf_vol, _sf_unit = _make_sync(
            self.spin_VF, self.spin_conc_F, self.combo_unit_F,
            lambda: self._n_nmol_F, _get_dropper_step)
        self.spin_conc_F.valueChanged.connect(_sf_conc)
        self.spin_VF.valueChanged.connect(_sf_vol)
        self.combo_unit_F.currentTextChanged.connect(_sf_unit)

        # Initialiser les volumes à partir des quantités et concentrations par défaut
        self._updating_fixed = True
        try:
            v_a1_init = _vol_from_n_conc(self._n_nmol_A1,
                                         float(self.spin_conc_A1.value()),
                                         str(self.combo_unit_A1.currentText()).strip())
            if v_a1_init is not None:
                self.spin_VA1.setValue(round(v_a1_init, 1))
            v_c_init = _vol_from_n_conc(self._n_nmol_C,
                                        float(self.spin_conc_C.value()),
                                        str(self.combo_unit_C.currentText()).strip())
            if v_c_init is not None:
                self.spin_VC.setValue(round(v_c_init, 1))
            v_f_init = _vol_from_n_conc(self._n_nmol_F,
                                        float(self.spin_conc_F.value()),
                                        str(self.combo_unit_F.currentText()).strip())
            if v_f_init is not None:
                self.spin_VF.setValue(round(v_f_init, 1))
        finally:
            self._updating_fixed = False

        layout.addWidget(box2)

        box3 = QGroupBox("Contraintes")
        form3 = QFormLayout(box3)

        self.spin_margin = QDoubleSpinBox(self); self.spin_margin.setDecimals(1); self.spin_margin.setRange(0.0, 1e9); self.spin_margin.setValue(20.0)
        self.spin_step = QDoubleSpinBox(self); self.spin_step.setDecimals(3); self.spin_step.setRange(0.001, 1e6); self.spin_step.setValue(1.0)

        # Cocher compte-goutte : applique un preset de valeurs, tout reste modifiable ensuite
        def _update_dropper_mode(state):
            self._apply_dropper_vol_mode(state)
            if state:
                self.spin_margin.setValue(0.0)
                self.spin_step.setValue(40.0)
                # A1 : 3 gouttes (= 120 µL) ; C et F : 1 goutte — le couplage recalcule la concentration
                self.spin_VA1.setValue(3.0)
                self.spin_VC.setValue(1.0)
                self.spin_VF.setValue(1.0)

        self.chk_dropper.toggled.connect(_update_dropper_mode)

        self.chk_zero = QCheckBox("Inclure un tube à 0 µL de titrant", self); self.chk_zero.setChecked(True)
        self.chk_dup = QCheckBox("Ajouter un tube contrôle (même volumes que l'équivalence, solution Contrôle)", self); self.chk_dup.setChecked(False)

        form3.addRow("Marge (µL)", self.spin_margin)
        form3.addRow("Pas pipette (µL)", self.spin_step)
        form3.addRow(self.chk_zero)
        form3.addRow(self.chk_dup)

        layout.addWidget(box3)

        self.btn_moles = QPushButton("Quantités de matière…", self)
        self.btn_moles.clicked.connect(self._on_edit_moles_clicked)
        layout.addWidget(self.btn_moles)

        self.btn_reset_defaults = QPushButton("Valeurs par défaut", self)
        self.btn_reset_defaults.setToolTip("Remet tous les paramètres aux valeurs d'usine.")
        self.btn_reset_defaults.clicked.connect(self._reset_to_defaults)
        layout.addWidget(self.btn_reset_defaults)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, self)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def _apply_dropper_vol_mode(self, dropper_on: bool) -> None:
        """Bascule spin_VA1 / spin_VC / spin_VF entre mode µL et mode gouttes."""
        step = float(self.spin_step.value()) if dropper_on else 0.0
        self._updating_fixed = True
        try:
            if dropper_on:
                # Récupérer les volumes actuels en µL et les convertir en nombre de gouttes
                cur_va1_uL = float(self.spin_VA1.value())
                cur_vc_uL = float(self.spin_VC.value())
                cur_vf_uL = float(self.spin_VF.value())
                for spin, lbl in [
                    (self.spin_VA1, self._lbl_VA1),
                    (self.spin_VC,  self._lbl_VC),
                    (self.spin_VF,  self._lbl_VF),
                ]:
                    spin.setDecimals(0)
                    spin.setSingleStep(1.0)
                    spin.setSuffix(" gouttes")
                    lbl.setText("N gouttes")
                # Convertir en gouttes (arrondi) — A1 : 3 gouttes par défaut
                self.spin_VA1.setValue(max(1, round(cur_va1_uL / step)) if step > 0 else 3)
                self.spin_VC.setValue(max(1, round(cur_vc_uL / step)) if step > 0 else 1)
                self.spin_VF.setValue(max(1, round(cur_vf_uL / step)) if step > 0 else 1)
            else:
                # Récupérer le nombre de gouttes et reconvertir en µL
                prev_step = float(self.spin_step.value())
                cur_va1_g = float(self.spin_VA1.value())
                cur_vc_g = float(self.spin_VC.value())
                cur_vf_g = float(self.spin_VF.value())
                for spin, lbl in [
                    (self.spin_VA1, self._lbl_VA1),
                    (self.spin_VC,  self._lbl_VC),
                    (self.spin_VF,  self._lbl_VF),
                ]:
                    spin.setDecimals(1)
                    spin.setSingleStep(1.0)
                    spin.setSuffix("")
                    lbl.setText("V (µL)")
                self.spin_VA1.setValue(round(cur_va1_g * prev_step, 1) if prev_step > 0 else cur_va1_g)
                self.spin_VC.setValue(round(cur_vc_g * prev_step, 1) if prev_step > 0 else cur_vc_g)
                self.spin_VF.setValue(round(cur_vf_g * prev_step, 1) if prev_step > 0 else cur_vf_g)
        finally:
            self._updating_fixed = False

    # Valeurs par défaut de référence (pour le bouton "Réinitialiser")
    _DEFAULTS = {
        "n_tubes": 11, "C0_nM": 1000.0, "V_echant_uL": 1000.0,
        "C_titrant_uM": 5.0, "Vtot_uL": DEFAULT_VTOT_UL,
        "margin_uL": 20.0, "pipette_step_uL": 1.0,
        "include_zero": True, "add_control": False, "compte_goutte": False,
        "_V_A1": 100.0, "_V_C": 30.0, "_V_D": 750.0, "_V_E": 750.0, "_V_F": 60.0,
    }
    _DEFAULTS_FIXED_STOCK = {
        "Solution A1": (70.0, "mM"),
        "Solution C": (10.0, "µM"), "Solution D": (0.5, "% en masse"),
        "Solution E": (1.0, "mM"), "Solution F": (1.0, "mM"),
    }

    def _apply_params(self, params: dict, fixed_stock: dict | None = None) -> None:
        """Remplit les widgets avec les paramètres fournis.

        Le signal toggled de chk_dropper est temporairement déconnecté pour éviter
        que cocher la case écrase les volumes sauvegardés.
        """
        # Bloquer le couplage C↔V pendant le chargement pour éviter les recalculs en cascade
        self._updating_fixed = True
        try:
            # Déconnecter le preset compte-goutte le temps du chargement
            try:
                self.chk_dropper.toggled.disconnect()
            except RuntimeError:
                pass

            self.spin_n.setValue(int(params.get("n_tubes", self.spin_n.value())))
            self.spin_C0.setValue(float(params.get("C0_nM", self.spin_C0.value())))
            self.spin_Vech.setValue(float(params.get("V_echant_uL", self.spin_Vech.value())))
            self.spin_Ctit.setValue(float(params.get("C_titrant_uM", self.spin_Ctit.value())))
            self.spin_Vtot.setValue(float(params.get("Vtot_uL", self.spin_Vtot.value())))
            self.spin_margin.setValue(float(params.get("margin_uL", self.spin_margin.value())))
            self.spin_step.setValue(float(params.get("pipette_step_uL", self.spin_step.value())))
            self.chk_zero.setChecked(bool(params.get("include_zero", self.chk_zero.isChecked())))
            self.chk_dup.setChecked(bool(params.get("add_control", params.get("duplicate_last", self.chk_dup.isChecked()))))
            dropper_on = bool(params.get("compte_goutte", self.chk_dropper.isChecked()))
            self.chk_dropper.setChecked(dropper_on)
            # Appliquer le mode d'affichage (gouttes ou µL) avant de charger les volumes
            self._apply_dropper_vol_mode(dropper_on)
            _step = float(params.get("pipette_step_uL", self.spin_step.value()))
            if "_V_A1" in params:
                v_a1 = float(params["_V_A1"])
                self.spin_VA1.setValue(round(v_a1 / _step) if dropper_on and _step > 0 else round(v_a1, 1))
            if "_V_C" in params:
                v_c = float(params["_V_C"])
                # Les paramètres sauvegardés stockent toujours des µL → convertir si mode gouttes
                self.spin_VC.setValue(round(v_c / _step) if dropper_on and _step > 0 else round(v_c, 1))
            if "_V_D" in params:
                self.spin_VD.setValue(float(params["_V_D"]))
            if "_V_E" in params:
                self.spin_VE.setValue(float(params["_V_E"]))
            if "_V_F" in params:
                v_f = float(params["_V_F"])
                self.spin_VF.setValue(round(v_f / _step) if dropper_on and _step > 0 else round(v_f, 1))
            if fixed_stock is not None:
                for letter, spin_conc, combo in [
                    ("Solution A1", self.spin_conc_A1, self.combo_unit_A1),
                    ("Solution C", self.spin_conc_C, self.combo_unit_C),
                    ("Solution D", self.spin_conc_D, self.combo_unit_D),
                    ("Solution E", self.spin_conc_E, self.combo_unit_E),
                    ("Solution F", self.spin_conc_F, self.combo_unit_F),
                ]:
                    if letter in fixed_stock:
                        c_val, u_val = fixed_stock[letter]
                        spin_conc.setValue(float(c_val))
                        combo.setCurrentText(str(u_val))

            # Reconstruire les quantités de référence depuis les valeurs chargées (C × V)
            n_a1 = self._conc_to_n_nmol(float(self.spin_conc_A1.value()),
                                        str(self.combo_unit_A1.currentText()).strip(),
                                        float(self.spin_VA1.value()))
            if n_a1 is not None and n_a1 > 0:
                self._n_nmol_A1 = n_a1
            n_c = self._conc_to_n_nmol(float(self.spin_conc_C.value()),
                                       str(self.combo_unit_C.currentText()).strip(),
                                       float(self.spin_VC.value()))
            if n_c is not None and n_c > 0:
                self._n_nmol_C = n_c
            n_f = self._conc_to_n_nmol(float(self.spin_conc_F.value()),
                                       str(self.combo_unit_F.currentText()).strip(),
                                       float(self.spin_VF.value()))
            if n_f is not None and n_f > 0:
                self._n_nmol_F = n_f
        finally:
            self._updating_fixed = False

        # Reconnecter le preset compte-goutte
        def _update_dropper_mode(state):
            self._apply_dropper_vol_mode(state)
            if state:
                self.spin_margin.setValue(0.0)
                self.spin_step.setValue(40.0)
                self.spin_VA1.setValue(3.0)
                self.spin_VC.setValue(1.0)
                self.spin_VF.setValue(1.0)

        self.chk_dropper.toggled.connect(_update_dropper_mode)

    def _reset_to_defaults(self) -> None:
        """Remet tous les paramètres aux valeurs d'usine."""
        self._apply_params(self._DEFAULTS, self._DEFAULTS_FIXED_STOCK)

    def _spin_to_uL(self, spin: QDoubleSpinBox) -> float:
        """Retourne la valeur du spin en µL, qu'on soit en mode gouttes ou µL."""
        val = float(spin.value())
        if self.chk_dropper.isChecked():
            step = float(self.spin_step.value())
            return val * step if step > 0 else val
        return val

    def params(self) -> dict:
        return {
            "n_tubes": int(self.spin_n.value()),
            "C0_nM": float(self.spin_C0.value()),
            "V_echant_uL": float(self.spin_Vech.value()),
            "C_titrant_uM": float(self.spin_Ctit.value()),
            "Vtot_uL": float(self.spin_Vtot.value()),
            # En mode compte-goutte, spin_VA1/spin_VC/spin_VF contiennent des gouttes → reconvertir en µL
            "V_Solution_A1_uL": self._spin_to_uL(self.spin_VA1),
            "V_Solution_C_uL": self._spin_to_uL(self.spin_VC),
            "V_Solution_D_uL": float(self.spin_VD.value()),
            "V_Solution_E_uL": float(self.spin_VE.value()),
            "V_Solution_F_uL": self._spin_to_uL(self.spin_VF),
            "margin_uL": float(self.spin_margin.value()),
            "pipette_step_uL": float(self.spin_step.value()),
            "include_zero": bool(self.chk_zero.isChecked()),
            "add_control": bool(self.chk_dup.isChecked()),
            "compte_goutte": bool(self.chk_dropper.isChecked()),
        }

    def fixed_solution_stocks(self) -> dict[str, tuple[float, str]]:
        return {
            "Solution A1": (float(self.spin_conc_A1.value()), str(self.combo_unit_A1.currentText()).strip()),
            "Solution C": (float(self.spin_conc_C.value()), str(self.combo_unit_C.currentText()).strip()),
            "Solution D": (float(self.spin_conc_D.value()), str(self.combo_unit_D.currentText()).strip()),
            "Solution E": (float(self.spin_conc_E.value()), str(self.combo_unit_E.currentText()).strip()),
            "Solution F": (float(self.spin_conc_F.value()), str(self.combo_unit_F.currentText()).strip()),
        }

    def _conc_to_n_nmol(self, conc_val: float, unit: str, vol_uL: float) -> float | None:
        if vol_uL <= 0:
            return None
        factor = MOLAR_UNIT_FACTORS.get(str(unit).strip())
        if factor is None:
            return None
        c_m = float(conc_val) * factor
        return c_m * float(vol_uL) * 1e3

    def _n_nmol_to_conc(self, n_nmol: float, unit: str, vol_uL: float) -> float | None:
        if vol_uL <= 0:
            return None
        factor = MOLAR_UNIT_FACTORS.get(str(unit).strip())
        if factor is None:
            return None
        c_m = float(n_nmol) / (float(vol_uL) * 1e3)
        return c_m / factor

    def _on_edit_moles_clicked(self) -> None:
        """Ouvre un dialogue pour fixer les quantités de matière (nmol)."""

        entries: list[dict] = []

        v_echant = float(self.spin_Vech.value())
        n_cu = self._conc_to_n_nmol(float(self.spin_C0.value()), "nM", v_echant)
        entries.append(
            {
                "key": "Echantillon",
                "label": "Echantillon (Cu)",
                "value": n_cu if n_cu is not None else None,
                "editable": n_cu is not None,
                "note": "n = C0 × V échantillon",
            }
        )

        def _add_solution(
            key: str,
            label: str,
            conc_spin: QDoubleSpinBox,
            unit_combo: QComboBox,
            vol_spin: QDoubleSpinBox,
        ) -> None:
            unit = str(unit_combo.currentText()).strip()
            vol = float(vol_spin.value())
            conc = float(conc_spin.value())
            n = self._conc_to_n_nmol(conc, unit, vol)
            note = ""
            editable = n is not None
            if n is None:
                if vol <= 0:
                    note = "Volume nul"
                else:
                    note = "Unité non molaire"
            entries.append(
                {
                    "key": key,
                    "label": label,
                    "value": n if n is not None else None,
                    "editable": editable,
                    "note": note,
                }
            )

        _add_solution("Solution A1", "Solution A1 (NH3+, 70 mM)", self.spin_conc_A1, self.combo_unit_A1, self.spin_VA1)
        _add_solution("Solution C", "Solution C", self.spin_conc_C, self.combo_unit_C, self.spin_VC)
        _add_solution("Solution D", "Solution D", self.spin_conc_D, self.combo_unit_D, self.spin_VD)
        _add_solution("Solution E", "Solution E", self.spin_conc_E, self.combo_unit_E, self.spin_VE)
        _add_solution("Solution F", "Solution F", self.spin_conc_F, self.combo_unit_F, self.spin_VF)

        dlg = MolesDialog(entries, parent=self)
        if dlg.exec() != QDialog.Accepted:
            return

        values = dlg.values()
        if "Echantillon" in values and v_echant > 0:
            new_c0 = self._n_nmol_to_conc(values["Echantillon"], "nM", v_echant)
            if new_c0 is not None:
                self.spin_C0.setValue(float(new_c0))

        def _apply_solution(
            key: str,
            conc_spin: QDoubleSpinBox,
            unit_combo: QComboBox,
            vol_spin: QDoubleSpinBox,
        ) -> None:
            if key not in values:
                return
            unit = str(unit_combo.currentText()).strip()
            vol = float(vol_spin.value())
            new_conc = self._n_nmol_to_conc(values[key], unit, vol)
            if new_conc is not None:
                conc_spin.setValue(float(new_conc))

        # Solutions D et E : comportement classique (met à jour la concentration)
        _apply_solution("Solution D", self.spin_conc_D, self.combo_unit_D, self.spin_VD)
        _apply_solution("Solution E", self.spin_conc_E, self.combo_unit_E, self.spin_VE)

        # Solutions A1, C et F : met à jour n_nmol de référence, puis recalcule le volume
        def _apply_anchored(key: str, n_attr: str,
                            conc_spin: QDoubleSpinBox, unit_combo: QComboBox,
                            vol_spin: QDoubleSpinBox) -> None:
            if key not in values:
                return
            new_n = float(values[key])
            setattr(self, n_attr, new_n)
            unit = str(unit_combo.currentText()).strip()
            factor = MOLAR_UNIT_FACTORS.get(unit)
            if factor is None:
                return
            c_m = float(conc_spin.value()) * factor
            if c_m <= 0:
                return
            new_vol = new_n / (c_m * 1e3)
            self._updating_fixed = True
            try:
                vol_spin.setValue(round(new_vol, 1))
            finally:
                self._updating_fixed = False

        _apply_anchored("Solution A1", "_n_nmol_A1",
                        self.spin_conc_A1, self.combo_unit_A1, self.spin_VA1)
        _apply_anchored("Solution C", "_n_nmol_C",
                        self.spin_conc_C, self.combo_unit_C, self.spin_VC)
        _apply_anchored("Solution F", "_n_nmol_F",
                        self.spin_conc_F, self.combo_unit_F, self.spin_VF)
    

class MetadataCreatorWidget(QWidget):
    titration_visibility_changed = Signal(bool)
    metadata_saved_status_changed = Signal(bool)

    """Remplaçant de MetadataPicker pour créer les métadonnées directement dans l'application.

    Cette version ne lit plus de fichiers Excel : elle permet de construire
    les deux DataFrames nécessaires au pipeline existant :

    - df_comp : composition des tubes (colonnes typiques : "Tube", "C (EGTA) (M)",
      "V cuvette (mL)", "n(EGTA) (mol)", ...)
    - df_map  : correspondance entre les noms de spectres et les tubes
      (colonnes typiques : "Spectrum name", "Tube", "Sample description", ...)

    Les attributs publics `df_comp` et `df_map` peuvent ensuite être utilisés
    à la place de ceux de `MetadataPickerWidget` pour appeler
    `build_combined_dataframe_from_df` dans le reste du programme.
    """

    # Modèles prédéfinis de tableaux des volumes
    VOLUME_PRESETS: dict[str, pd.DataFrame] = {
       "Titration classique (historique)": pd.DataFrame(
            {
                "Réactif": [
                    "Echantillon",
                    "Contrôle",
                    "Solution A1",
                    "Solution A2",
                    "Solution B",
                    "Solution C",
                    "Solution D",
                    "Solution E",
                    "Solution F",
                ],
                "Concentration": ["0,5", "", "70", "7", "2", "2", "0,5", "1", "1"],
                "Unité": ["µM", "", "mM", "mM", "µM", "µM", "% en masse", "mM", "mM"],
                "Tube 1":  [1000, 0, 100, 400,   0,  30, 750, 750,  60],
                "Tube 2":  [1000, 0, 100, 311,  89,  30, 750, 750,  60],
                "Tube 3":  [1000, 0, 100, 221, 179,  30, 750, 750,  60],
                "Tube 4":  [1000, 0, 100, 199, 201,  30, 750, 750,  60],
                "Tube 5":  [1000, 0, 100, 177, 223,  30, 750, 750,  60],
                "Tube 6":  [1000, 0, 100, 154, 246,  30, 750, 750,  60],
                "Tube 7":  [1000, 0, 100, 132, 268,  30, 750, 750,  60],
                "Tube 8":  [1000, 0, 100,  88, 312,  30, 750, 750,  60],
                "Tube 9":  [1000, 0, 100,   0, 402,  30, 750, 750,  60],
                "Tube 10": [1000, 0, 100,   0, 446,  30, 750, 750,  60],
                "Tube 11": [   0, 1000, 100,   0, 446,  30, 750, 750,  60],
            }
        ),
        "Titration classique compte-gouttes (actuel)": pd.DataFrame(
            {
                "Réactif": [
                    "Echantillon",
                    "Contrôle",
                    "Solution A1",
                    "Solution A2",
                    "Solution B",
                    "Solution C",
                    "Solution D",
                    "Solution E",
                    "Solution F",
                ],
                "Concentration": ["1000", "", "58,333333", "7", "5", "7,5", "0,5", "1", "1,5"],
                "Unité": ["nM", "", "mM", "mM", "µM", "µM", "% en masse", "mM", "mM"],
                "Tube 1":  [1000, 0, 3, 11,  0, 1, 750, 750, 1],
                "Tube 2":  [1000, 0, 3, 10,  1, 1, 750, 750, 1],
                "Tube 3":  [1000, 0, 3,  9,  2, 1, 750, 750, 1],
                "Tube 4":  [1000, 0, 3,  8,  3, 1, 750, 750, 1],
                "Tube 5":  [1000, 0, 3,  7,  4, 1, 750, 750, 1],
                "Tube 6":  [1000, 0, 3,  6,  5, 1, 750, 750, 1],
                "Tube 7":  [1000, 0, 3,  5,  6, 1, 750, 750, 1],
                "Tube 8":  [1000, 0, 3,  4,  7, 1, 750, 750, 1],
                "Tube 9":  [1000, 0, 3,  3,  8, 1, 750, 750, 1],
                "Tube 10": [1000, 0, 3,  2,  9, 1, 750, 750, 1],
                "Tube 11": [1000, 0, 3,  1, 10, 1, 750, 750, 1],
                "Pas (µL)": [0, 0, 40, 40, 40, 40, 0, 0, 40],
            }
        ),
        "Titration Angelina": pd.DataFrame(
            {
                "Réactif": [
                    "Echantillon",
                    "Contrôle",
                    "Solution B",
                    "Solution C",
                    "Solution D",
                    "Solution E",
                ],
                "Concentration": ["0,5", "", "4", "2", "0,5", "1",],
                "Unité": ["µM", "", "mM",  "µM", "% en masse","mM"],
                "Tube 1":  [1000, 0, 500,   0, 750, 750],
                "Tube 2":  [1000, 0, 411,  89, 750, 750],
                "Tube 3":  [1000, 0, 321, 179, 750, 750],
                "Tube 4":  [1000, 0, 299, 201, 750, 750],
                "Tube 5":  [1000, 0, 277, 223, 750, 750],
                "Tube 6":  [1000, 0, 254, 246, 750, 750],
                "Tube 7":  [1000, 0, 232, 268, 750, 750],
                "Tube 8":  [1000, 0, 188, 312, 750, 750],
                "Tube 9":  [1000, 0,  98, 402, 750, 750],
                "Tube 10": [1000, 0,  54, 446, 750, 750],
                "Tube 11": [1000, 0,  54, 446, 750, 750],
            }
        ),
    }

    DEBIT_FLUX_COLUMNS = [
        "Largeur route",
        "Longueur 6 arches",
        "Hauteur eau",
        "Volume total (m3)",
        "Vitesse (s)",
        "Débit (m3/s)",
    ]
    DEBIT_FLUX_PREFIX = "Débit / flux - "

    def __init__(self, parent=None) -> None:
        super().__init__(parent)

        self.df_comp: pd.DataFrame | None = None # tableau de composition des tubes
        self.df_map: pd.DataFrame | None = None # tableau de correspondance spectres ↔ tubes
        # Indique que la correspondance spectres↔tubes doit être revue (champs d'en-tête ou volumes modifiés)
        self.map_dirty: bool = False
        self._metadata_dirty: bool = True
        self._protocol_states: dict = {}   # état des cases du protocole de paillasse
        # Options d'affichage pour la correspondance spectres ↔ tubes
        self._map_include_brb: bool = True
        self._map_include_control: bool = True
        # Derniers paramètres utilisés dans le dialogue gaussien (mémorisés entre ouvertures)
        self._last_gaussian_params: dict | None = None
        self._last_gaussian_fixed_stock: dict | None = None
        # Cible de contrôle pour la somme des volumes (µL) dans le tableau
        self._sum_target_uL: float = DEFAULT_VTOT_UL
        self._spec_start_index: int = 0
        self._ammonium_test_enabled: bool = False
        self._ammonium_test_values: dict[str, str] = {"test_1": "", "test_2": "", "ph": "", "operator": ""}
        self._debit_flux_values: dict[str, str] = {column: "" for column in self.DEBIT_FLUX_COLUMNS}
        self._fillable_fields: list = []
        self._extra_sampler_rows: list[dict] = []
        self._reverse_geocode_cache: dict[tuple[float, float], tuple[str, str]] = {}
        self._last_reverse_geocode_key: tuple[float, float] | None = None
        self._last_reverse_geocode_at: float = 0.0
        self._titration_name_manual: bool = False

        layout = QVBoxLayout(self)
        section_title_style = "color: #2f75b5; font-weight: 800; font-size: 15px;"

        # Première action proposée : charger une fiche existante.
        top_actions = QHBoxLayout()
        self.btn_load_meta = QPushButton("Charger des métadonnées…", self)
        self.btn_load_meta.setStyleSheet("background-color: #0057b8; color: white; font-weight: 700;")
        self.btn_load_meta.clicked.connect(self._on_load_metadata_clicked)
        top_actions.addWidget(self.btn_load_meta)
        top_actions.addStretch(1)
        layout.addLayout(top_actions)

        # Champs d'en-tête pour la manipulation : prélèvement puis titration.
        header_layout = QVBoxLayout()
        sample_title = QLabel("Prélèvement de l'échantillon :", self)
        sample_title.setStyleSheet(section_title_style)
        header_layout.addWidget(sample_title)

        # Ligne 1 : préleveur(s)
        row1 = QHBoxLayout()
        row1.addWidget(QLabel("Préleveur·se :", self))
        self.edit_sampler = QLineEdit(self)
        self.edit_sampler.setPlaceholderText("Prénom puis nom")
        row1.addWidget(self.edit_sampler)
        self.edit_sampler.textChanged.connect(self._on_sample_components_changed)

        row1.addWidget(QLabel("Association :", self))
        self.edit_sampler_association = QLineEdit(self)
        self.edit_sampler_association.setPlaceholderText("association")
        self.edit_sampler_association.textChanged.connect(self._on_header_field_changed)
        row1.addWidget(self.edit_sampler_association)
        self.edit_sample_associations = self.edit_sampler_association

        self.btn_add_sampler = QPushButton("+1 préleveur·se", self)
        self.btn_add_sampler.setToolTip("Ajouter une personne supplémentaire au prélèvement.")
        self.btn_add_sampler.clicked.connect(self._on_add_sampler_clicked)
        row1.addWidget(self.btn_add_sampler)

        row1.addStretch(1)

        header_layout.addLayout(row1)

        self.extra_samplers_layout = QVBoxLayout()
        header_layout.addLayout(self.extra_samplers_layout)

        # Ligne 2 : date + heure + lieu du prélèvement
        row2 = QHBoxLayout()
        row2.addWidget(QLabel("Date du prélèvement :", self))
        self.edit_sample_date = QDateEdit(self)
        self.edit_sample_date.setCalendarPopup(True)
        self.edit_sample_date.setDisplayFormat("dd/MM/yyyy")
        self.edit_sample_date.setMinimumDate(OPTIONAL_DATE_SENTINEL)
        self.edit_sample_date.setSpecialValueText("à renseigner")
        self.edit_sample_date.setDate(OPTIONAL_DATE_SENTINEL)
        self.edit_sample_date.dateChanged.connect(self._on_sample_components_changed)
        self._set_default_on_interaction(self.edit_sample_date, "current_date")
        row2.addWidget(self.edit_sample_date)

        row2.addWidget(QLabel("Heure du prélèvement :", self))
        self.edit_sample_time = QTimeEdit(self)
        self.edit_sample_time.setDisplayFormat("HH:mm")
        self.edit_sample_time.setMinimumTime(OPTIONAL_TIME_SENTINEL)
        self.edit_sample_time.setSpecialValueText("à renseigner")
        self.edit_sample_time.setTime(OPTIONAL_TIME_SENTINEL)
        self.edit_sample_time.timeChanged.connect(self._on_sample_components_changed)
        self._set_default_on_interaction(self.edit_sample_time, "current_time")
        row2.addWidget(self.edit_sample_time)

        row2.addWidget(QLabel("Lieu du prélèvement :", self))
        self.edit_sample_location = QLineEdit(self)
        self.edit_sample_location.setPlaceholderText("ex : site / point de prélèvement")
        row2.addWidget(self.edit_sample_location)
        self.edit_sample_location.textChanged.connect(self._on_header_field_changed)

        header_layout.addLayout(row2)

        row_sample_name = QHBoxLayout()
        row_sample_name.addWidget(QLabel("Nom du prélèvement :", self))
        self.edit_manip = QLineEdit(self)
        self.edit_manip.setReadOnly(True)
        self.edit_manip.setPlaceholderText("généré automatiquement")
        self.edit_manip.setFocusPolicy(Qt.NoFocus)
        self.edit_manip.setStyleSheet(
            "background-color: #263238; color: #ffffff; border: 1px solid #6fa8d6; "
            "font-weight: 700; padding: 3px 6px;"
        )
        self.edit_manip.setToolTip(
            "Généré depuis les initiales prénom+nom de chaque préleveur·se, "
            "puis _AAAAMMJJ_HHhMM."
        )
        row_sample_name.addWidget(self.edit_manip, 1)
        header_layout.addLayout(row_sample_name)

        row_gps = QHBoxLayout()
        row_gps.addWidget(QLabel("Latitude :", self))
        self.edit_sample_lat = QLineEdit(self)
        self.edit_sample_lat.setPlaceholderText("ex : 46.548861 ou 46°32'55.9\"N")
        self.edit_sample_lat.setToolTip(
            "Format standard : degrés décimaux WGS84, entre -90 et 90. "
            "Le format degrés/minutes/secondes est accepté et converti."
        )
        row_gps.addWidget(self.edit_sample_lat)
        self.edit_sample_lat.textChanged.connect(self._on_header_field_changed)
        self.edit_sample_lat.editingFinished.connect(self._normalize_sample_gps_fields)

        row_gps.addWidget(QLabel("Longitude :", self))
        self.edit_sample_lon = QLineEdit(self)
        self.edit_sample_lon.setPlaceholderText("ex : 0.368861 ou 0°22'07.9\"E")
        self.edit_sample_lon.setToolTip(
            "Format standard : degrés décimaux WGS84, entre -180 et 180. "
            "Le format degrés/minutes/secondes est accepté et converti."
        )
        row_gps.addWidget(self.edit_sample_lon)
        self.edit_sample_lon.textChanged.connect(self._on_header_field_changed)
        self.edit_sample_lon.editingFinished.connect(self._normalize_sample_gps_fields)

        self.btn_sample_map = QPushButton("Carte / pointer…", self)
        self.btn_sample_map.setToolTip(
            "Ouvre une carte interne : cliquez sur le point de prélèvement pour remplir latitude et longitude."
        )
        self.btn_sample_map.clicked.connect(self._on_view_sample_map_clicked)
        row_gps.addWidget(self.btn_sample_map)
        header_layout.addLayout(row_gps)

        row_admin = QHBoxLayout()
        row_admin.addWidget(QLabel("Département :", self))
        self.edit_sample_department = QLineEdit(self)
        self.edit_sample_department.setPlaceholderText("automatique si GPS + internet")
        self.edit_sample_department.textChanged.connect(self._on_header_field_changed)
        row_admin.addWidget(self.edit_sample_department)

        row_admin.addWidget(QLabel("Commune :", self))
        self.edit_sample_commune = QLineEdit(self)
        self.edit_sample_commune.setPlaceholderText("automatique si GPS + internet")
        self.edit_sample_commune.textChanged.connect(self._on_header_field_changed)
        row_admin.addWidget(self.edit_sample_commune)
        header_layout.addLayout(row_admin)

        row_water = QHBoxLayout()
        row_water.addWidget(QLabel("Type d'eau :", self))
        self.combo_water_type = QComboBox(self)
        self.combo_water_type.addItems(["Choisir...", "Eau douce", "Eau de mer", "Eau estuarienne"])
        self.combo_water_type.currentTextChanged.connect(self._on_water_type_changed)
        row_water.addWidget(self.combo_water_type)

        self.lbl_tide_coefficient = QLabel("Coefficient de marée :", self)
        row_water.addWidget(self.lbl_tide_coefficient)
        self.spin_tide_coefficient = QSpinBox(self)
        self.spin_tide_coefficient.setRange(0, 200)
        self.spin_tide_coefficient.setSpecialValueText("non renseigné")
        self.spin_tide_coefficient.valueChanged.connect(self._on_header_field_changed)
        row_water.addWidget(self.spin_tide_coefficient)

        self.lbl_high_tide_time = QLabel("Heure de pleine mer :", self)
        row_water.addWidget(self.lbl_high_tide_time)
        self.edit_high_tide_time = QLineEdit(self)
        self.edit_high_tide_time.setPlaceholderText("ex : 14:35")
        self.edit_high_tide_time.editingFinished.connect(self._normalize_high_tide_time)
        self.edit_high_tide_time.textChanged.connect(self._on_header_field_changed)
        row_water.addWidget(self.edit_high_tide_time)

        row_water.addWidget(QLabel("Température de l'eau :", self))
        self.spin_water_temperature = QDoubleSpinBox(self)
        self.spin_water_temperature.setRange(-10.0, 50.0)
        self.spin_water_temperature.setDecimals(1)
        self.spin_water_temperature.setSuffix(" °C")
        self.spin_water_temperature.setSpecialValueText("non renseigné")
        self.spin_water_temperature.setValue(self.spin_water_temperature.minimum())
        self.spin_water_temperature.valueChanged.connect(self._on_header_field_changed)
        self._set_default_on_interaction(self.spin_water_temperature, 20.0)
        row_water.addWidget(self.spin_water_temperature)

        self.btn_debit_flux = QPushButton("Débit / flux…", self)
        self.btn_debit_flux.setToolTip("Ouvre le tableau débit / flux pour les prélèvements en eau douce.")
        self.btn_debit_flux.clicked.connect(self._on_debit_flux_clicked)
        row_water.addWidget(self.btn_debit_flux)

        row_water.addStretch(1)
        header_layout.addLayout(row_water)
        self._refresh_tide_fields_enabled()
        self._refresh_water_type_style()

        row_weather = QHBoxLayout()
        row_weather.addWidget(QLabel("Météo (contexte de prélèvement) :", self))
        self.combo_weather = QComboBox(self)
        self.combo_weather.addItems([
            "Choisir...", "Ensoleillé", "Nuageux", "Couvert",
            "Petite pluie", "Pluie forte", "Orage",
        ])
        self.combo_weather.currentTextChanged.connect(self._on_header_field_changed)
        row_weather.addWidget(self.combo_weather)

        row_weather.addWidget(QLabel("Température de l'air :", self))
        self.spin_air_temperature = QDoubleSpinBox(self)
        self.spin_air_temperature.setRange(-30.0, 60.0)
        self.spin_air_temperature.setDecimals(1)
        self.spin_air_temperature.setSuffix(" °C")
        self.spin_air_temperature.setSpecialValueText("non renseigné")
        self.spin_air_temperature.setValue(self.spin_air_temperature.minimum())
        self.spin_air_temperature.valueChanged.connect(self._on_header_field_changed)
        self._set_default_on_interaction(self.spin_air_temperature, 20.0)
        row_weather.addWidget(self.spin_air_temperature)

        row_weather.addWidget(QLabel("Pluie dernières 24 h :", self))
        self.spin_rainfall_24h = QDoubleSpinBox(self)
        self.spin_rainfall_24h.setRange(-1.0, 500.0)
        self.spin_rainfall_24h.setDecimals(1)
        self.spin_rainfall_24h.setSuffix(" mm")
        self.spin_rainfall_24h.setSpecialValueText("non renseigné")
        self.spin_rainfall_24h.setValue(self.spin_rainfall_24h.minimum())
        self.spin_rainfall_24h.valueChanged.connect(self._on_header_field_changed)
        row_weather.addWidget(self.spin_rainfall_24h)
        row_weather.addStretch(1)
        header_layout.addLayout(row_weather)

        row_conditions = QHBoxLayout()
        row_conditions.addWidget(QLabel("Commentaire sur les conditions :", self))
        self.edit_conditions_comment = QTextEdit(self)
        self.edit_conditions_comment.setPlaceholderText("Décrivez ici toute observation sur les conditions environnementales…")
        self.edit_conditions_comment.setFixedHeight(70)
        self.edit_conditions_comment.textChanged.connect(self._on_header_field_changed)
        row_conditions.addWidget(self.edit_conditions_comment)
        header_layout.addLayout(row_conditions)

        layout.addLayout(header_layout)

        measures_layout = QVBoxLayout()
        measures_title = QLabel("Mesures effectuées :", self)
        measures_title.setStyleSheet(section_title_style)
        measures_layout.addWidget(measures_title)
        row_measures = QHBoxLayout()
        self.chk_ammonium_test = QCheckBox("Test ammonium", self)
        self.chk_ammonium_test.toggled.connect(self._on_ammonium_test_toggled)
        row_measures.addWidget(self.chk_ammonium_test)
        self.lbl_ammonium_operator = QLabel("Réalisé par :", self)
        self.lbl_ammonium_operator.setVisible(False)
        row_measures.addWidget(self.lbl_ammonium_operator)
        self.edit_ammonium_operator = QLineEdit(self)
        self.edit_ammonium_operator.setPlaceholderText("Prénom puis nom")
        self.edit_ammonium_operator.setVisible(False)
        self.edit_ammonium_operator.textChanged.connect(self._on_header_field_changed)
        row_measures.addWidget(self.edit_ammonium_operator)
        self.lbl_ammonium_test1 = QLabel("Test 1 (grossier) :", self)
        self.lbl_ammonium_test1.setVisible(False)
        row_measures.addWidget(self.lbl_ammonium_test1)
        self.edit_ammonium_test1 = QLineEdit(self)
        self.edit_ammonium_test1.setPlaceholderText("valeur")
        self.edit_ammonium_test1.setMaximumWidth(80)
        self.edit_ammonium_test1.setVisible(False)
        self.edit_ammonium_test1.textChanged.connect(self._on_header_field_changed)
        row_measures.addWidget(self.edit_ammonium_test1)
        self.lbl_ammonium_test2 = QLabel("Test 2 (précis) :", self)
        self.lbl_ammonium_test2.setVisible(False)
        row_measures.addWidget(self.lbl_ammonium_test2)
        self.edit_ammonium_test2 = QLineEdit(self)
        self.edit_ammonium_test2.setPlaceholderText("valeur")
        self.edit_ammonium_test2.setMaximumWidth(80)
        self.edit_ammonium_test2.setVisible(False)
        self.edit_ammonium_test2.textChanged.connect(self._on_header_field_changed)
        row_measures.addWidget(self.edit_ammonium_test2)
        self.lbl_ammonium_ph = QLabel("pH :", self)
        self.lbl_ammonium_ph.setVisible(False)
        row_measures.addWidget(self.lbl_ammonium_ph)
        self.edit_ammonium_ph = QLineEdit(self)
        self.edit_ammonium_ph.setPlaceholderText("valeur")
        self.edit_ammonium_ph.setMaximumWidth(60)
        self.edit_ammonium_ph.setVisible(False)
        self.edit_ammonium_ph.textChanged.connect(self._on_header_field_changed)
        row_measures.addWidget(self.edit_ammonium_ph)
        row_measures.addStretch(1)
        measures_layout.addLayout(row_measures)

        # Turbidité
        row_turbidity = QHBoxLayout()
        self.chk_turbidity = QCheckBox("Turbidité mesurée", self)
        self.chk_turbidity.toggled.connect(self._on_turbidity_toggled)
        row_turbidity.addWidget(self.chk_turbidity)
        self.lbl_turbidity_val = QLabel("Valeur :", self)
        self.lbl_turbidity_val.setVisible(False)
        row_turbidity.addWidget(self.lbl_turbidity_val)
        self.spin_turbidity = QDoubleSpinBox(self)
        self.spin_turbidity.setRange(0.0, 10000.0)
        self.spin_turbidity.setDecimals(1)
        self.spin_turbidity.setSuffix(" NTU")
        self.spin_turbidity.setSpecialValueText("non renseigné")
        self.spin_turbidity.setValue(self.spin_turbidity.minimum())
        self.spin_turbidity.setVisible(False)
        self.spin_turbidity.valueChanged.connect(self._on_header_field_changed)
        row_turbidity.addWidget(self.spin_turbidity)
        row_turbidity.addStretch(1)
        measures_layout.addLayout(row_turbidity)

        # Conductivité
        row_conductivity = QHBoxLayout()
        self.chk_conductivity = QCheckBox("Conductivité mesurée", self)
        self.chk_conductivity.toggled.connect(self._on_conductivity_toggled)
        row_conductivity.addWidget(self.chk_conductivity)
        self.lbl_conductivity_val = QLabel("Valeur :", self)
        self.lbl_conductivity_val.setVisible(False)
        row_conductivity.addWidget(self.lbl_conductivity_val)
        self.spin_conductivity = QDoubleSpinBox(self)
        self.spin_conductivity.setRange(0.0, 200000.0)
        self.spin_conductivity.setDecimals(1)
        self.spin_conductivity.setSuffix(" µS/cm")
        self.spin_conductivity.setSpecialValueText("non renseigné")
        self.spin_conductivity.setValue(self.spin_conductivity.minimum())
        self.spin_conductivity.setVisible(False)
        self.spin_conductivity.valueChanged.connect(self._on_header_field_changed)
        row_conductivity.addWidget(self.spin_conductivity)
        row_conductivity.addStretch(1)
        measures_layout.addLayout(row_conductivity)

        # pH de l'eau
        row_ph_water = QHBoxLayout()
        self.chk_ph_water = QCheckBox("pH de l'eau mesuré", self)
        self.chk_ph_water.toggled.connect(self._on_ph_water_toggled)
        row_ph_water.addWidget(self.chk_ph_water)
        self.lbl_ph_water_val = QLabel("Valeur :", self)
        self.lbl_ph_water_val.setVisible(False)
        row_ph_water.addWidget(self.lbl_ph_water_val)
        self.spin_ph_water = QDoubleSpinBox(self)
        self.spin_ph_water.setRange(0.0, 14.0)
        self.spin_ph_water.setDecimals(2)
        self.spin_ph_water.setSpecialValueText("non renseigné")
        self.spin_ph_water.setValue(self.spin_ph_water.minimum())
        self.spin_ph_water.setVisible(False)
        self.spin_ph_water.valueChanged.connect(self._on_header_field_changed)
        row_ph_water.addWidget(self.spin_ph_water)
        row_ph_water.addStretch(1)
        measures_layout.addLayout(row_ph_water)

        # Oxygène dissous
        row_dissolved_o2 = QHBoxLayout()
        self.chk_dissolved_o2 = QCheckBox("Oxygène dissous mesuré", self)
        self.chk_dissolved_o2.toggled.connect(self._on_dissolved_o2_toggled)
        row_dissolved_o2.addWidget(self.chk_dissolved_o2)
        self.lbl_dissolved_o2_val = QLabel("Valeur :", self)
        self.lbl_dissolved_o2_val.setVisible(False)
        row_dissolved_o2.addWidget(self.lbl_dissolved_o2_val)
        self.spin_dissolved_o2 = QDoubleSpinBox(self)
        self.spin_dissolved_o2.setRange(0.0, 30.0)
        self.spin_dissolved_o2.setDecimals(2)
        self.spin_dissolved_o2.setSuffix(" mg/L")
        self.spin_dissolved_o2.setSpecialValueText("non renseigné")
        self.spin_dissolved_o2.setValue(self.spin_dissolved_o2.minimum())
        self.spin_dissolved_o2.setVisible(False)
        self.spin_dissolved_o2.valueChanged.connect(self._on_header_field_changed)
        row_dissolved_o2.addWidget(self.spin_dissolved_o2)
        row_dissolved_o2.addStretch(1)
        measures_layout.addLayout(row_dissolved_o2)


        row_bacterio = QHBoxLayout()
        self.chk_bacterio_analysis = QCheckBox("Analyses bactériologiques", self)
        self.chk_bacterio_analysis.toggled.connect(self._on_bacterio_analysis_toggled)
        row_bacterio.addWidget(self.chk_bacterio_analysis)

        self.lbl_bacterio_deposit_day = QLabel("Jour de dépôt :", self)
        row_bacterio.addWidget(self.lbl_bacterio_deposit_day)
        self.edit_bacterio_deposit_day = QDateEdit(self)
        self.edit_bacterio_deposit_day.setCalendarPopup(True)
        self.edit_bacterio_deposit_day.setDisplayFormat("dd/MM/yyyy")
        self.edit_bacterio_deposit_day.setMinimumDate(OPTIONAL_DATE_SENTINEL)
        self.edit_bacterio_deposit_day.setSpecialValueText("à renseigner")
        self.edit_bacterio_deposit_day.setDate(OPTIONAL_DATE_SENTINEL)
        self.edit_bacterio_deposit_day.dateChanged.connect(self._on_header_field_changed)
        self._set_default_on_interaction(self.edit_bacterio_deposit_day, "current_date")
        row_bacterio.addWidget(self.edit_bacterio_deposit_day)

        self.lbl_bacterio_deposit_time = QLabel("Heure du dépôt :", self)
        row_bacterio.addWidget(self.lbl_bacterio_deposit_time)
        self.edit_bacterio_deposit_time = QTimeEdit(self)
        self.edit_bacterio_deposit_time.setDisplayFormat("HH:mm")
        self.edit_bacterio_deposit_time.setMinimumTime(OPTIONAL_TIME_SENTINEL)
        self.edit_bacterio_deposit_time.setSpecialValueText("à renseigner")
        self.edit_bacterio_deposit_time.setTime(OPTIONAL_TIME_SENTINEL)
        self.edit_bacterio_deposit_time.timeChanged.connect(self._on_header_field_changed)
        self._set_default_on_interaction(self.edit_bacterio_deposit_time, "current_time")
        row_bacterio.addWidget(self.edit_bacterio_deposit_time)

        self.lbl_bacterio_company = QLabel("Entreprise de mesure :", self)
        row_bacterio.addWidget(self.lbl_bacterio_company)
        self.edit_bacterio_company = QLineEdit(self)
        self.edit_bacterio_company.setPlaceholderText("ex : laboratoire / entreprise")
        self.edit_bacterio_company.textChanged.connect(self._on_header_field_changed)
        row_bacterio.addWidget(self.edit_bacterio_company)
        row_bacterio.addStretch(1)
        measures_layout.addLayout(row_bacterio)

        row_bacterio_results = QHBoxLayout()
        self.lbl_bacterio_results = QLabel("Si résultat :", self)
        row_bacterio_results.addWidget(self.lbl_bacterio_results)

        self.lbl_bacterio_ecoli = QLabel("Résultats E.Coli (npp/ml) :", self)
        row_bacterio_results.addWidget(self.lbl_bacterio_ecoli)
        self.edit_bacterio_ecoli = QLineEdit(self)
        self.edit_bacterio_ecoli.setPlaceholderText("valeur")
        self.edit_bacterio_ecoli.textChanged.connect(self._on_header_field_changed)
        row_bacterio_results.addWidget(self.edit_bacterio_ecoli)

        self.lbl_bacterio_enterococci = QLabel("Résultats Entérocoques intestinaux (npp/100ml) :", self)
        row_bacterio_results.addWidget(self.lbl_bacterio_enterococci)
        self.edit_bacterio_enterococci = QLineEdit(self)
        self.edit_bacterio_enterococci.setPlaceholderText("valeur")
        self.edit_bacterio_enterococci.textChanged.connect(self._on_header_field_changed)
        row_bacterio_results.addWidget(self.edit_bacterio_enterococci)
        row_bacterio_results.addStretch(1)
        measures_layout.addLayout(row_bacterio_results)
        self._refresh_bacterio_fields_visible()

        row_titration_measure = QHBoxLayout()
        self.chk_titration_done = QCheckBox("Titration du cuivre", self)
        self.chk_titration_done.toggled.connect(self._on_titration_toggled)
        row_titration_measure.addWidget(self.chk_titration_done)
        row_titration_measure.addStretch(1)
        measures_layout.addLayout(row_titration_measure)

        layout.addLayout(measures_layout)

        # Compatibilité avec l'ancien nommage : ces attributs pointent désormais
        # vers les champs de titration.
        titration_layout = QVBoxLayout()
        titration_title = QLabel("Information sur la titration :", self)
        titration_title.setStyleSheet(section_title_style)
        titration_layout.addWidget(titration_title)

        row_titration_people = QHBoxLayout()
        row_titration_people.addWidget(QLabel("Coordinateur·ice :", self))
        self.edit_coordinator = QLineEdit(self)
        self.edit_coordinator.setPlaceholderText("Prénom puis nom")
        row_titration_people.addWidget(self.edit_coordinator)
        self.edit_coordinator.textChanged.connect(self._on_titration_components_changed)

        row_titration_people.addWidget(QLabel("Opérateur·ice :", self))
        self.edit_operator = QLineEdit(self)
        self.edit_operator.setPlaceholderText("Prénom puis nom")
        row_titration_people.addWidget(self.edit_operator)
        self.edit_operator.textChanged.connect(self._on_titration_components_changed)
        titration_layout.addLayout(row_titration_people)

        row_titration_place = QHBoxLayout()
        row_titration_place.addWidget(QLabel("Lieu de la titration :", self))
        self.edit_titration_location = QLineEdit(self)
        self.edit_titration_location.setPlaceholderText("ex : Laboratoire / Terrain")
        row_titration_place.addWidget(self.edit_titration_location)
        self.edit_titration_location.textChanged.connect(self._on_titration_info_changed)

        row_titration_place.addWidget(QLabel("Date de la titration :", self))
        self.edit_titration_date = QDateEdit(self)
        self.edit_titration_date.setCalendarPopup(True)
        self.edit_titration_date.setDisplayFormat("dd/MM/yyyy")
        self.edit_titration_date.setMinimumDate(OPTIONAL_DATE_SENTINEL)
        self.edit_titration_date.setSpecialValueText("à renseigner")
        self.edit_titration_date.setDate(OPTIONAL_DATE_SENTINEL)
        self.edit_titration_date.dateChanged.connect(self._on_titration_components_changed)
        self._set_default_on_interaction(self.edit_titration_date, "current_date")
        row_titration_place.addWidget(self.edit_titration_date)

        row_titration_place.addWidget(QLabel("Heure de la titration :", self))
        self.edit_titration_time = QTimeEdit(self)
        self.edit_titration_time.setDisplayFormat("HH:mm")
        self.edit_titration_time.setMinimumTime(OPTIONAL_TIME_SENTINEL)
        self.edit_titration_time.setSpecialValueText("à renseigner")
        self.edit_titration_time.setTime(OPTIONAL_TIME_SENTINEL)
        self.edit_titration_time.timeChanged.connect(self._on_titration_components_changed)
        self._set_default_on_interaction(self.edit_titration_time, "current_time")
        row_titration_place.addWidget(self.edit_titration_time)
        titration_layout.addLayout(row_titration_place)

        row_titration_name = QHBoxLayout()
        row_titration_name.addWidget(QLabel("Nom de la manip de titration :", self))
        self.edit_titration_manip = QLineEdit(self)
        self.edit_titration_manip.setReadOnly(True)
        self.edit_titration_manip.setPlaceholderText("généré automatiquement")
        self.edit_titration_manip.setFocusPolicy(Qt.NoFocus)
        self.edit_titration_manip.setMinimumWidth(460)
        self.edit_titration_manip.setStyleSheet(
            "background-color: #263238; color: #ffffff; border: 1px solid #6fa8d6; "
            "font-weight: 700; padding: 3px 6px;"
        )
        self.edit_titration_manip.setToolTip(
            "Généré depuis les initiales prénom+nom de l'opérateur·ice puis du/de la coordinateur·ice, "
            "puis _AAAAMMJJ_HHhMM. Le bouton Modifier permet de saisir un nom manuel."
        )
        self.edit_titration_manip.textChanged.connect(self._on_titration_manual_name_text_changed)
        row_titration_name.addWidget(self.edit_titration_manip, 1)
        self.btn_toggle_titration_name_manual = QPushButton("Modifier", self)
        self.btn_toggle_titration_name_manual.setToolTip("Saisir ou quitter le mode nom manuel pour la titration.")
        self.btn_toggle_titration_name_manual.clicked.connect(self._on_toggle_titration_name_manual)
        row_titration_name.addWidget(self.btn_toggle_titration_name_manual)
        titration_layout.addLayout(row_titration_name)

        self.edit_location = self.edit_titration_location
        self.edit_date = self.edit_titration_date

        if hasattr(self, "edit_sampler") and self.edit_sampler.text().strip():
            self._refresh_manip_name()
        self._refresh_titration_manip_name()
        self.titration_details_widget = QWidget(self)
        self.titration_details_widget.setLayout(titration_layout)
        layout.addWidget(self.titration_details_widget)

        self.titration_workflow_widget = QWidget(self)
        titration_workflow_layout = QVBoxLayout(self.titration_workflow_widget)
        titration_workflow_layout.setContentsMargins(0, 0, 0, 0)
        titration_workflow_layout.setSpacing(6)

        glossary_html = (
            "<b>Glossaire des solutions utilisées :</b><ul>"
            "<li><b>Echantillon</b> : Echantillon d'eau à analyser (contient les ions Cu²⁺)</li>"
            "<li><b>Solution A1</b> : Tampon NH3+ concentré (70 mM) — volume fixe par tube. "
            "Maintient les particules de cuivre stables en solution.</li>"
            "<li><b>Solution A2</b> : Tampon NH3+ (7 mM) — milieu de base, contrôle du pH. "
            "Volume complémentaire à Solution B (A2 + B = constante par tube).</li>"
            "<li><b>Solution B</b> : Titrant (variable expérimentale principale, volume croissant de tube en tube)</li>"
            "<li><b>Solution C</b> : Indicateur SERS (molécule rapporteur suivie par spectroscopie, volume fixe)</li>"
            "<li><b>Solution D</b> : PEG (agent de crowding / modification du milieu, volume fixe)</li>"
            "<li><b>Solution E</b> : NPs (nanoparticules SERS, volume fixe)</li>"
            "<li><b>Solution F</b> : Crosslinker (ex. spermine, agent de réticulation, volume fixe)</li>"
            "</ul>"
            "Les tableaux sont éditables comme dans Excel mais sont stockés en interne sous forme de DataFrame pandas."
        )
        glossary_box = QTextBrowser(self)
        glossary_box.setHtml(glossary_html)
        glossary_box.setOpenExternalLinks(False)
        glossary_box.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        glossary_box.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        glossary_box.setFixedHeight(145)
        glossary_box.setStyleSheet("QTextBrowser { padding: 4px; }")
        titration_workflow_layout.addWidget(glossary_box)

        creation_info_label = QLabel(
            "<b>Création des métadonnées directement dans le programme</b><br>"
            "Utilisez les boutons ci-dessous pour créer / éditer :<ul>"
            "<li>le tableau des volumes (avec concentrations) pour chaque tube</li>"
            "<li>le tableau de correspondance entre les noms de spectres et les tubes</li>"
            "</ul>",
            self,
        )
        creation_info_label.setWordWrap(True)
        creation_info_label.setContentsMargins(0, 8, 0, 0)
        titration_workflow_layout.addWidget(creation_info_label)

        # --- Sélecteur de modèle de tableau des volumes ---
        preset_layout = QHBoxLayout()
        preset_layout.addWidget(QLabel("Modèle de tableau des volumes :", self))

        self.combo_volume_preset = QComboBox(self)
        self.combo_volume_preset.addItems(self.VOLUME_PRESETS.keys())
        self.combo_volume_preset.setCurrentText("Titration classique compte-gouttes (actuel)")
        preset_layout.addWidget(self.combo_volume_preset)

        # Le changement de modèle ne s'applique pas automatiquement : il faut valider.
        self.btn_validate_volume_preset = QPushButton("Valider le modèle", self)
        self.btn_validate_volume_preset.clicked.connect(self._on_validate_volume_preset_clicked)
        preset_layout.addWidget(self.btn_validate_volume_preset)

        self.btn_export_proto  = QPushButton("Feuille de protocole…", self)
        self.btn_export_proto.clicked.connect(self._on_export_protocol_clicked)
        preset_layout.addWidget(self.btn_export_proto)

        preset_layout.addStretch(1)

        self.btn_generate_volumes = QPushButton("Générer un tableau de volume…", self)
        self.btn_generate_volumes.setToolTip("Option avancée : générer un nouveau tableau de volumes paramétrable.")
        self.btn_generate_volumes.clicked.connect(self._on_generate_volumes_clicked)
        preset_layout.addWidget(self.btn_generate_volumes)
        titration_workflow_layout.addLayout(preset_layout)

        # --- Boutons pour ouvrir les éditeurs de tableaux ---
        btns = QHBoxLayout()
        self.btn_edit_comp = QPushButton("Créer / éditer le tableau des volumes", self)
        self.btn_edit_map = QPushButton("Créer / éditer la correspondance spectres ↔ tubes", self)
        self.btn_show_conc = QPushButton("Concentration dans les cuvettes (info)", self)

        # Boutons enregistrer / charger sur une seconde ligne
        self.btn_save_meta     = QPushButton("Enregistrer les métadonnées…", self)

        self.btn_edit_comp.clicked.connect(self._on_edit_comp_clicked)
        self.btn_edit_map.clicked.connect(self._on_edit_map_clicked)
        self.btn_show_conc.clicked.connect(self._on_show_conc_clicked)
        self.btn_save_meta.clicked.connect(self._on_save_metadata_clicked)

        # Première ligne
        btns.addWidget(self.btn_edit_comp)
        btns.addWidget(self.btn_edit_map)
        btns.addWidget(self.btn_show_conc)

        titration_workflow_layout.addLayout(btns)

        # Couleurs des boutons selon l'état (rouge si non défini, vert si prêt)
        self._refresh_button_states()

        # --- État ---
        self.lbl_status_comp = QLabel("Tableau des volumes : non défini", self)
        self.lbl_status_map = QLabel("Tableau de correspondance : non défini", self)
        titration_workflow_layout.addWidget(self.lbl_status_comp)
        titration_workflow_layout.addWidget(self.lbl_status_map)
        layout.addWidget(self.titration_workflow_widget)

        layout.addStretch(1)

        bottom_save = QHBoxLayout()
        bottom_save.addStretch(1)
        bottom_save.addWidget(self.btn_save_meta)
        bottom_save.addStretch(1)
        layout.addLayout(bottom_save)

        self._setup_fillable_field_styles()
        self._refresh_titration_fields_visible()

    def _setup_fillable_field_styles(self) -> None:
        self._fillable_fields = [
            self.edit_sampler,
            self.edit_sample_date,
            self.edit_sample_time,
            self.edit_sample_location,
            self.edit_sample_lat,
            self.edit_sample_lon,
            self.edit_sample_department,
            self.edit_sample_commune,
            self.edit_sampler_association,
            self.combo_water_type,
            self.spin_tide_coefficient,
            self.edit_high_tide_time,
            self.spin_water_temperature,
            self.combo_weather,
            self.spin_air_temperature,
            self.spin_rainfall_24h,
            self.edit_conditions_comment,
            self.edit_bacterio_deposit_day,
            self.edit_bacterio_deposit_time,
            self.edit_bacterio_company,
            self.edit_bacterio_ecoli,
            self.edit_bacterio_enterococci,
            self.edit_coordinator,
            self.edit_operator,
            self.edit_titration_location,
            self.edit_titration_date,
            self.edit_titration_time,
            self.combo_volume_preset,
        ]
        self._refresh_fillable_fields_style()
        for widget in self._fillable_fields:
            _install_field_fill_style(widget)

    def _refresh_fillable_fields_style(self) -> None:
        for widget in getattr(self, "_fillable_fields", []):
            _apply_field_fill_style(widget)

    def _set_default_on_interaction(self, widget, default_value) -> None:
        if widget is None:
            return
        widget.setProperty("_default_on_interaction", default_value)
        widget.installEventFilter(self)

    def eventFilter(self, obj, event) -> bool:
        default_value = obj.property("_default_on_interaction") if hasattr(obj, "property") else None
        if default_value is not None and event.type() in (QEvent.Type.FocusIn, QEvent.Type.MouseButtonPress):
            if not _field_has_value(obj):
                if isinstance(obj, QDateEdit):
                    obj.setDate(QDate.currentDate())
                elif isinstance(obj, QTimeEdit):
                    obj.setTime(QTime.currentTime())
                elif isinstance(obj, (QSpinBox, QDoubleSpinBox)):
                    obj.setValue(float(default_value))
                _apply_field_fill_style(obj)
        return super().eventFilter(obj, event)

    def _sampler_names(self) -> list[str]:
        names = []
        if hasattr(self, "edit_sampler"):
            first = self.edit_sampler.text().strip()
            if first:
                names.append(first)
        for row in getattr(self, "_extra_sampler_rows", []):
            edit = row.get("edit")
            if edit is not None and edit.text().strip():
                names.append(edit.text().strip())
        return names

    def _sampler_associations(self) -> list[str]:
        associations = []
        if hasattr(self, "edit_sampler_association"):
            associations.append(self.edit_sampler_association.text().strip())
        for row in getattr(self, "_extra_sampler_rows", []):
            edit = row.get("association_edit")
            associations.append(edit.text().strip() if edit is not None else "")
        return associations

    def _on_add_sampler_clicked(self) -> None:
        self._add_sampler_field()
        self._on_header_field_changed()

    def _add_sampler_field(self, text: str = "", association: str = "") -> QLineEdit:
        index = len(self._extra_sampler_rows) + 2
        row_widget = QWidget(self)
        row_layout = QHBoxLayout(row_widget)
        row_layout.setContentsMargins(0, 0, 0, 0)
        label = QLabel(f"Préleveur·se {index} :", row_widget)
        row_layout.addWidget(label)
        edit = QLineEdit(row_widget)
        edit.setPlaceholderText("Prénom puis nom")
        edit.textChanged.connect(self._on_sample_components_changed)
        row_layout.addWidget(edit)

        row_layout.addWidget(QLabel("Association :", row_widget))
        association_edit = QLineEdit(row_widget)
        association_edit.setPlaceholderText("association")
        association_edit.textChanged.connect(self._on_header_field_changed)
        row_layout.addWidget(association_edit)

        btn_remove = QPushButton("Retirer", row_widget)
        btn_remove.clicked.connect(lambda _checked=False, w=row_widget: self._remove_sampler_field(w))
        row_layout.addWidget(btn_remove)
        row_layout.addStretch(1)
        self.extra_samplers_layout.addWidget(row_widget)
        self._extra_sampler_rows.append({
            "widget": row_widget,
            "label": label,
            "edit": edit,
            "association_edit": association_edit,
            "button": btn_remove,
        })
        if hasattr(self, "_fillable_fields"):
            self._fillable_fields.extend([edit, association_edit])
            _install_field_fill_style(edit)
            _install_field_fill_style(association_edit)
        edit.setText(str(text or ""))
        association_edit.setText(str(association or ""))
        return edit

    def _remove_sampler_field(self, row_widget: QWidget) -> None:
        removed_edits = []
        for row in self._extra_sampler_rows:
            if row.get("widget") is row_widget:
                removed_edits.extend([row.get("edit"), row.get("association_edit")])
        self._extra_sampler_rows = [
            row for row in self._extra_sampler_rows if row.get("widget") is not row_widget
        ]
        if hasattr(self, "_fillable_fields"):
            self._fillable_fields = [w for w in self._fillable_fields if w not in removed_edits]
        row_widget.setParent(None)
        row_widget.deleteLater()
        for idx, row in enumerate(self._extra_sampler_rows, start=2):
            label = row.get("label")
            if label is not None:
                label.setText(f"Préleveur·se {idx} :")
        self._on_sample_components_changed()

    def _clear_extra_sampler_fields(self) -> None:
        removed_edits = []
        for row in getattr(self, "_extra_sampler_rows", []):
            removed_edits.extend([row.get("edit"), row.get("association_edit")])
        if hasattr(self, "_fillable_fields"):
            self._fillable_fields = [w for w in self._fillable_fields if w not in removed_edits]
        for row in list(getattr(self, "_extra_sampler_rows", [])):
            widget = row.get("widget")
            if widget is not None:
                widget.setParent(None)
                widget.deleteLater()
        self._extra_sampler_rows = []

    def reset_all(self) -> None:
        """Remet tous les champs à zéro : textes vidés, cases décochées, dates/heures en
        'à renseigner', tableaux chargés supprimés, états visuels mis à jour."""

        # --- Champs texte ---
        for attr in (
            "edit_sampler", "edit_sampler_association",
            "edit_sample_location", "edit_sample_lat", "edit_sample_lon",
            "edit_sample_department", "edit_sample_commune",
            "edit_ammonium_operator", "edit_ammonium_test1",
            "edit_ammonium_test2", "edit_ammonium_ph",
            "edit_bacterio_company", "edit_bacterio_ecoli", "edit_bacterio_enterococci",
            "edit_coordinator", "edit_operator",
            "edit_titration_location", "edit_high_tide_time",
        ):
            w = getattr(self, attr, None)
            if w is not None:
                w.blockSignals(True)
                w.setText("")
                w.blockSignals(False)

        comment = getattr(self, "edit_conditions_comment", None)
        if comment is not None:
            comment.blockSignals(True)
            comment.clear()
            comment.blockSignals(False)

        # --- Dates → "à renseigner" ---
        for attr in ("edit_sample_date", "edit_bacterio_deposit_day", "edit_titration_date"):
            w = getattr(self, attr, None)
            if w is not None:
                w.blockSignals(True)
                w.setDate(OPTIONAL_DATE_SENTINEL)
                w.blockSignals(False)

        # --- Heures → "à renseigner" ---
        for attr in ("edit_sample_time", "edit_bacterio_deposit_time", "edit_titration_time"):
            w = getattr(self, attr, None)
            if w is not None:
                w.blockSignals(True)
                w.setTime(OPTIONAL_TIME_SENTINEL)
                w.blockSignals(False)

        # --- Combos → premier item ("Choisir...") ---
        for attr in ("combo_water_type", "combo_weather"):
            w = getattr(self, attr, None)
            if w is not None:
                w.blockSignals(True)
                w.setCurrentIndex(0)
                w.blockSignals(False)

        # --- SpinBox → valeur minimale (special value text) ---
        for attr in (
            "spin_tide_coefficient",
            "spin_water_temperature", "spin_air_temperature", "spin_rainfall_24h",
            "spin_turbidity", "spin_conductivity", "spin_ph_water", "spin_dissolved_o2",
        ):
            w = getattr(self, attr, None)
            if w is not None:
                w.blockSignals(True)
                w.setValue(w.minimum())
                w.blockSignals(False)

        # --- Cases à cocher → décochées (signaux bloqués ; visibilités gérées explicitement après) ---
        for attr in (
            "chk_ammonium_test", "chk_bacterio_analysis", "chk_titration_done",
            "chk_turbidity", "chk_conductivity", "chk_ph_water", "chk_dissolved_o2",
        ):
            w = getattr(self, attr, None)
            if w is not None:
                w.blockSignals(True)
                w.setChecked(False)
                w.blockSignals(False)

        # --- Préleveurs supplémentaires ---
        self._clear_extra_sampler_fields()

        # --- Nom manuel de la titration → retour au mode auto ---
        self._set_titration_name_manual_mode(False)

        # --- Tableaux et état interne ---
        self.df_comp = None
        self.df_map = None
        self._protocol_states = {}
        self._ammonium_test_enabled = False
        self._ammonium_test_values = {"test_1": "", "test_2": "", "ph": "", "operator": ""}
        self._debit_flux_values = {column: "" for column in self.DEBIT_FLUX_COLUMNS}
        self.map_dirty = False
        self._spec_start_index = 0
        self._last_titration_visibility = None  # force la ré-émission du signal

        # --- Labels de statut ---
        if hasattr(self, "lbl_status_comp"):
            self.lbl_status_comp.setText("Tableau des volumes : non défini")
        if hasattr(self, "lbl_status_map"):
            self.lbl_status_map.setText("Tableau de correspondance : non défini")

        # --- Visibilités des champs dépendant des cases à cocher ---
        self._on_ammonium_test_toggled(False)
        self._on_turbidity_toggled(False)
        self._on_conductivity_toggled(False)
        self._on_ph_water_toggled(False)
        self._on_dissolved_o2_toggled(False)
        self._refresh_bacterio_fields_visible()
        self._refresh_titration_fields_visible()

        # --- Rafraîchissement visuel final ---
        self._refresh_manip_name()
        self._refresh_titration_manip_name(force=True)
        self._refresh_fillable_fields_style()
        self._refresh_button_states()
        self._mark_metadata_dirty(refresh=False)

    def _refresh_button_states(self) -> None:
        """Met à jour la couleur des boutons selon l'état des DataFrames.

        - rouge si le tableau n'est pas défini
        - vert si le tableau est prêt (df non vide)
        """
        red = "background-color: #d9534f; color: white; font-weight: 600;"
        green = "background-color: #5cb85c; color: white; font-weight: 600;"
        neutral = ""  # laisse le style par défaut

        comp_ready = self.df_comp is not None and isinstance(self.df_comp, pd.DataFrame) and not self.df_comp.empty
        map_ready = (
            self.df_map is not None
            and isinstance(self.df_map, pd.DataFrame)
            and not self.df_map.empty
            and not self.map_dirty
        )

        # Bouton volumes
        if comp_ready:
            self.btn_edit_comp.setStyleSheet(green)
        else:
            self.btn_edit_comp.setStyleSheet(red)

        # Bouton correspondance spectres↔tubes
        if map_ready:
            self.btn_edit_map.setStyleSheet(green)
        else:
            self.btn_edit_map.setStyleSheet(red)

        # Bouton concentrations : vert si le tableau des volumes existe (calcul possible),
        # sinon rouge.
        if comp_ready:
            self.btn_show_conc.setStyleSheet(green)
        else:
            self.btn_show_conc.setStyleSheet(red)

        # Bouton protocole : prêt dès que le tableau des volumes existe.
        if comp_ready:
            self.btn_export_proto.setStyleSheet(green)
        else:
            self.btn_export_proto.setStyleSheet(red)

        # Bouton sauvegarde : rouge dès qu'une modification n'a pas été enregistrée,
        # vert uniquement après un chargement ou un enregistrement réussi.
        metadata_saved = not getattr(self, "_metadata_dirty", True)
        if metadata_saved:
            self.btn_save_meta.setStyleSheet(green)
        else:
            self.btn_save_meta.setStyleSheet(red)

        # Mettre à jour les tooltips
        self.btn_edit_comp.setToolTip("Vert = tableau des volumes défini" if comp_ready else "Rouge = à créer / éditer")
        self.btn_edit_map.setToolTip("Vert = correspondance définie" if map_ready else "Rouge = à créer / éditer")
        self.btn_save_meta.setToolTip(
            "Vert = métadonnées enregistrées" if metadata_saved else "Rouge = modifications à enregistrer"
        )
        if comp_ready:
            self.btn_show_conc.setToolTip("Calculable (volumes définis)")
            self.btn_export_proto.setToolTip("Vert = feuille de protocole prête")
        else:
            self.btn_show_conc.setToolTip("Rouge = définir d'abord les volumes")
            self.btn_export_proto.setToolTip("Rouge = définir d'abord le tableau des volumes")
# ------------------------------------------------------------------
# Helpers – recalcul à partir d'un tableau des concentrations édité
# ------------------------------------------------------------------
    def _mark_metadata_dirty(self, refresh: bool = True) -> None:
        self._metadata_dirty = True
        if refresh and hasattr(self, "btn_save_meta"):
            self._refresh_button_states()
        self.metadata_saved_status_changed.emit(False)

    def _mark_metadata_saved(self) -> None:
        self._metadata_dirty = False
        if hasattr(self, "btn_save_meta"):
            self._refresh_button_states()
        self.metadata_saved_status_changed.emit(True)

    def has_unsaved_metadata(self) -> bool:
        return bool(getattr(self, "_metadata_dirty", True))

    def _metadata_save_warnings(self) -> list[str]:
        warnings: list[str] = []
        titration_done = bool(self.chk_titration_done.isChecked()) if hasattr(self, "chk_titration_done") else False
        comp_ready = self.df_comp is not None and isinstance(self.df_comp, pd.DataFrame) and not self.df_comp.empty

        if titration_done:
            if not comp_ready:
                warnings.append("La titration est cochée, mais aucun tableau des volumes n'est défini.")
            if not self._protocol_states:
                warnings.append("La feuille de protocole n'a pas encore été validée.")

        if self._ammonium_test_enabled:
            missing = []
            if not getattr(self, "edit_ammonium_operator", None) or not self.edit_ammonium_operator.text().strip():
                missing.append("réalisateur·ice")
            if not getattr(self, "edit_ammonium_test1", None) or not self.edit_ammonium_test1.text().strip():
                missing.append("test 1")
            if not getattr(self, "edit_ammonium_test2", None) or not self.edit_ammonium_test2.text().strip():
                missing.append("test 2")
            if not getattr(self, "edit_ammonium_ph", None) or not self.edit_ammonium_ph.text().strip():
                missing.append("pH")
            if missing:
                warnings.append("Test ammonium coché, mais champ(s) manquant(s) : " + ", ".join(missing) + ".")

        if hasattr(self, "chk_bacterio_analysis") and self.chk_bacterio_analysis.isChecked():
            values = self._header_values()
            required = {
                "Jour de dépôt": values.get("Jour de dépôt", ""),
                "Heure du dépôt": values.get("Heure du dépôt", ""),
                "Entreprise de mesure": values.get("Entreprise de mesure", ""),
                "Résultats E.Coli (npp/ml)": values.get("Résultats E.Coli (npp/ml)", ""),
                "Résultats Entérocoques intestinaux (npp/100ml)": values.get(
                    "Résultats Entérocoques intestinaux (npp/100ml)", ""
                ),
            }
            missing = [label for label, value in required.items() if not str(value or "").strip()]
            if missing:
                warnings.append("Analyses bactériologiques cochées, mais champ(s) manquant(s) : " + ", ".join(missing) + ".")
        return warnings

    @staticmethod
    def _date_time_text(widget) -> str:
        if widget is None:
            return ""
        if isinstance(widget, QDateEdit) and widget.specialValueText() and widget.date() == widget.minimumDate():
            return ""
        if isinstance(widget, QDateTimeEdit) and widget.specialValueText() and widget.dateTime() == widget.minimumDateTime():
            return ""
        if isinstance(widget, QDateEdit):
            return widget.date().toString("dd/MM/yyyy")
        try:
            return widget.dateTime().toString("dd/MM/yyyy HH:mm")
        except Exception:
            try:
                return widget.date().toString("dd/MM/yyyy")
            except Exception:
                return ""

    @staticmethod
    def _set_date_time_text(widget, text: str) -> None:
        if widget is None:
            return
        txt = str(text or "").strip()
        if not txt:
            return
        qdt = QDateTime.fromString(txt, "dd/MM/yyyy HH:mm")
        if qdt.isValid():
            if isinstance(widget, QDateEdit):
                widget.setDate(qdt.date())
                return
            try:
                widget.setDateTime(qdt)
            except Exception:
                try:
                    widget.setDate(qdt.date())
                except Exception:
                    pass
            return
        qd = QDate.fromString(txt, "dd/MM/yyyy")
        if qd.isValid():
            if isinstance(widget, QDateEdit):
                widget.setDate(qd)
                return
            try:
                widget.setDateTime(QDateTime(qd, widget.time()))
            except Exception:
                widget.setDate(qd)

    @staticmethod
    def _time_text(widget) -> str:
        if widget is None:
            return ""
        if isinstance(widget, QTimeEdit) and widget.specialValueText() and widget.time() == widget.minimumTime():
            return ""
        try:
            return widget.time().toString("HH:mm")
        except Exception:
            try:
                return str(widget.text()).strip()
            except Exception:
                return ""

    @staticmethod
    def _set_time_text(widget, text: str) -> None:
        if widget is None:
            return
        txt = str(text or "").strip()
        if not txt:
            return
        if " " in txt:
            txt = txt.rsplit(" ", 1)[-1]
        normalized = MetadataCreatorWidget._normalize_time_text(txt)
        qt = QTime.fromString(normalized, "HH:mm")
        if qt.isValid():
            try:
                widget.setTime(qt)
                return
            except Exception:
                pass
        try:
            widget.setText(normalized)
        except Exception:
            pass

    @staticmethod
    def _normalize_time_text(text: str | None) -> str:
        txt = str(text or "").strip().lower().replace(" ", "")
        if not txt:
            return ""
        txt = txt.replace("h", ":").replace(".", ":")
        if re.fullmatch(r"\d{3,4}", txt):
            txt = f"{txt[:-2]}:{txt[-2:]}"
        match = re.fullmatch(r"(\d{1,2})(?::(\d{1,2}))?", txt)
        if not match:
            return str(text or "").strip()
        hour = int(match.group(1))
        minute = int(match.group(2) or 0)
        if not (0 <= hour <= 23 and 0 <= minute <= 59):
            return str(text or "").strip()
        return f"{hour:02d}:{minute:02d}"

    def _normalize_high_tide_time(self) -> None:
        if not hasattr(self, "edit_high_tide_time"):
            return
        txt = self.edit_high_tide_time.text().strip()
        if not txt:
            return
        normalized = self._normalize_time_text(txt)
        if normalized != txt:
            self.edit_high_tide_time.setText(normalized)

    def _sample_date_text(self) -> str:
        return self._date_time_text(getattr(self, "edit_sample_date", None))

    def _sample_time_text(self) -> str:
        return self._time_text(getattr(self, "edit_sample_time", None))

    def _sample_date_time_text(self) -> str:
        date_text = self._sample_date_text()
        time_text = self._sample_time_text()
        if date_text and time_text:
            return f"{date_text} {time_text}"
        return ""

    def _titration_date_text(self) -> str:
        return self._date_time_text(getattr(self, "edit_titration_date", None))

    def _titration_time_text(self) -> str:
        return self._time_text(getattr(self, "edit_titration_time", None))

    def _titration_date_time_text(self) -> str:
        date_text = self._titration_date_text()
        time_text = self._titration_time_text()
        if date_text and time_text:
            return f"{date_text} {time_text}"
        return ""

    def _set_sample_date_time_text(self, text: str) -> None:
        txt = str(text or "").strip()
        if not txt:
            return
        if " " in txt:
            date_part, time_part = txt.rsplit(" ", 1)
        else:
            date_part, time_part = txt, ""
        if hasattr(self, "edit_sample_date"):
            self._set_date_time_text(self.edit_sample_date, date_part)
        if time_part and hasattr(self, "edit_sample_time"):
            self._set_time_text(self.edit_sample_time, time_part)

    def _set_titration_date_time_text(self, text: str) -> None:
        txt = str(text or "").strip()
        if not txt:
            return
        if " " in txt:
            date_part, time_part = txt.rsplit(" ", 1)
        else:
            date_part, time_part = txt, ""
        if hasattr(self, "edit_titration_date"):
            self._set_date_time_text(self.edit_titration_date, date_part)
        if time_part and hasattr(self, "edit_titration_time"):
            self._set_time_text(self.edit_titration_time, time_part)

    def _normalize_bacterio_deposit_time(self) -> None:
        if not hasattr(self, "edit_bacterio_deposit_time"):
            return
        txt = self._time_text(self.edit_bacterio_deposit_time)
        if not txt:
            return
        normalized = self._normalize_time_text(txt)
        if normalized != txt:
            self._set_time_text(self.edit_bacterio_deposit_time, normalized)

    def _is_tidal_water_type(self) -> bool:
        if not hasattr(self, "combo_water_type"):
            return False
        water_type = self._norm_text_key(self.combo_water_type.currentText())
        return water_type in {"eau de mer", "eau salee", "eau estuarienne", "eau esturienne"}

    def _is_fresh_water_type(self) -> bool:
        if not hasattr(self, "combo_water_type"):
            return False
        return self._norm_text_key(self.combo_water_type.currentText()) == "eau douce"

    def _is_water_type_selected(self) -> bool:
        if not hasattr(self, "combo_water_type"):
            return False
        return self._norm_text_key(self.combo_water_type.currentText()) not in {"", "choisir..."}

    def _refresh_water_type_style(self) -> None:
        if not hasattr(self, "combo_water_type"):
            return
        _apply_field_fill_style(self.combo_water_type)

    def _refresh_tide_fields_enabled(self) -> None:
        visible = self._is_tidal_water_type()
        for widget_name in (
            "lbl_tide_coefficient",
            "spin_tide_coefficient",
            "lbl_high_tide_time",
            "edit_high_tide_time",
        ):
            widget = getattr(self, widget_name, None)
            if widget is not None:
                widget.setVisible(visible)
                widget.setEnabled(visible)
        btn_debit_flux = getattr(self, "btn_debit_flux", None)
        if btn_debit_flux is not None:
            btn_debit_flux.setVisible(self._is_fresh_water_type())
            btn_debit_flux.setEnabled(self._is_fresh_water_type())

    def _on_water_type_changed(self, *args) -> None:
        self._refresh_tide_fields_enabled()
        self._refresh_water_type_style()
        self._on_header_field_changed()

    def _refresh_bacterio_fields_visible(self) -> None:
        visible = bool(self.chk_bacterio_analysis.isChecked()) if hasattr(self, "chk_bacterio_analysis") else False
        for widget_name in (
            "lbl_bacterio_deposit_day",
            "edit_bacterio_deposit_day",
            "lbl_bacterio_deposit_time",
            "edit_bacterio_deposit_time",
            "lbl_bacterio_company",
            "edit_bacterio_company",
            "lbl_bacterio_results",
            "lbl_bacterio_ecoli",
            "edit_bacterio_ecoli",
            "lbl_bacterio_enterococci",
            "edit_bacterio_enterococci",
        ):
            widget = getattr(self, widget_name, None)
            if widget is not None:
                widget.setVisible(visible)

    def _on_bacterio_analysis_toggled(self, checked: bool) -> None:
        self._refresh_bacterio_fields_visible()
        self._on_header_field_changed()

    def _refresh_titration_fields_visible(self) -> None:
        visible = bool(self.chk_titration_done.isChecked()) if hasattr(self, "chk_titration_done") else False
        for attr in ("titration_details_widget", "titration_workflow_widget"):
            widget = getattr(self, attr, None)
            if widget is not None:
                widget.setVisible(visible)
        if getattr(self, "_last_titration_visibility", None) != visible:
            self._last_titration_visibility = visible
            self.titration_visibility_changed.emit(visible)

    def _on_titration_toggled(self, checked: bool) -> None:
        self._refresh_titration_fields_visible()
        self._on_header_field_changed()

    @staticmethod
    def _person_initials(name: str) -> str:
        # Saisie attendue : "Prénom puis nom".
        parts = re.findall(r"[A-Za-zÀ-ÖØ-öø-ÿ0-9-]+", str(name or ""))
        if len(parts) >= 2:
            return f"{parts[0][0]}{parts[-1][0]}".upper()
        if len(parts) == 1:
            return parts[0][0].upper()
        return ""

    @staticmethod
    def _parse_coord_value(text: str | None) -> float | None:
        raw = str(text or "").strip()
        txt = raw.replace(",", ".")
        if not txt:
            return None

        hemisphere_match = re.search(r"([NSEW])", txt, flags=re.IGNORECASE)
        has_dms_marker = bool(re.search(r"[°º'’′\"”″]", txt))
        if hemisphere_match or has_dms_marker:
            numbers = re.findall(r"[-+]?\d+(?:\.\d+)?", txt)
            if numbers:
                deg = float(numbers[0])
                minutes = float(numbers[1]) if len(numbers) >= 2 else 0.0
                seconds = float(numbers[2]) if len(numbers) >= 3 else 0.0
                sign = -1.0 if deg < 0 else 1.0
                if hemisphere_match and hemisphere_match.group(1).upper() in ("S", "W"):
                    sign = -1.0
                elif hemisphere_match and hemisphere_match.group(1).upper() in ("N", "E"):
                    sign = 1.0
                return sign * (abs(deg) + minutes / 60.0 + seconds / 3600.0)

        try:
            return float(txt)
        except ValueError:
            return None

    @classmethod
    def _format_coord_value(cls, text: str | None, min_value: float, max_value: float) -> str:
        val = cls._parse_coord_value(text)
        if val is None or val < min_value or val > max_value:
            return str(text or "").strip()
        return f"{val:.6f}"

    @classmethod
    def _split_gps_pair(cls, text: str | None) -> tuple[str, str]:
        raw = str(text or "")
        hemisphere_chunks: list[str] = []
        start = 0
        for match in re.finditer(r"[NSEW]", raw, flags=re.IGNORECASE):
            hemisphere_chunks.append(raw[start:match.end()])
            start = match.end()

        lat = ""
        lon = ""
        for chunk in hemisphere_chunks:
            val = cls._parse_coord_value(chunk)
            if val is None:
                continue
            hemisphere = re.search(r"([NSEW])", chunk, flags=re.IGNORECASE)
            if hemisphere and hemisphere.group(1).upper() in ("N", "S"):
                lat = cls._format_coord_value(str(val), -90.0, 90.0)
            elif hemisphere and hemisphere.group(1).upper() in ("E", "W"):
                lon = cls._format_coord_value(str(val), -180.0, 180.0)
        if lat or lon:
            return lat, lon

        numbers = re.findall(r"[-+]?\d+(?:[,.]\d+)?", raw)
        if len(numbers) < 2:
            return "", ""
        if len(numbers) >= 6 and re.search(r"[°º'’′\"”″]", raw):
            lat_raw = f"{numbers[0]}°{numbers[1]}'{numbers[2]}\""
            lon_raw = f"{numbers[3]}°{numbers[4]}'{numbers[5]}\""
            lat = cls._format_coord_value(lat_raw, -90.0, 90.0)
            lon = cls._format_coord_value(lon_raw, -180.0, 180.0)
            return lat, lon
        lat = cls._format_coord_value(numbers[0], -90.0, 90.0)
        lon = cls._format_coord_value(numbers[1], -180.0, 180.0)
        return lat, lon

    def _normalize_coord_field(self, widget: QLineEdit, min_value: float, max_value: float) -> None:
        if widget is None:
            return
        txt = widget.text().strip()
        if not txt:
            return
        formatted = self._format_coord_value(txt, min_value, max_value)
        if formatted != txt:
            widget.setText(formatted)

    def _sample_gps_decimal_pair(self) -> tuple[float, float] | None:
        if not hasattr(self, "edit_sample_lat") or not hasattr(self, "edit_sample_lon"):
            return None
        lat = self._parse_coord_value(self.edit_sample_lat.text())
        lon = self._parse_coord_value(self.edit_sample_lon.text())
        if lat is None or lon is None or not (-90 <= lat <= 90) or not (-180 <= lon <= 180):
            return None
        return round(float(lat), 6), round(float(lon), 6)

    @staticmethod
    def _admin_from_nominatim_payload(payload: dict) -> tuple[str, str]:
        address = payload.get("address", {}) if isinstance(payload, dict) else {}
        department = (
            address.get("county")
            or address.get("state_district")
            or address.get("region")
            or ""
        )
        commune = (
            address.get("city")
            or address.get("town")
            or address.get("village")
            or address.get("municipality")
            or address.get("commune")
            or address.get("hamlet")
            or address.get("locality")
            or ""
        )
        return str(department or "").strip(), str(commune or "").strip()

    def _reverse_geocode_sample_admin(self, lat: float, lon: float) -> tuple[str, str] | None:
        key = (round(float(lat), 6), round(float(lon), 6))
        if key in self._reverse_geocode_cache:
            return self._reverse_geocode_cache[key]

        now = time.monotonic()
        if self._last_reverse_geocode_key == key and now - self._last_reverse_geocode_at < 30.0:
            return None
        if now - self._last_reverse_geocode_at < 1.0:
            return None

        params = urllib.parse.urlencode({
            "format": "jsonv2",
            "addressdetails": 1,
            "lat": f"{lat:.6f}",
            "lon": f"{lon:.6f}",
            "zoom": 18,
            "accept-language": "fr",
        })
        request = urllib.request.Request(
            f"https://nominatim.openstreetmap.org/reverse?{params}",
            headers={
                "User-Agent": "Ramanalyze/1.0 (desktop metadata geocoder)",
                "Accept": "application/json",
            },
        )
        self._last_reverse_geocode_at = now
        self._last_reverse_geocode_key = key
        try:
            with urllib.request.urlopen(request, timeout=3) as response:
                payload = json.loads(response.read().decode("utf-8"))
        except Exception:
            return None

        admin = self._admin_from_nominatim_payload(payload)
        if any(admin):
            self._reverse_geocode_cache[key] = admin
        return admin

    def _try_autofill_sample_admin_from_gps(self) -> None:
        pair = self._sample_gps_decimal_pair()
        if pair is None:
            return
        admin = self._reverse_geocode_sample_admin(*pair)
        if not admin:
            return
        department, commune = admin
        if department and hasattr(self, "edit_sample_department"):
            self.edit_sample_department.setText(department)
        if commune and hasattr(self, "edit_sample_commune"):
            self.edit_sample_commune.setText(commune)

    def _normalize_sample_gps_fields(self) -> None:
        if not hasattr(self, "edit_sample_lat") or not hasattr(self, "edit_sample_lon"):
            return

        lat_txt = self.edit_sample_lat.text().strip()
        lon_txt = self.edit_sample_lon.text().strip()
        for txt in (lat_txt, lon_txt):
            pair_lat, pair_lon = self._split_gps_pair(txt)
            if pair_lat and pair_lon:
                self.edit_sample_lat.setText(pair_lat)
                self.edit_sample_lon.setText(pair_lon)
                self._try_autofill_sample_admin_from_gps()
                return

        self._normalize_coord_field(self.edit_sample_lat, -90.0, 90.0)
        self._normalize_coord_field(self.edit_sample_lon, -180.0, 180.0)
        self._try_autofill_sample_admin_from_gps()

    def _on_view_sample_map_clicked(self) -> None:
        self._normalize_sample_gps_fields()
        lat = self._parse_coord_value(self.edit_sample_lat.text() if hasattr(self, "edit_sample_lat") else "")
        lon = self._parse_coord_value(self.edit_sample_lon.text() if hasattr(self, "edit_sample_lon") else "")

        if lat is not None and not (-90 <= lat <= 90):
            lat = None
        if lon is not None and not (-180 <= lon <= 180):
            lon = None

        if QWebEngineView is None or QWebChannel is None:
            if lat is None or lon is None:
                QMessageBox.warning(
                    self,
                    "Carte interne indisponible",
                    "Le composant QtWebEngine n'est pas disponible, et aucune coordonnée valide n'est renseignée.",
                )
                return
            url = f"https://www.openstreetmap.org/?mlat={lat:.6f}&mlon={lon:.6f}#map=16/{lat:.6f}/{lon:.6f}"
            if not QDesktopServices.openUrl(QUrl(url)):
                QMessageBox.warning(self, "Carte indisponible", "Impossible d'ouvrir la carte dans le navigateur.")
            return

        dlg = MapPickerDialog(lat, lon, self)
        if dlg.exec() != QDialog.Accepted or not dlg.has_coordinates():
            return
        new_lat, new_lon = dlg.coordinates()
        self.edit_sample_lat.setText(f"{new_lat:.6f}")
        self.edit_sample_lon.setText(f"{new_lon:.6f}")
        self._try_autofill_sample_admin_from_gps()
        self._on_header_field_changed()

    def _build_manip_name(self) -> str:
        initials = "".join(
            self._person_initials(name) for name in self._sampler_names()
        )
        sample_date = getattr(self, "edit_sample_date", None)
        sample_time = getattr(self, "edit_sample_time", None)
        if not initials or not _field_has_value(sample_date) or not _field_has_value(sample_time):
            return ""
        dt = QDateTime(sample_date.date(), sample_time.time())
        return f"{initials}_{dt.toString('yyyyMMdd')}_{dt.toString('HH')}h{dt.toString('mm')}"

    def _build_titration_manip_name(self) -> str:
        oper = self._person_initials(self.edit_operator.text() if hasattr(self, "edit_operator") else "")
        coord = self._person_initials(self.edit_coordinator.text() if hasattr(self, "edit_coordinator") else "")
        titration_date = getattr(self, "edit_titration_date", None)
        titration_time = getattr(self, "edit_titration_time", None)
        if not oper or not coord or not _field_has_value(titration_date) or not _field_has_value(titration_time):
            return ""
        dt = QDateTime(titration_date.date(), titration_time.time())
        return f"{oper}{coord}_{dt.toString('yyyyMMdd')}_{dt.toString('HH')}h{dt.toString('mm')}"

    @staticmethod
    def _is_yes_value(value: str | None) -> bool:
        return str(value or "").strip().lower() in {"oui", "yes", "true", "1", "x"}

    def _set_titration_name_manual_mode(self, manual: bool) -> None:
        self._titration_name_manual = bool(manual)
        if not hasattr(self, "edit_titration_manip"):
            return
        self.edit_titration_manip.setReadOnly(not manual)
        self.edit_titration_manip.setFocusPolicy(Qt.StrongFocus if manual else Qt.NoFocus)
        if hasattr(self, "btn_toggle_titration_name_manual"):
            self.btn_toggle_titration_name_manual.setText("Auto" if manual else "Modifier")
            self.btn_toggle_titration_name_manual.setToolTip(
                "Revenir au nom généré automatiquement."
                if manual
                else "Saisir un nom manuel pour la titration."
            )
        if manual:
            self.edit_titration_manip.setStyleSheet(
                "background-color: rgba(92, 184, 92, 77); color: #ffffff; border: 1px solid #6fa8d6; "
                "font-weight: 700; padding: 3px 6px;"
            )
        else:
            self.edit_titration_manip.setStyleSheet(
                "background-color: #263238; color: #ffffff; border: 1px solid #6fa8d6; "
                "font-weight: 700; padding: 3px 6px;"
            )

    def _on_toggle_titration_name_manual(self) -> None:
        manual = not bool(getattr(self, "_titration_name_manual", False))
        self._set_titration_name_manual_mode(manual)
        if manual:
            if not self.edit_titration_manip.text().strip():
                self.edit_titration_manip.setText(self._build_titration_manip_name())
            self.edit_titration_manip.setFocus()
            self.edit_titration_manip.selectAll()
        else:
            self._refresh_titration_manip_name(force=True)
        self._on_titration_info_changed()

    def _on_titration_manual_name_text_changed(self, *args) -> None:
        if getattr(self, "_titration_name_manual", False):
            self._on_titration_info_changed()

    def _refresh_manip_name(self) -> None:
        if not hasattr(self, "edit_manip"):
            return
        name = self._build_manip_name()
        self.edit_manip.blockSignals(True)
        try:
            self.edit_manip.setText(name)
        finally:
            self.edit_manip.blockSignals(False)
        if name:
            self._on_manip_name_changed()

    def _refresh_titration_manip_name(self, force: bool = False) -> None:
        if not hasattr(self, "edit_titration_manip"):
            return
        if getattr(self, "_titration_name_manual", False) and not force:
            return
        name = self._build_titration_manip_name()
        self.edit_titration_manip.blockSignals(True)
        try:
            self.edit_titration_manip.setText(name)
        finally:
            self.edit_titration_manip.blockSignals(False)

    def _on_sample_components_changed(self, *args) -> None:
        self._refresh_manip_name()
        self._on_header_field_changed()

    def _on_titration_components_changed(self, *args) -> None:
        self._refresh_titration_manip_name()
        self._on_titration_info_changed()

    def _on_manip_components_changed(self, *args) -> None:
        self._on_sample_components_changed(*args)

    def _on_titration_info_changed(self, *args) -> None:
        self._refresh_fillable_fields_style()
        self._mark_metadata_dirty(refresh=False)
        if self.df_map is not None and isinstance(self.df_map, pd.DataFrame) and not self.df_map.empty:
            self.map_dirty = True
            # Mettre à jour les lignes d'en-tête (valeurs comme coordinateur, lieu…)
            self.df_map = self._apply_header_values_to_df_map(self.df_map)
            # Mettre à jour les noms de spectres avec le nom de la manip de titration.
            # edit_titration_manip est déjà à jour ici (auto : refreshed avant cet appel ;
            # manuel : en cours de frappe).
            titration_widget = getattr(self, "edit_titration_manip", None)
            titration_name = titration_widget.text().strip() if titration_widget else ""
            if titration_name:
                self._apply_spectrum_names_to_df_map(titration_name, int(self._spec_start_index))
        self._refresh_button_states()

    def _on_header_field_changed(self, *args) -> None:
        """Marque les métadonnées comme modifiées sans invalider la correspondance spectres ↔ tubes."""
        self._refresh_fillable_fields_style()
        self._mark_metadata_dirty(refresh=False)
        self._refresh_button_states()

    def _header_values(self, manip_name: str | None = None) -> dict[str, str]:
        name = manip_name if manip_name is not None else (
            self.edit_manip.text().strip() if hasattr(self, "edit_manip") else ""
        )
        titration_name = (
            self.edit_titration_manip.text().strip() if hasattr(self, "edit_titration_manip") else ""
        )
        titration_done = bool(self.chk_titration_done.isChecked()) if hasattr(self, "chk_titration_done") else False
        if not titration_done:
            titration_name = ""
        ammonium_enabled = bool(self._ammonium_test_enabled)
        ammonium = "Oui" if ammonium_enabled else "Non"
        bacterio_enabled = bool(self.chk_bacterio_analysis.isChecked()) if hasattr(self, "chk_bacterio_analysis") else False
        bacterio = "Oui" if bacterio_enabled else "Non"
        bacterio_day = ""
        bacterio_time = ""
        bacterio_company = ""
        bacterio_ecoli = ""
        bacterio_enterococci = ""
        if bacterio_enabled:
            bacterio_day = self._date_time_text(getattr(self, "edit_bacterio_deposit_day", None))
            bacterio_time = self._time_text(getattr(self, "edit_bacterio_deposit_time", None))
            bacterio_company = (
                self.edit_bacterio_company.text().strip() if hasattr(self, "edit_bacterio_company") else ""
            )
            bacterio_ecoli = (
                self.edit_bacterio_ecoli.text().strip() if hasattr(self, "edit_bacterio_ecoli") else ""
            )
            bacterio_enterococci = (
                self.edit_bacterio_enterococci.text().strip()
                if hasattr(self, "edit_bacterio_enterococci")
                else ""
            )
        tidal = self._is_tidal_water_type()
        tide_coefficient = ""
        if tidal and hasattr(self, "spin_tide_coefficient") and self.spin_tide_coefficient.value() > 0:
            tide_coefficient = str(self.spin_tide_coefficient.value())
        high_tide_time = ""
        if tidal and hasattr(self, "edit_high_tide_time"):
            high_tide_time = self._normalize_time_text(self.edit_high_tide_time.text())
        water_type = ""
        if hasattr(self, "combo_water_type") and self._is_water_type_selected():
            water_type = self.combo_water_type.currentText().strip()
        sampler_names = self._sampler_names()
        sampler_associations = self._sampler_associations()
        extra_sampler_values = {
            f"Préleveur·se {idx}": sampler_names[idx - 1] if len(sampler_names) >= idx else ""
            for idx in range(2, 7)
        }
        extra_association_values = {
            f"Association préleveur·se {idx}": sampler_associations[idx - 1] if len(sampler_associations) >= idx else ""
            for idx in range(2, 7)
        }
        debit_flux_values = {
            f"{self.DEBIT_FLUX_PREFIX}{column}": (
                str(self._debit_flux_values.get(column, "") or "").strip()
                if self._is_fresh_water_type()
                else ""
            )
            for column in self.DEBIT_FLUX_COLUMNS
        }
        return {
            "Nom du prélèvement": str(name or "").strip(),
            "Préleveur·se": self.edit_sampler.text().strip() if hasattr(self, "edit_sampler") else "",
            "Préleveur·ses": " ; ".join(sampler_names),
            **extra_sampler_values,
            "Association préleveur·se": (
                self.edit_sampler_association.text().strip()
                if hasattr(self, "edit_sampler_association")
                else ""
            ),
            **extra_association_values,
            "Association(s) du prélèvement": " ; ".join(a for a in sampler_associations if a),
            "Lieu du prélèvement": self.edit_sample_location.text().strip() if hasattr(self, "edit_sample_location") else "",
            "Latitude du prélèvement": self._format_coord_value(
                self.edit_sample_lat.text() if hasattr(self, "edit_sample_lat") else "",
                -90.0,
                90.0,
            ),
            "Longitude du prélèvement": self._format_coord_value(
                self.edit_sample_lon.text() if hasattr(self, "edit_sample_lon") else "",
                -180.0,
                180.0,
            ),
            "Département": self.edit_sample_department.text().strip() if hasattr(self, "edit_sample_department") else "",
            "Commune": self.edit_sample_commune.text().strip() if hasattr(self, "edit_sample_commune") else "",
            "Date du prélèvement": self._sample_date_text(),
            "Heure du prélèvement": self._sample_time_text(),
            "Date/heure du prélèvement": self._sample_date_time_text(),
            "Type d'eau": water_type,
            "Température de l'eau": (
                str(round(self.spin_water_temperature.value(), 1))
                if hasattr(self, "spin_water_temperature") and _field_has_value(self.spin_water_temperature)
                else ""
            ),
            **debit_flux_values,
            "Coefficient de marée": tide_coefficient,
            "Heure de pleine mer": high_tide_time,
            "Temps au moment de la mesure": (
                self.combo_weather.currentText()
                if hasattr(self, "combo_weather") and _field_has_value(self.combo_weather)
                else ""
            ),
            "Température de l'air": (
                str(round(self.spin_air_temperature.value(), 1))
                if hasattr(self, "spin_air_temperature") and _field_has_value(self.spin_air_temperature)
                else ""
            ),
            "Pluie dernières 24 h (mm)": (
                str(round(self.spin_rainfall_24h.value(), 1))
                if hasattr(self, "spin_rainfall_24h") and _field_has_value(self.spin_rainfall_24h)
                else ""
            ),
            "Commentaire sur les conditions": (
                self.edit_conditions_comment.toPlainText().strip()
                if hasattr(self, "edit_conditions_comment")
                else ""
            ),
            "Test ammonium réalisé": ammonium,
            "Test ammonium réalisé par": (
                self.edit_ammonium_operator.text().strip() if hasattr(self, "edit_ammonium_operator") else ""
            ) if ammonium_enabled else "",
            "Test ammonium 1 (grossier)": (
                self.edit_ammonium_test1.text().strip() if hasattr(self, "edit_ammonium_test1") else ""
            ) if ammonium_enabled else "",
            "Test ammonium 2 (précis)": (
                self.edit_ammonium_test2.text().strip() if hasattr(self, "edit_ammonium_test2") else ""
            ) if ammonium_enabled else "",
            "pH (ammonium)": (
                self.edit_ammonium_ph.text().strip() if hasattr(self, "edit_ammonium_ph") else ""
            ) if ammonium_enabled else "",
            "Turbidité mesurée": (
                "Oui" if hasattr(self, "chk_turbidity") and self.chk_turbidity.isChecked() else "Non"
            ),
            "Turbidité (NTU)": (
                str(round(self.spin_turbidity.value(), 1))
                if hasattr(self, "chk_turbidity") and self.chk_turbidity.isChecked()
                and hasattr(self, "spin_turbidity") and _field_has_value(self.spin_turbidity)
                else ""
            ),
            "Conductivité mesurée": (
                "Oui" if hasattr(self, "chk_conductivity") and self.chk_conductivity.isChecked() else "Non"
            ),
            "Conductivité (µS/cm)": (
                str(round(self.spin_conductivity.value(), 1))
                if hasattr(self, "chk_conductivity") and self.chk_conductivity.isChecked()
                and hasattr(self, "spin_conductivity") and _field_has_value(self.spin_conductivity)
                else ""
            ),
            "pH de l'eau mesuré": (
                "Oui" if hasattr(self, "chk_ph_water") and self.chk_ph_water.isChecked() else "Non"
            ),
            "pH de l'eau": (
                str(round(self.spin_ph_water.value(), 2))
                if hasattr(self, "chk_ph_water") and self.chk_ph_water.isChecked()
                and hasattr(self, "spin_ph_water") and _field_has_value(self.spin_ph_water)
                else ""
            ),
            "Oxygène dissous mesuré": (
                "Oui" if hasattr(self, "chk_dissolved_o2") and self.chk_dissolved_o2.isChecked() else "Non"
            ),
            "Oxygène dissous (mg/L)": (
                str(round(self.spin_dissolved_o2.value(), 2))
                if hasattr(self, "chk_dissolved_o2") and self.chk_dissolved_o2.isChecked()
                and hasattr(self, "spin_dissolved_o2") and _field_has_value(self.spin_dissolved_o2)
                else ""
            ),
            "Analyses bactériologiques": bacterio,
            "Jour de dépôt": bacterio_day,
            "Heure du dépôt": bacterio_time,
            "Entreprise de mesure": bacterio_company,
            "Résultats E.Coli (npp/ml)": bacterio_ecoli,
            "Résultats Entérocoques intestinaux (npp/100ml)": bacterio_enterococci,
            "Titration du cuivre": "Oui" if titration_done else "Non",
            "Nom de la manip de titration": str(titration_name or "").strip() if titration_done else "",
            "Nom de la manip de titration manuel": (
                "Oui" if titration_done and getattr(self, "_titration_name_manual", False) else "Non"
            ),
            "Date de la titration": self._titration_date_text() if titration_done else "",
            "Heure de la titration": self._titration_time_text() if titration_done else "",
            "Date/heure de la titration": self._titration_date_time_text() if titration_done else "",
            "Lieu de la titration": (
                self.edit_titration_location.text().strip()
                if titration_done and hasattr(self, "edit_titration_location")
                else ""
            ),
            "Coordinateur": (
                self.edit_coordinator.text().strip()
                if titration_done and hasattr(self, "edit_coordinator")
                else ""
            ),
            "Opérateur": (
                self.edit_operator.text().strip()
                if titration_done and hasattr(self, "edit_operator")
                else ""
            ),
        }

    @staticmethod
    def _header_storage_key(label: str) -> str:
        key = str(label or "").strip().rstrip(":").strip()
        aliases = {
            "Nom de la manip": "Nom du prélèvement",
            "Nom du prélèvement": "Nom du prélèvement",
            "Date": "Date/heure de la titration",
            "Lieu": "Lieu de la titration",
            "Préleveur/préleveuse": "Préleveur·se",
            "Préleveur": "Préleveur·se",
            "Préleveuse": "Préleveur·se",
            "Preleveur": "Préleveur·se",
            "Preleveuse": "Préleveur·se",
            "Préleveurs": "Préleveur·ses",
            "Préleveur·ses": "Préleveur·ses",
            "Preleveurs": "Préleveur·ses",
            "Préleveur 2": "Préleveur·se 2",
            "Préleveuse 2": "Préleveur·se 2",
            "Preleveur 2": "Préleveur·se 2",
            "Préleveur 3": "Préleveur·se 3",
            "Préleveuse 3": "Préleveur·se 3",
            "Preleveur 3": "Préleveur·se 3",
            "Association 1": "Association préleveur·se",
            "Association préleveur·se 1": "Association préleveur·se",
            "Association preleveur 1": "Association préleveur·se",
            "Association préleveur": "Association préleveur·se",
            "Association preleveur": "Association préleveur·se",
            "Association 2": "Association préleveur·se 2",
            "Association préleveur 2": "Association préleveur·se 2",
            "Association preleveur 2": "Association préleveur·se 2",
            "Association 3": "Association préleveur·se 3",
            "Association préleveur 3": "Association préleveur·se 3",
            "Association preleveur 3": "Association préleveur·se 3",
            "Association": "Association(s) du prélèvement",
            "Associations": "Association(s) du prélèvement",
            "Association(s)": "Association(s) du prélèvement",
            "Association prélèvement": "Association(s) du prélèvement",
            "Association prelevement": "Association(s) du prélèvement",
            "Associations prélèvement": "Association(s) du prélèvement",
            "Associations prelevement": "Association(s) du prélèvement",
            "Lat": "Latitude du prélèvement",
            "Latitude": "Latitude du prélèvement",
            "Latitude GPS": "Latitude du prélèvement",
            "Latitude prélèvement": "Latitude du prélèvement",
            "Latitude prelevement": "Latitude du prélèvement",
            "Lon": "Longitude du prélèvement",
            "Long": "Longitude du prélèvement",
            "Longitude": "Longitude du prélèvement",
            "Longitude GPS": "Longitude du prélèvement",
            "Longitude prélèvement": "Longitude du prélèvement",
            "Longitude prelevement": "Longitude du prélèvement",
            "Departement": "Département",
            "Département du prélèvement": "Département",
            "Departement du prelevement": "Département",
            "Commune du prélèvement": "Commune",
            "Commune du prelevement": "Commune",
            "Date prélèvement": "Date du prélèvement",
            "Date prelevement": "Date du prélèvement",
            "Date de prélèvement": "Date du prélèvement",
            "Date de prelevement": "Date du prélèvement",
            "Heure prélèvement": "Heure du prélèvement",
            "Heure prelevement": "Heure du prélèvement",
            "Heure de prélèvement": "Heure du prélèvement",
            "Heure de prelevement": "Heure du prélèvement",
            "Date titration": "Date de la titration",
            "Date de titration": "Date de la titration",
            "Heure titration": "Heure de la titration",
            "Heure de titration": "Heure de la titration",
            "Type eau": "Type d'eau",
            "Type d eau": "Type d'eau",
            "Type deau": "Type d'eau",
            "Type d'eau du prélèvement": "Type d'eau",
            "Type d'eau du prelevement": "Type d'eau",
            "Largeur route": "Débit / flux - Largeur route",
            "Longueur 6 arches": "Débit / flux - Longueur 6 arches",
            "Hauteur eau": "Débit / flux - Hauteur eau",
            "Volume total (m3)": "Débit / flux - Volume total (m3)",
            "Vitesse (s)": "Débit / flux - Vitesse (s)",
            "Débit (m3/s)": "Débit / flux - Débit (m3/s)",
            "Debit (m3/s)": "Débit / flux - Débit (m3/s)",
            "Débit/flux - Largeur route": "Débit / flux - Largeur route",
            "Débit/flux - Longueur 6 arches": "Débit / flux - Longueur 6 arches",
            "Débit/flux - Hauteur eau": "Débit / flux - Hauteur eau",
            "Débit/flux - Volume total (m3)": "Débit / flux - Volume total (m3)",
            "Débit/flux - Vitesse (s)": "Débit / flux - Vitesse (s)",
            "Débit/flux - Débit (m3/s)": "Débit / flux - Débit (m3/s)",
            "Debit / flux - Largeur route": "Débit / flux - Largeur route",
            "Debit / flux - Longueur 6 arches": "Débit / flux - Longueur 6 arches",
            "Debit / flux - Hauteur eau": "Débit / flux - Hauteur eau",
            "Debit / flux - Volume total (m3)": "Débit / flux - Volume total (m3)",
            "Debit / flux - Vitesse (s)": "Débit / flux - Vitesse (s)",
            "Debit / flux - Debit (m3/s)": "Débit / flux - Débit (m3/s)",
            "Coefficient maree": "Coefficient de marée",
            "Coef marée": "Coefficient de marée",
            "Coef maree": "Coefficient de marée",
            "Coefficient de maree": "Coefficient de marée",
            "Heure pleine mer": "Heure de pleine mer",
            "Pleine mer": "Heure de pleine mer",
            "Ammonium test 1": "Test ammonium 1 (grossier)",
            "Ammonium test 2": "Test ammonium 2 (précis)",
            "Test 1 ammonium": "Test ammonium 1 (grossier)",
            "Test 2 ammonium": "Test ammonium 2 (précis)",
            "Réalisateur test ammonium": "Test ammonium réalisé par",
            "Réalisatrice test ammonium": "Test ammonium réalisé par",
            "Test ammonium realise par": "Test ammonium réalisé par",
            "pH ammonium": "pH (ammonium)",
            "pH": "pH (ammonium)",
            "Turbidite": "Turbidité (NTU)",
            "Turbidité": "Turbidité (NTU)",
            "Turbidity": "Turbidité (NTU)",
            "Turbidité mesuree": "Turbidité mesurée",
            "Turbidite mesuree": "Turbidité mesurée",
            "Turbidité mesurée": "Turbidité mesurée",
            "Conductivite": "Conductivité (µS/cm)",
            "Conductivité": "Conductivité (µS/cm)",
            "Conductivity": "Conductivité (µS/cm)",
            "Conductivité mesuree": "Conductivité mesurée",
            "Conductivite mesuree": "Conductivité mesurée",
            "Conductivité mesurée": "Conductivité mesurée",
            "pH eau": "pH de l'eau",
            "pH de l eau": "pH de l'eau",
            "pH water": "pH de l'eau",
            "pH eau mesure": "pH de l'eau mesuré",
            "pH de l eau mesure": "pH de l'eau mesuré",
            "pH de l'eau mesuré": "pH de l'eau mesuré",
            "Oxygene dissous": "Oxygène dissous (mg/L)",
            "Oxygène dissous": "Oxygène dissous (mg/L)",
            "O2 dissous": "Oxygène dissous (mg/L)",
            "Dissolved O2": "Oxygène dissous (mg/L)",
            "Oxygene dissous mesure": "Oxygène dissous mesuré",
            "Oxygène dissous mesuré": "Oxygène dissous mesuré",
            "O2 dissous mesure": "Oxygène dissous mesuré",
            "Analyse bactériologique": "Analyses bactériologiques",
            "Analyses bacteriologiques": "Analyses bactériologiques",
            "Analyse bacteriologique": "Analyses bactériologiques",
            "Bactériologie": "Analyses bactériologiques",
            "Bacteriologie": "Analyses bactériologiques",
            "Jour depot": "Jour de dépôt",
            "Jour dépôt": "Jour de dépôt",
            "Jour de depot": "Jour de dépôt",
            "Heure depot": "Heure du dépôt",
            "Heure dépôt": "Heure du dépôt",
            "Heure du depot": "Heure du dépôt",
            "Entreprise mesure": "Entreprise de mesure",
            "Entreprise de mesures": "Entreprise de mesure",
            "Résultat E.Coli": "Résultats E.Coli (npp/ml)",
            "Résultats E.Coli": "Résultats E.Coli (npp/ml)",
            "Resultat E.Coli": "Résultats E.Coli (npp/ml)",
            "Resultats E.Coli": "Résultats E.Coli (npp/ml)",
            "E.Coli": "Résultats E.Coli (npp/ml)",
            "E. Coli": "Résultats E.Coli (npp/ml)",
            "Résultat Entérocoques intestinaux": "Résultats Entérocoques intestinaux (npp/100ml)",
            "Résultats Entérocoques intestinaux": "Résultats Entérocoques intestinaux (npp/100ml)",
            "Resultat Enterocoques intestinaux": "Résultats Entérocoques intestinaux (npp/100ml)",
            "Resultats Enterocoques intestinaux": "Résultats Entérocoques intestinaux (npp/100ml)",
            "Entérocoques intestinaux": "Résultats Entérocoques intestinaux (npp/100ml)",
            "Enterocoques intestinaux": "Résultats Entérocoques intestinaux (npp/100ml)",
            "Titration": "Titration du cuivre",
            "Titration réalisée": "Titration du cuivre",
            "Titration realisee": "Titration du cuivre",
            "Titration faite": "Titration du cuivre",
            "Nom de la titration": "Nom de la manip de titration",
            "Nom manip titration": "Nom de la manip de titration",
            "Nom titration manuel": "Nom de la manip de titration manuel",
            "Nom de la titration manuel": "Nom de la manip de titration manuel",
            "Nom manip titration manuel": "Nom de la manip de titration manuel",
            "Coordinateur·ice": "Coordinateur",
            "Opérateur·ice": "Opérateur",
            "Temperature de l'eau": "Température de l'eau",
            "Temp eau": "Température de l'eau",
            "Température eau": "Température de l'eau",
            "Température de l eau": "Température de l'eau",
            "Meteo": "Temps au moment de la mesure",
            "Météo": "Temps au moment de la mesure",
            "Temps mesure": "Temps au moment de la mesure",
            "Temps au moment": "Temps au moment de la mesure",
            "Temperature de l'air": "Température de l'air",
            "Temp air": "Température de l'air",
            "Température air": "Température de l'air",
            "Température de l air": "Température de l'air",
            "Pluie 24h": "Pluie dernières 24 h (mm)",
            "Pluie dernieres 24h": "Pluie dernières 24 h (mm)",
            "Pluie dernières 24h": "Pluie dernières 24 h (mm)",
            "Rainfall 24h": "Pluie dernières 24 h (mm)",
            "Pluie 24 h": "Pluie dernières 24 h (mm)",
            "Conditions commentaire": "Commentaire sur les conditions",
            "Commentaire conditions": "Commentaire sur les conditions",
            "Observations conditions": "Commentaire sur les conditions",
        }
        return aliases.get(key, key)

    def _header_rows_dataframe(self, manip_name: str | None = None) -> pd.DataFrame:
        values = self._header_values(manip_name)
        rows = [
            ["Nom du prélèvement", values.get("Nom du prélèvement", "")],
            ["Nom de la manip", values.get("Nom de la manip de titration", "")],
            ["Coordinateur", values.get("Coordinateur", "")],
            ["Opérateur", values.get("Opérateur", "")],
            ["Date de la titration", values.get("Date de la titration", "")],
            ["Heure de la titration", values.get("Heure de la titration", "")],
            ["Lieu de la titration", values.get("Lieu de la titration", "")],
            ["Date/heure de la titration", values.get("Date/heure de la titration", "")],
        ]
        return pd.DataFrame(rows, columns=["Nom du spectre", "Tube"])

    def _metadata_index_dataframe(self) -> pd.DataFrame:
        values = self._header_values()
        rows = [
            {
                "Clé": self._header_storage_key(label),
                "Champ": label,
                "Valeur": value,
            }
            for label, value in values.items()
        ]
        return pd.DataFrame(rows, columns=["Clé", "Champ", "Valeur"])

    def _df_map_with_current_header(self, df: pd.DataFrame) -> pd.DataFrame:
        if df is None or df.empty or not {"Nom du spectre", "Tube"}.issubset(df.columns):
            return df
        col_ns = df["Nom du spectre"].astype(str).str.strip()
        col_tube = df["Tube"].astype(str).str.strip()
        hdr_mask = (col_ns == "Nom du spectre") & (col_tube == "Tube")
        mapping_rows = (
            df.loc[df.index > hdr_mask[hdr_mask].index[-1], ["Nom du spectre", "Tube"]].copy()
            if hdr_mask.any()
            else df[["Nom du spectre", "Tube"]].copy()
        )
        blank_row = pd.DataFrame([["", ""]], columns=["Nom du spectre", "Tube"])
        internal_header = pd.DataFrame([["Nom du spectre", "Tube"]], columns=["Nom du spectre", "Tube"])
        return pd.concat(
            [self._header_rows_dataframe(), blank_row, internal_header, mapping_rows.reset_index(drop=True)],
            ignore_index=True,
        )

    @staticmethod
    def _clean_loaded_volume_dataframe(df: pd.DataFrame) -> pd.DataFrame:
        if not isinstance(df, pd.DataFrame):
            return pd.DataFrame()
        out = df.dropna(how="all").copy()
        if out.empty:
            return out
        unnamed = out.columns.astype(str).str.match(r"^Unnamed")
        if unnamed.any():
            out = out.loc[:, ~unnamed]
        return out

    def _header_values_from_metadata_index_sheet(self, xls: pd.ExcelFile) -> dict[str, str]:
        """Relit les métadonnées depuis la feuille technique clé/valeur."""
        if SHEET_METADATA_INDEX not in xls.sheet_names:
            return {}
        try:
            df = pd.read_excel(xls, sheet_name=SHEET_METADATA_INDEX, dtype=object)
        except Exception:
            return {}
        if not isinstance(df, pd.DataFrame) or df.empty:
            return {}

        def _clean_cell(value) -> str:
            if value is None:
                return ""
            try:
                if pd.isna(value):
                    return ""
            except Exception:
                pass
            if isinstance(value, pd.Timestamp):
                if value.time().hour == 0 and value.time().minute == 0 and value.time().second == 0:
                    return value.strftime("%d/%m/%Y")
                return value.strftime("%d/%m/%Y %H:%M")
            text = str(value).strip()
            return "" if text.lower() in {"nan", "nat", "none", "<na>"} else text

        normalized_columns = {self._header_storage_key(col): col for col in df.columns}
        label_col = (
            normalized_columns.get("Champ")
            or normalized_columns.get("Libellé")
            or normalized_columns.get("Label")
            or normalized_columns.get("Clé")
        )
        value_col = normalized_columns.get("Valeur") or normalized_columns.get("Value") or normalized_columns.get("Tube")
        if not label_col or not value_col:
            return {}

        values: dict[str, str] = {}
        for _, row in df.iterrows():
            label = _clean_cell(row.get(label_col, ""))
            value = _clean_cell(row.get(value_col, ""))
            if label and value:
                values[self._header_storage_key(label)] = value
        return values

    def _header_values_from_measure_summary_sheet(self, xls: pd.ExcelFile) -> dict[str, str]:
        """Relit les champs principaux depuis la feuille terrain lisible.

        Cette feuille n'est pas le stockage principal, mais elle permet de
        récupérer les fichiers enregistrés pendant la période où l'en-tête
        rechargeable de la feuille Correspondance était incomplet.
        """
        if SHEET_MEASURES not in xls.sheet_names:
            return {}
        try:
            raw = pd.read_excel(xls, sheet_name=SHEET_MEASURES, header=None, dtype=object)
        except Exception:
            return {}

        def _cell(row: int, col: int) -> str:
            try:
                value = raw.iat[row - 1, col - 1]
            except Exception:
                return ""
            if value is None:
                return ""
            try:
                if pd.isna(value):
                    return ""
            except Exception:
                pass
            if isinstance(value, pd.Timestamp):
                if value.time().hour == 0 and value.time().minute == 0 and value.time().second == 0:
                    return value.strftime("%d/%m/%Y")
                return value.strftime("%d/%m/%Y %H:%M")
            text = str(value).strip()
            return "" if text.lower() in {"nan", "nat", "none", "<na>"} else text

        def _key(value) -> str:
            text = mm.norm_text_key(value)
            text = text.replace("°", " ").replace("·", " ").replace("/", " ")
            text = re.sub(r"[^a-z0-9]+", " ", text)
            return re.sub(r"\s+", " ", text).strip()

        known_label_keys = {
            _key(label)
            for label in (
                "Nom du prélèvement",
                "Type d'eau",
                "Département",
                "Commune",
                "Lieu du prélèvement",
                "Nom du lieu",
                "Latitude",
                "Latitude GPS",
                "Longitude",
                "Longitude GPS",
                "Lieu de la titration",
                "Lieu titration",
                "Date de la titration",
                "Date titration",
                "Heure de la titration",
                "Heure titration",
                "Date du prélèvement",
                "Heure du prélèvement",
                "Préleveur·se",
                "Association",
                "Titration du cuivre",
                "Analyses bactériologiques",
                "Test ammonium réalisé",
                "Turbidité mesurée",
                "Conductivité mesurée",
                "pH eau mesuré",
                "pH de l'eau mesuré",
                "Oxygène dissous mesuré",
                "Coefficient de marée",
                "Heure pleine mer",
                "Heure de pleine mer",
                "Réalisé par",
                "Test 1 grossier",
                "Test 2 précis",
                "pH ammonium",
                "Jour de dépôt",
                "Heure du dépôt",
                "Entreprise de mesure",
                "Météo",
                "T° eau (°C)",
                "T° air (°C)",
                "Pluie 24 h (mm)",
                "E. coli (npp/ml)",
                "Entérocoques (npp/100ml)",
                "Turbidité (NTU)",
                "Conductivité (µS/cm)",
                "pH de l'eau",
                "Oxygène dissous (mg/L)",
                "Nom manip titration",
                "Largeur route",
                "Longueur 6 arches",
                "Hauteur eau",
                "Volume total (m3)",
                "Vitesse (s)",
                "Débit (m3/s)",
                "Commentaires",
            )
        }

        def _cell0(row_idx: int, col_idx: int) -> str:
            return _cell(row_idx + 1, col_idx + 1)

        def _value_right_of_label(*labels: str, max_scan: int = 12) -> str:
            wanted = {_key(label) for label in labels}
            for row_idx in range(min(len(raw), 40)):
                for col_idx in range(raw.shape[1]):
                    if _key(_cell0(row_idx, col_idx)).rstrip(" :") not in wanted:
                        continue
                    for scan_col in range(col_idx + 1, min(raw.shape[1], col_idx + 1 + max_scan)):
                        val = _cell0(row_idx, scan_col)
                        if not val:
                            continue
                        val_key = _key(val).rstrip(" :")
                        if val_key in known_label_keys and val_key not in wanted:
                            break
                        return val
            return ""

        def _values_right_of_label_prefix(*prefixes: str) -> list[str]:
            prefix_keys = tuple(_key(prefix) for prefix in prefixes)
            values_out: list[str] = []
            for row_idx in range(min(len(raw), 40)):
                for col_idx in range(raw.shape[1] - 1):
                    label_key = _key(_cell0(row_idx, col_idx))
                    if not label_key:
                        continue
                    if any(label_key == prefix or label_key.startswith(f"{prefix} ") for prefix in prefix_keys):
                        val = _cell0(row_idx, col_idx + 1)
                        if val:
                            values_out.append(val)
            return values_out

        def _find_measure_header_rows() -> tuple[int, int, int]:
            for row_idx in range(max(0, len(raw) - 1)):
                row_keys = [_key(_cell0(row_idx, col_idx)) for col_idx in range(raw.shape[1])]
                next_keys = [_key(_cell0(row_idx + 1, col_idx)) for col_idx in range(raw.shape[1])]
                current_joined = " ".join(row_keys)
                joined = " ".join(row_keys + next_keys)
                new_group_score = sum(
                    marker in current_joined
                    for marker in (
                        "contexte terrain",
                        "test ammonium",
                        "analyses bacteriologiques",
                        "autres mesures",
                        "titration du cuivre",
                        "debit flux",
                    )
                )
                new_sub_score = sum(
                    marker in " ".join(next_keys)
                    for marker in (
                        "meteo",
                        "t eau",
                        "test 1 grossier",
                        "jour de depot",
                        "ph eau mesure",
                        "largeur route",
                        "commentaires",
                    )
                )
                legacy_current_score = sum(
                    marker in current_joined
                    for marker in (
                        "point gps",
                        "nature de l eau",
                        "heure du prelevement",
                        "analyse e coli",
                        "resultats e coli",
                        "debit flux",
                    )
                )
                legacy_score = sum(
                    marker in joined
                    for marker in (
                        "point gps",
                        "nature de l eau",
                        "heure du prelevement",
                        "analyse e coli",
                        "resultats e coli",
                        "debit flux",
                    )
                )
                is_new_table = new_group_score >= 3 and new_sub_score >= 2
                is_legacy_table = legacy_current_score >= 2 and (
                    legacy_score >= 3 or ("lat" in next_keys and "lon" in next_keys and "point gps" in joined)
                )
                if not (is_new_table or is_legacy_table):
                    continue
                data_row = row_idx + 2
                for candidate in range(row_idx + 2, min(len(raw), row_idx + 12)):
                    filled = sum(1 for col_idx in range(raw.shape[1]) if _cell0(candidate, col_idx))
                    if filled >= 2:
                        data_row = candidate
                        break
                return row_idx, row_idx + 1, data_row
            return 4, 5, 6

        top_header_row, sub_header_row, data_row = _find_measure_header_rows()

        headers: list[str] = []
        for col_idx in range(raw.shape[1]):
            top = _cell0(top_header_row, col_idx)
            sub = _cell0(sub_header_row, col_idx)
            headers.append(_key(f"{top} {sub}".strip()))

        def _header_matches(header: str, *patterns: str) -> bool:
            if not header:
                return False
            for pattern in patterns:
                pat = _key(pattern)
                if not pat:
                    continue
                if pat == "association":
                    if header == pat or header.startswith(f"{pat} "):
                        return True
                    continue
                if header == pat or pat in header:
                    return True
            return False

        def _col_value(*patterns: str, exclude: tuple[str, ...] = (), fallback: str = "") -> str:
            exclude_keys = tuple(_key(item) for item in exclude)
            for col_idx, header in enumerate(headers):
                if any(ex and ex in header for ex in exclude_keys):
                    continue
                if _header_matches(header, *patterns):
                    value = _cell0(data_row, col_idx)
                    if value:
                        return value
            return fallback

        def _col_values(*patterns: str, fallback_values: list[str] | None = None) -> list[str]:
            values_out: list[str] = []
            for col_idx, header in enumerate(headers):
                if _header_matches(header, *patterns):
                    val = _cell0(data_row, col_idx)
                    if val:
                        values_out.append(val)
            if values_out:
                return values_out
            return fallback_values or []

        sampler_values = _col_values(
            "préleveur",
            "preleveur",
            fallback_values=[_cell(7, 10), _cell(7, 12), _cell(7, 14)],
        )
        association_values = _col_values(
            "association",
            fallback_values=[_cell(7, 11), _cell(7, 13), _cell(7, 15)],
        )
        if not any(sampler_values):
            sampler_values = _values_right_of_label_prefix("Préleveur·se", "Preleveur")
        if not any(sampler_values):
            sampler_values = [p.strip() for p in re.split(r"[;\n]+", _value_right_of_label("Bénévoles engagé·es")) if p.strip()]
        if not any(association_values):
            association_values = _values_right_of_label_prefix("Association")
        if not any(association_values):
            association_values = [
                p.strip()
                for p in re.split(r"[;\n]+", _value_right_of_label("De (ou des) l'association(s)"))
                if p.strip()
            ]

        bacterio_markers = [
            _col_value("analyse e coli", fallback=_cell(7, 19)),
            _col_value("analyse enterocoques", fallback=_cell(7, 20)),
            _col_value("resultats e coli", fallback=_cell(7, 25)),
            _col_value("resultats enterocoques", fallback=_cell(7, 26)),
        ]
        bacterio_status = _col_value("analyses bacteriologiques", "analyse bacteriologique")
        bacterio_done = any(str(v).strip() for v in bacterio_markers)

        values: dict[str, str] = {
            "Nom du prélèvement": _value_right_of_label("Nom du prélèvement"),
            "Département": _value_right_of_label("Département") or _cell(2, 4),
            "Commune": _value_right_of_label("Commune") or _cell(2, 9),
            "Lieu du prélèvement": _value_right_of_label("Lieu du prélèvement", "Nom du lieu")
            or _col_value("point", exclude=("point gps",), fallback=_cell(7, 2))
            or _cell(2, 19),
            "Latitude du prélèvement": _value_right_of_label("Latitude du prélèvement", "Latitude GPS", "Latitude", "Lat")
            or _col_value("point gps lat", "latitude", "lat", fallback=_cell(7, 3)),
            "Longitude du prélèvement": _value_right_of_label(
                "Longitude du prélèvement",
                "Longitude GPS",
                "Longitude",
                "Lon",
            )
            or _col_value("point gps lon", "longitude", "lon", fallback=_cell(7, 4)),
            "Type d'eau": _value_right_of_label("Type d'eau", "Type de lieu", "Nature de l'eau")
            or _col_value("nature de l eau", "type d eau", fallback=_cell(7, 5))
            or _cell(2, 14),
            "Date du prélèvement": _value_right_of_label("Date du prélèvement", "Date de prélèvement")
            or _col_value("date", "date prelevement", fallback=_cell(7, 6)),
            "Coefficient de marée": _value_right_of_label("Coefficient de marée", "Coefficient marée")
            or _col_value("coefficient de maree", "coefficient maree", fallback=_cell(7, 7)),
            "Heure de pleine mer": _value_right_of_label("Heure de pleine mer", "Heure pleine mer")
            or _col_value("heure pleine mer", "heure de pleine mer", fallback=_cell(7, 8)),
            "Heure du prélèvement": _value_right_of_label("Heure du prélèvement", "Heure de prélèvement")
            or _col_value("heure du prelevement", fallback=_cell(7, 9)),
            "Test ammonium réalisé": _value_right_of_label("Test ammonium réalisé", "Test ammonium")
            or _col_value("test ammonium realise", "test ammonium"),
            "Test ammonium réalisé par": _value_right_of_label("Test ammonium réalisé par")
            or _col_value("realise par", "réalisé par"),
            "Test ammonium 1 (grossier)": _col_value("test 1 grossier", "ammonium test 1"),
            "Test ammonium 2 (précis)": _col_value("test 2 precis", "ammonium test 2"),
            "pH (ammonium)": _col_value("ph ammonium"),
            "Analyses bactériologiques": _value_right_of_label(
                "Analyses bactériologiques",
                "Analyse bactériologique",
            )
            or bacterio_status
            or ("Oui" if bacterio_done else ""),
            "Jour de dépôt": _col_value("jour de depot", "jour depot", fallback=_cell(7, 16)),
            "Heure du dépôt": _col_value("heure du depot", "heure depot", fallback=_cell(7, 17)),
            "Entreprise de mesure": _col_value("entreprise de mesure", fallback=_cell(7, 18)),
            "Temps au moment de la mesure": _col_value("meteo", "contexte de prelevement", fallback=_cell(7, 21)),
            "Température de l'eau": _col_value("t eau", "temperature de l eau", fallback=_cell(7, 22)),
            "Température de l'air": _col_value("t air", "temperature de l air", fallback=_cell(7, 23)),
            "Pluie dernières 24 h (mm)": _col_value("pluviometrie", "mm 24h", "pluie", fallback=_cell(7, 24)),
            "Résultats E.Coli (npp/ml)": _col_value("resultats e coli", "e coli", fallback=_cell(7, 25)),
            "Résultats Entérocoques intestinaux (npp/100ml)": _col_value(
                "resultats enterocoques",
                "enterocoques",
                fallback=_cell(7, 26),
            ),
            "Turbidité mesurée": _value_right_of_label("Turbidité mesurée")
            or _col_value("turbidite mesuree"),
            "Turbidité (NTU)": _col_value("turbidite ntu"),
            "Conductivité mesurée": _value_right_of_label("Conductivité mesurée")
            or _col_value("conductivite mesuree"),
            "Conductivité (µS/cm)": _col_value("conductivite us cm", "conductivite s cm"),
            "pH de l'eau mesuré": _value_right_of_label("pH eau mesuré", "pH de l'eau mesuré")
            or _col_value("ph eau mesure", "ph de l eau mesure"),
            "pH de l'eau": _col_value("ph de l eau", "ph eau", exclude=("mesure",)),
            "Oxygène dissous mesuré": _value_right_of_label("Oxygène dissous mesuré", "O2 dissous mesuré")
            or _col_value("oxygene dissous mesure", "o2 dissous mesure"),
            "Oxygène dissous (mg/L)": _col_value("oxygene dissous mg l", "o2 dissous"),
            "Titration du cuivre": _value_right_of_label("Titration du cuivre", "Titration")
            or _col_value("titration du cuivre", "titration"),
            "Nom de la manip de titration": _value_right_of_label("Nom de la manip", "Nom de la manip de titration")
            or _col_value(
                "nom manip titration",
                "nom de la manip de titration",
                "nom de la manip",
            ),
            "Date de la titration": _value_right_of_label("Date de la titration", "Date titration"),
            "Heure de la titration": _value_right_of_label("Heure de la titration", "Heure titration"),
            "Lieu de la titration": _value_right_of_label("Lieu de la titration", "Lieu titration")
            or _col_value("lieu titration", "lieu de la titration"),
            "Débit / flux - Largeur route": _col_value("largeur route", fallback=_cell(7, 27)),
            "Débit / flux - Longueur 6 arches": _col_value("longueur 6 arches", fallback=_cell(7, 28)),
            "Débit / flux - Hauteur eau": _col_value("hauteur eau", fallback=_cell(7, 29)),
            "Débit / flux - Volume total (m3)": _col_value("volume total", fallback=_cell(7, 30)),
            "Débit / flux - Vitesse (s)": _col_value("vitesse", fallback=_cell(7, 31)),
            "Débit / flux - Débit (m3/s)": _col_value("debit m3 s", fallback=_cell(7, 32)),
            "Commentaire sur les conditions": _col_value("commentaires", "commentaire", fallback=_cell(7, 33)),
        }

        sampler_values = [v for v in sampler_values if str(v).strip()]
        association_values = [v for v in association_values if str(v).strip()]
        if sampler_values:
            values["Préleveur·se"] = sampler_values[0]
            values["Préleveur·ses"] = " ; ".join(sampler_values)
            for idx, sampler in enumerate(sampler_values[1:6], start=2):
                values[f"Préleveur·se {idx}"] = sampler
        if association_values:
            values["Association préleveur·se"] = association_values[0]
            values["Association(s) du prélèvement"] = " ; ".join(association_values)
            for idx, association in enumerate(association_values[1:6], start=2):
                values[f"Association préleveur·se {idx}"] = association

        return {label: value for label, value in values.items() if str(value or "").strip()}

    def _merge_header_values_into_df_map(self, df: pd.DataFrame, values: dict[str, str]) -> pd.DataFrame:
        if not values:
            return df
        default_cols = ["Nom du spectre", "Tube"]
        clean_values = {
            self._header_storage_key(label): (label, str(value).strip())
            for label, value in values.items()
            if str(value or "").strip()
        }
        if not clean_values:
            return df

        def _clean_cell(value) -> str:
            if value is None:
                return ""
            try:
                if pd.isna(value):
                    return ""
            except Exception:
                pass
            text = str(value).strip()
            return "" if text.lower() in {"nan", "nat", "none", "<na>"} else text

        if df is None or df.empty or not set(default_cols).issubset(df.columns):
            rows = [[label, value] for label, value in values.items() if str(value or "").strip()]
            return pd.DataFrame(rows, columns=default_cols)

        col_ns = df["Nom du spectre"].map(_clean_cell)
        col_tube = df["Tube"].map(_clean_cell)
        hdr_mask = (col_ns == "Nom du spectre") & (col_tube == "Tube")
        if hdr_mask.any():
            first_hdr_idx = hdr_mask[hdr_mask].index[0]
            header_rows = df.loc[df.index < first_hdr_idx, default_cols].copy()
            mapping_rows = df.loc[df.index > first_hdr_idx, default_cols].copy()
        else:
            header_rows = df[default_cols].copy()
            mapping_rows = pd.DataFrame(columns=default_cols)

        merged_rows: list[list[str]] = []
        seen_keys: set[str] = set()
        for _, row in header_rows.iterrows():
            label = _clean_cell(row.get("Nom du spectre", ""))
            current_value = _clean_cell(row.get("Tube", ""))
            if not label:
                continue
            key = self._header_storage_key(label)
            if key in clean_values and not current_value:
                current_value = clean_values[key][1]
            merged_rows.append([label, current_value])
            seen_keys.add(key)

        for key, (label, value) in clean_values.items():
            if key not in seen_keys:
                merged_rows.append([label, value])

        merged_header = pd.DataFrame(merged_rows, columns=default_cols)
        if hdr_mask.any() or not mapping_rows.empty:
            blank_row = pd.DataFrame([["", ""]], columns=default_cols)
            internal_header = pd.DataFrame([["Nom du spectre", "Tube"]], columns=default_cols)
            return pd.concat(
                [merged_header, blank_row, internal_header, mapping_rows.reset_index(drop=True)],
                ignore_index=True,
            )
        return merged_header

    def _apply_header_values_to_df_map(self, df: pd.DataFrame, manip_name: str | None = None) -> pd.DataFrame:
        if df is None or df.empty or not {"Nom du spectre", "Tube"}.issubset(df.columns):
            return df
        values = self._header_values(manip_name)
        out = df.copy()
        for idx in out.index:
            label = str(out.at[idx, "Nom du spectre"]).strip()
            storage_key = self._header_storage_key(label)
            if storage_key in values:
                out.at[idx, "Tube"] = values[storage_key]
        return out

    def _on_ammonium_test_toggled(self, checked: bool) -> None:
        self._ammonium_test_enabled = bool(checked)
        for attr in (
            "lbl_ammonium_operator", "edit_ammonium_operator",
            "lbl_ammonium_test1", "edit_ammonium_test1",
            "lbl_ammonium_test2", "edit_ammonium_test2",
            "lbl_ammonium_ph", "edit_ammonium_ph",
        ):
            w = getattr(self, attr, None)
            if w is not None:
                w.setVisible(bool(checked))
        self._on_header_field_changed()

    def _on_turbidity_toggled(self, checked: bool) -> None:
        for attr in ("lbl_turbidity_val", "spin_turbidity"):
            w = getattr(self, attr, None)
            if w is not None:
                w.setVisible(bool(checked))
        self._on_header_field_changed()

    def _on_conductivity_toggled(self, checked: bool) -> None:
        for attr in ("lbl_conductivity_val", "spin_conductivity"):
            w = getattr(self, attr, None)
            if w is not None:
                w.setVisible(bool(checked))
        self._on_header_field_changed()

    def _on_ph_water_toggled(self, checked: bool) -> None:
        for attr in ("lbl_ph_water_val", "spin_ph_water"):
            w = getattr(self, attr, None)
            if w is not None:
                w.setVisible(bool(checked))
        self._on_header_field_changed()

    def _on_dissolved_o2_toggled(self, checked: bool) -> None:
        for attr in ("lbl_dissolved_o2_val", "spin_dissolved_o2"):
            w = getattr(self, attr, None)
            if w is not None:
                w.setVisible(bool(checked))
        self._on_header_field_changed()


    def _on_debit_flux_clicked(self) -> None:
        dlg = DebitFluxDialog(self.DEBIT_FLUX_COLUMNS, self._debit_flux_values, self)
        if dlg.exec() != QDialog.Accepted:
            return
        self._debit_flux_values = dlg.values()
        self._on_header_field_changed()

    @staticmethod
    def _to_float(x):
        return mm.to_float(x)

    @staticmethod
    def _normalize_tube_label(label: str) -> str:
        return mm.normalize_tube_label(label)
    
    # ------------------------------------------------------------------
    # Gestion du tableau des volumes (avec concentrations)
    # ------------------------------------------------------------------
    def _on_edit_map_clicked(self) -> None:
        """
        Ouvre le dialog d'édition pour la correspondance spectres ↔ tubes.

        Le modèle par défaut suit la structure Excel d'origine avec :
        - les lignes d'en-tête (nom du prélèvement, date, lieu, coordinateur, opérateur),
        - une ligne vide,
        - une ligne d'en-tête interne "Nom du spectre" / "Tube",
        - puis les lignes de correspondance NomSpectre ↔ Tube.

        Les noms de spectres sont automatiquement générés à partir du nom de la manip
        sous la forme NomManip_00, NomManip_01, … (indice à 2 chiffres) lors de la
        première création du tableau.
        """
        default_cols = ["Nom du spectre", "Tube"]

        # 1) Récupération du nom de base pour les spectres.
        # Priorité : nom de la manip de titration (car les spectres sont mesurés
        # pendant la titration). Fallback : nom du prélèvement.
        try:
            titration_widget = getattr(self, "edit_titration_manip", None)
            titration_name = titration_widget.text().strip() if titration_widget else ""
            titration_done = (
                hasattr(self, "chk_titration_done") and self.chk_titration_done.isChecked()
            )
            if titration_done and titration_name:
                manip_name = titration_name
            else:
                manip_name = self.edit_manip.text().strip() if hasattr(self, "edit_manip") else ""
        except Exception:
            manip_name = ""

        if not manip_name:
            QMessageBox.warning(
                self,
                "Nom de la manip manquant",
                "Veuillez renseigner le préleveur·se et la date/heure de prélèvement, "
                "ou le coordinateur·ice et la date/heure de titration, "
                "avant de créer la correspondance spectres ↔ tubes.",
            )
            return

        # Si df_map existe, on met à jour les lignes d'en-tête pour refléter l'UI
        if self.df_map is not None and not self.df_map.empty:
            try:
                self.df_map = self._apply_header_values_to_df_map(self.df_map, manip_name)
            except Exception:
                pass

        def _extract_mapping_rows(existing: pd.DataFrame | None) -> pd.DataFrame | None:
            if existing is None or not isinstance(existing, pd.DataFrame) or existing.empty:
                return None
            if not {"Nom du spectre", "Tube"}.issubset(existing.columns):
                return None
            hdr_idx = None
            for idx, row in existing.iterrows():
                left = str(row.get("Nom du spectre", "")).strip()
                right = str(row.get("Tube", "")).strip()
                if left == "Nom du spectre" and right == "Tube":
                    hdr_idx = idx
                    break
            if hdr_idx is not None:
                return existing.loc[hdr_idx + 1 :].reset_index(drop=True)
            return existing.reset_index(drop=True)

        def _is_brb_label(label: str) -> bool:
            return "brb" in self._norm_text_key(label)

        def _build_initial_df(
            include_brb: bool,
            include_control: bool,
            existing_df: pd.DataFrame | None,
        ) -> tuple[pd.DataFrame, pd.DataFrame]:
            # 2) Lignes d'en-tête construites à partir de l'UI (toujours recalculées)
            df_header = self._header_rows_dataframe(manip_name)
            blank_row = pd.DataFrame([["", ""]], columns=default_cols)
            internal_header = pd.DataFrame([["Nom du spectre", "Tube"]], columns=default_cols)

            # Nombre de tubes : on s'aligne sur le tableau des volumes si disponible
            n_tubes = self._infer_n_tubes_for_mapping()
            tube_labels_default = self._default_tube_labels_for_mapping(
                n_tubes,
                include_brb_override=include_brb,
                include_control_override=include_control,
            )

            start_index = int(self._spec_start_index)

            def _make_mapping_rows(labels: list[str]) -> pd.DataFrame:
                names = [f"{manip_name}_{i:02d}" for i in range(start_index, start_index + len(labels))]
                return pd.DataFrame({"Nom du spectre": names, "Tube": labels}, columns=default_cols)

            # --- DataFrame RESET (valeurs par défaut) ---
            mapping_rows_default = _make_mapping_rows(tube_labels_default)
            reset_df = pd.concat(
                [df_header, blank_row, internal_header, mapping_rows_default],
                ignore_index=True,
            )

            # 3) Récupération de la partie "correspondance" existante si elle a déjà été éditée
            mapping_rows = _extract_mapping_rows(existing_df)

            # Si aucune correspondance existante, on génère un brouillon par défaut
            if mapping_rows is None:
                mapping_rows = _make_mapping_rows(tube_labels_default)
            else:
                # Si le nombre de tubes a changé (ex : 15), on complète automatiquement
                # pour que l'utilisateur voie toutes les lignes nécessaires.
                def _clean_text(v) -> str:
                    if v is None or (isinstance(v, float) and pd.isna(v)) or pd.isna(v):
                        return ""
                    s = str(v).strip()
                    return "" if s == "<NA>" else s

                existing_by_tube: dict[str, dict[str, str]] = {}
                extra_rows: list[dict] = []
                desired_keys = {self._tube_merge_key_for_mapping(t, n_tubes) for t in tube_labels_default}

                for _, row in mapping_rows.iterrows():
                    tube_raw = _clean_text(row.get("Tube", ""))
                    name_raw = _clean_text(row.get("Nom du spectre", ""))
                    if (not include_brb) and _is_brb_label(tube_raw):
                        continue
                    tube_key = self._tube_merge_key_for_mapping(tube_raw, n_tubes)
                    if not tube_key:
                        continue
                    if tube_key in desired_keys:
                        # Conserver la première ligne rencontrée pour ce tube (sans écraser les valeurs)
                        if tube_key not in existing_by_tube:
                            existing_by_tube[tube_key] = {"Nom du spectre": name_raw, "Tube": tube_raw}
                        else:
                            # Compléter le nom si la première occurrence était vide
                            if name_raw and not existing_by_tube[tube_key].get("Nom du spectre"):
                                existing_by_tube[tube_key]["Nom du spectre"] = name_raw
                    else:
                        # Conserver les lignes hors "tubes par défaut" (ne pas les perdre)
                        if tube_raw or name_raw:
                            extra_rows.append({"Nom du spectre": name_raw, "Tube": tube_raw})

                aligned_rows: list[dict] = []
                for i, tube_label in enumerate(tube_labels_default):
                    tube_key = self._tube_merge_key_for_mapping(tube_label, n_tubes)
                    existing = existing_by_tube.get(tube_key, {})
                    name = f"{manip_name}_{start_index + i:02d}"
                    existing_tube = existing.get("Tube") or ""
                    # Forcer l'affichage du dernier tube comme "Contrôle" si demandé,
                    # et inversement forcer "Tube N" si le contrôle n'est pas demandé.
                    if self._is_control_label(tube_label):
                        tube_val = "Contrôle"
                    else:
                        tube_val = tube_label if self._is_control_label(existing_tube) else (existing_tube or tube_label)
                    aligned_rows.append({"Nom du spectre": name, "Tube": tube_val})

                # Si le nombre de tubes change, le dernier "Contrôle" peut changer d'index et
                # récupérer un ancien nom déjà utilisé (ex: ..._13 en Tube 13 ET en Contrôle).
                # Dans ce cas, on renumérote automatiquement comme le bouton Reset.
                names_seen: set[str] = set()
                has_dup = False
                for r in aligned_rows:
                    nm = str(r.get("Nom du spectre", "")).strip()
                    if not nm:
                        continue
                    if nm in names_seen:
                        has_dup = True
                        break
                    names_seen.add(nm)

                need_renumber = has_dup or (len(existing_by_tube) != len(desired_keys))
                if need_renumber:
                    for i in range(len(aligned_rows)):
                        aligned_rows[i]["Nom du spectre"] = f"{manip_name}_{start_index + i:02d}"

                mapping_rows = pd.DataFrame(aligned_rows, columns=default_cols)
                if extra_rows:
                    mapping_rows = pd.concat(
                        [mapping_rows, pd.DataFrame(extra_rows, columns=default_cols)],
                        ignore_index=True,
                    )

            # 4) DataFrame initial passé au TableEditorDialog
            initial_df = pd.concat(
                [df_header, blank_row, internal_header, mapping_rows],
                ignore_index=True,
            )
            return initial_df, reset_df

        # Valeurs par défaut des options BRB / Contrôle
        include_brb = self._map_include_brb
        include_control = self._map_include_control
        current_mapping = _extract_mapping_rows(self.df_map)
        if current_mapping is not None and not current_mapping.empty:
            tubes = current_mapping["Tube"].astype(str).str.strip()
            include_brb = any(_is_brb_label(t) for t in tubes)
            include_control = any(self._is_control_label(t) for t in tubes)

        initial_df, reset_df = _build_initial_df(include_brb, include_control, self.df_map)

        # 5) Ouverture du dialogue d'édition
        dlg = TableEditorDialog(
            title="Correspondance spectres ↔ tubes",
            columns=default_cols,
            initial_df=initial_df,
            parent=self,
            reset_df=reset_df,
            reset_button_text="Reset (valeurs par défaut)",
            show_rename_col=False,
        )
        # Options : inclure Contrôle BRB / Contrôle (dernier tube)
        opts_layout = QHBoxLayout()
        opts_layout.addWidget(QLabel("Départ index spectres :", dlg))
        spin_start = QSpinBox(dlg)
        spin_start.setRange(0, 999)
        spin_start.setValue(int(self._spec_start_index))
        spin_start.setToolTip("Index de départ (0 => _00, 7 => _07).")
        opts_layout.addWidget(spin_start)
        chk_brb = QCheckBox('Inclure "Contrôle BRB"', dlg)
        chk_ctrl = QCheckBox('Inclure "Contrôle" (dernier tube)', dlg)
        chk_brb.setChecked(bool(include_brb))
        chk_ctrl.setChecked(bool(include_control))
        opts_layout.addWidget(chk_brb)
        opts_layout.addWidget(chk_ctrl)
        opts_layout.addStretch(1)
        try:
            dlg.layout().insertLayout(1, opts_layout)
        except Exception:
            pass

        def _rebuild_mapping_table() -> None:
            self._spec_start_index = int(spin_start.value())
            inc_brb = chk_brb.isChecked()
            inc_ctrl = chk_ctrl.isChecked()
            try:
                current_df = dlg.to_dataframe()
            except Exception:
                current_df = self.df_map
            new_df, new_reset = _build_initial_df(inc_brb, inc_ctrl, current_df)
            dlg._reset_df = new_reset
            try:
                dlg._load_from_dataframe(new_df)
            except Exception:
                pass

        chk_brb.toggled.connect(_rebuild_mapping_table)
        chk_ctrl.toggled.connect(_rebuild_mapping_table)
        spin_start.valueChanged.connect(_rebuild_mapping_table)
        if dlg.exec() != QDialog.Accepted:
            return
        self._spec_start_index = int(spin_start.value())

        # 6) Récupération du tableau édité et MISE À JOUR forcée des lignes d'en-tête
        df = dlg.to_dataframe()
        if df.empty:
            QMessageBox.warning(self, "Tableau vide", "Le tableau de correspondance est vide.")
            return

        # On force les valeurs des lignes d'en-tête à partir de l'UI, pour éviter
        # tout décalage : le clic sur le bouton reflète immédiatement les derniers champs saisis.
        df = self._apply_header_values_to_df_map(df, manip_name)

        # Mémoriser les options BRB / Contrôle
        self._map_include_brb = bool(chk_brb.isChecked())
        self._map_include_control = bool(chk_ctrl.isChecked())

        self.df_map = df
        self.map_dirty = False
        self._mark_metadata_dirty(refresh=False)
        self.lbl_status_map.setText(
            f"Tableau de correspondance : {df.shape[0]} ligne(s), {df.shape[1]} colonne(s)"
        )
        self._refresh_button_states()

    def _apply_spectrum_names_to_df_map(self, manip_name: str, start_index: int) -> None:
        """Met à jour la colonne 'Nom du spectre' dans df_map sans modifier la colonne 'Tube'.

        Règles :
        - On ne touche jamais aux lignes d'en-tête (nom du prélèvement, date, etc.).
        - Si une ligne interne d'en-tête "Nom du spectre" / "Tube" existe, on met à jour uniquement
        les lignes situées après cette ligne.
        - Si cette ligne n'existe pas, on met à jour uniquement les lignes qui ressemblent à un nom
        de spectre (pattern *_NN).
        - Si df_map n'existe pas encore, on crée un mapping minimal avec tubes par défaut.
        """
        manip_name = (manip_name or "").strip()
        if not manip_name:
            return

        # --- Création minimaliste si df_map n'existe pas encore ---
        if self.df_map is None or not isinstance(self.df_map, pd.DataFrame) or self.df_map.empty:
            tube_labels_default = self._default_tube_labels_for_mapping()
            n = len(tube_labels_default)
            names = [f"{manip_name}_{i:02d}" for i in range(start_index, start_index + n)]
            self.df_map = pd.DataFrame(
                {
                    "Nom du spectre": ["Nom du spectre"] + names,
                    "Tube": ["Tube"] + tube_labels_default,
                }
            )
            return

        df = self.df_map.copy()
        if not {"Nom du spectre", "Tube"}.issubset(df.columns):
            return

        # Cherche la ligne d'en-tête interne "Nom du spectre" / "Tube"
        hdr_idx = None
        for ridx, row in df.iterrows():
            left = str(row.get("Nom du spectre", "")).strip()
            right = str(row.get("Tube", "")).strip()
            if left == "Nom du spectre" and right == "Tube":
                hdr_idx = ridx
                break

        # Déterminer les indices des lignes à mettre à jour
        if hdr_idx is not None:
            mapping_idx = [i for i in df.index if i > hdr_idx]
        else:
            # Fallback : ne mettre à jour que les lignes qui ressemblent à un nom de spectre (ex: XYZ_00)
            col_ns = df["Nom du spectre"].astype(str).str.strip()
            mask = col_ns.str.match(r"^.+_\d+$", na=False)
            mapping_idx = df.index[mask].tolist()

        if not mapping_idx:
            return

        names = [f"{manip_name}_{i:02d}" for i in range(start_index, start_index + len(mapping_idx))]
        for k, ridx in enumerate(mapping_idx):
            df.at[ridx, "Nom du spectre"] = names[k]

        # Important : ne jamais toucher à la colonne Tube
        self.df_map = df
        
    def _on_manip_name_changed(self, *args) -> None:
        """Met à jour les noms de spectres générés automatiquement."""
        manip_name = self.edit_manip.text().strip() if hasattr(self, "edit_manip") else ""
        start_index = int(self._spec_start_index)

        if not manip_name:
            return

        # Si la titration est active et qu'un nom de manip de titration est défini,
        # les noms de spectres suivent la titration — ne pas les écraser avec le nom
        # du prélèvement.
        if hasattr(self, "chk_titration_done") and self.chk_titration_done.isChecked():
            titration_widget = getattr(self, "edit_titration_manip", None)
            if titration_widget and titration_widget.text().strip():
                self._mark_metadata_dirty(refresh=False)
                self._refresh_button_states()
                return

        had_map = self.df_map is not None and isinstance(self.df_map, pd.DataFrame) and not self.df_map.empty
        was_dirty = bool(self.map_dirty)
        self._apply_spectrum_names_to_df_map(manip_name, start_index)

        self.map_dirty = True if not had_map else was_dirty
        self._mark_metadata_dirty(refresh=False)
        self._refresh_button_states()

    def _on_validate_volume_preset_clicked(self) -> None:
        """Applique le modèle sélectionné au tableau des volumes (df_comp)."""
        name = self.combo_volume_preset.currentText().strip()
        if not name or name not in self.VOLUME_PRESETS:
            QMessageBox.warning(self, "Modèle introuvable", "Le modèle sélectionné est introuvable.")
            return

        self.df_comp = self.VOLUME_PRESETS[name].copy()
        if hasattr(self, "chk_titration_done"):
            self.chk_titration_done.setChecked(True)
        try:
            conc_df = mm.compute_concentration_table(self.df_comp)
            tube_cols = self._get_tube_columns(self.df_comp)
            if tube_cols and not conc_df.empty:
                self._sum_target_uL = float(conc_df.at[0, tube_cols[0]])
        except Exception:
            self._sum_target_uL = DEFAULT_VTOT_UL
        self.lbl_status_comp.setText(
            f"Tableau des volumes : {self.df_comp.shape[0]} ligne(s), {self.df_comp.shape[1]} colonne(s)"
        )

        QMessageBox.information(
            self,
            "Modèle appliqué",
            f"Le modèle '{name}' a été appliqué au tableau des volumes.\n\n"
            "Vous pouvez maintenant ouvrir 'Créer / éditer le tableau des volumes' pour l'ajuster.",
        )
        self._mark_metadata_dirty(refresh=False)
        self._sync_df_map_with_df_comp()
        self._refresh_button_states()

    def _apply_manip_name_to_df_map(self, manip_name: str) -> None:
        """
        Applique le nom du prélèvement à df_map :
        - met à jour la ligne 'Nom du prélèvement'
        - régénère tous les noms de spectres sous la forme NomManip_XX
        sans modifier l'affectation des tubes.
        """
        if self.df_map is None or self.df_map.empty:
            return
        if "Nom du spectre" not in self.df_map.columns or "Tube" not in self.df_map.columns:
            return

        df = self.df_map.copy()
        col_ns = df["Nom du spectre"].astype(str).str.strip()

        # 1) Mettre à jour la ligne d'en-tête "Nom du prélèvement"
        mask_manip = col_ns.isin([
            "Nom du prélèvement",
            "Nom du prélèvement :",
            "Nom de la manip",
            "Nom de la manip :",
        ])
        if mask_manip.any():
            df.loc[mask_manip, "Nom du spectre"] = "Nom du prélèvement"
            df.loc[mask_manip, "Tube"] = manip_name

        # 2) Trouver la ligne d'en-tête interne "Nom du spectre"
        header_mask = col_ns == "Nom du spectre"
        if not header_mask.any():
            self.df_map = df
            self._mark_metadata_dirty()
            return

        last_header_idx = header_mask[header_mask].index[-1]

        # 3) Régénérer les noms des spectres après cette ligne
        mapping_idx = df.index[df.index > last_header_idx].tolist()
        for i, ridx in enumerate(mapping_idx):
            df.at[ridx, "Nom du spectre"] = f"{manip_name}_{i:02d}"

        self.df_map = df

        # Mettre à jour le label d'état si présent
        if hasattr(self, "lbl_status_map"):
            self.lbl_status_map.setText(
                f"Tableau de correspondance : {df.shape[0]} ligne(s), {df.shape[1]} colonne(s)"
            )
        self._mark_metadata_dirty()

    def _on_generate_volumes_clicked(self) -> None:
        """Ouvre le dialogue de génération des volumes selon une distribution gaussienne."""

        dlg = GaussianVolumesDialog(self)
        # Restaurer les derniers paramètres utilisés si disponibles
        if self._last_gaussian_params is not None:
            dlg._apply_params(self._last_gaussian_params, self._last_gaussian_fixed_stock)
        if dlg.exec() != QDialog.Accepted:
            return

        params = dlg.params()
        fixed_stock = dlg.fixed_solution_stocks()
        # Mémoriser les paramètres pour la prochaine ouverture
        self._last_gaussian_params = dict(params)
        self._last_gaussian_params["_V_A1"] = dlg._spin_to_uL(dlg.spin_VA1)
        self._last_gaussian_params["_V_C"] = dlg._spin_to_uL(dlg.spin_VC)
        self._last_gaussian_params["_V_D"] = dlg.spin_VD.value()
        self._last_gaussian_params["_V_E"] = dlg.spin_VE.value()
        self._last_gaussian_params["_V_F"] = dlg._spin_to_uL(dlg.spin_VF)
        self._last_gaussian_fixed_stock = dict(fixed_stock)
        compte_goutte = bool(params.pop("compte_goutte", False))
        add_control   = bool(params.pop("add_control", False))

        if compte_goutte:
            params["max_Vtot_uL"] = 3400.0

        # duplicate_last n'est plus utilisé par le modèle
        params.pop("duplicate_last", None)

        try:
            df_gen, meta = mm.sers_gaussian_volumes(**params)
        except Exception as e:
            QMessageBox.critical(self, "Erreur génération", f"Impossible de générer les volumes :\n{e}")
            return

        # Mettre à jour la cible de volume APRÈS génération : en mode compte-goutte
        # le Vtot réel peut être supérieur au Vtot demandé.
        self._sum_target_uL = float(meta.get("Vtot_real_uL", params.get("Vtot_uL", DEFAULT_VTOT_UL)))

        # Toujours définir tube_cols une seule fois
        n_tubes = int(df_gen["Tube"].max())
        tube_cols = [f"Tube {i}" for i in range(1, n_tubes + 1)]

        # --- Initialiser df_comp si aucun modèle n'a été validé ---
        if self.df_comp is None or not isinstance(self.df_comp, pd.DataFrame) or self.df_comp.empty:
            cC, uC = fixed_stock.get("Solution C", (2.0, "µM"))
            cD, uD = fixed_stock.get("Solution D", (0.5, "% en masse"))
            cE, uE = fixed_stock.get("Solution E", (1.0, "mM"))
            cF, uF = fixed_stock.get("Solution F", (1.0, "mM"))
            cA1, uA1 = fixed_stock.get("Solution A1", (70.0, "mM"))
            base_rows = [
                {"Réactif": "Echantillon", "Concentration": float(params["C0_nM"]), "Unité": "nM"},
                {"Réactif": "Contrôle",    "Concentration": float(params["C0_nM"]), "Unité": "nM"},
                {"Réactif": "Solution A1", "Concentration": float(cA1),             "Unité": str(uA1)},
                {"Réactif": "Solution A2", "Concentration": 7.0,                    "Unité": "mM"},
                {"Réactif": "Solution B",  "Concentration": float(params["C_titrant_uM"]), "Unité": "µM"},
                {"Réactif": "Solution C",  "Concentration": float(cC), "Unité": str(uC)},
                {"Réactif": "Solution D",  "Concentration": float(cD), "Unité": str(uD)},
                {"Réactif": "Solution E",  "Concentration": float(cE), "Unité": str(uE)},
                {"Réactif": "Solution F",  "Concentration": float(cF), "Unité": str(uF)},
            ]
            df = pd.DataFrame(base_rows)
            for c in tube_cols:
                df[c] = 0.0

            # Supprimer les colonnes Tube au-delà de n_tubes (pour que la taille corresponde exactement)
            existing = self._get_tube_columns(df)
            extra = [c for c in existing if c not in tube_cols]
            if extra:
                df = df.drop(columns=extra)
            self.df_comp = df
            if hasattr(self, "chk_titration_done"):
                self.chk_titration_done.setChecked(True)

        # Copie de travail
        df = self.df_comp.copy()

        # Harmoniser les colonnes de concentration/unité si besoin
        if "Concentration" not in df.columns and "Conc" in df.columns:
            df = df.rename(columns={"Conc": "Concentration"})
        if "Unité" not in df.columns and "Unit" in df.columns:
            df = df.rename(columns={"Unit": "Unité"})
        for c in ("Concentration", "Unité"):
            if c not in df.columns:
                df[c] = pd.NA

        # Vérif structure minimale
        if "Réactif" not in df.columns:
            QMessageBox.critical(self, "Erreur", "La colonne 'Réactif' est introuvable dans le tableau des volumes.")
            return

        # Aligner les colonnes Tube exactement sur le nombre demandé
        existing_tube_cols = self._get_tube_columns(df)
        extra = [c for c in existing_tube_cols if c not in tube_cols]
        if extra:
            df = df.drop(columns=extra)
        for c in tube_cols:
            if c not in df.columns:
                df[c] = 0.0

        def ensure_row(reactif_name: str) -> int:
            mask = df["Réactif"].astype(str).str.strip().str.lower() == reactif_name.strip().lower()
            if mask.any():
                return int(df.index[mask][0])
            new = {col: pd.NA for col in df.columns}
            new["Réactif"] = reactif_name
            df.loc[len(df)] = new
            return int(df.index[-1])

        # A1 (tampon concentré fixe), A2 (tampon variable), B (titrant)
        idx_A1 = ensure_row("Solution A1")
        idx_A = ensure_row("Solution A2")
        idx_B = ensure_row("Solution B")
        idx_sample = ensure_row("Echantillon")
        idx_control = ensure_row("Contrôle")
        idx_C = ensure_row("Solution C")
        idx_D = ensure_row("Solution D")
        idx_E = ensure_row("Solution E")
        idx_F = ensure_row("Solution F")

        # Renseigner automatiquement les concentrations stock (saisie utilisateur dans le dialogue)
        df.at[idx_sample, "Concentration"] = float(params["C0_nM"])
        df.at[idx_sample, "Unité"] = "nM"

        df.at[idx_B, "Concentration"] = float(params["C_titrant_uM"])
        df.at[idx_B, "Unité"] = "µM"

        # Solutions fixes : concentration stock + unité (saisie dans le dialogue)
        cA1, uA1 = fixed_stock.get("Solution A1", (70.0, "mM"))
        cC, uC = fixed_stock.get("Solution C", (2.0, "µM"))
        cD, uD = fixed_stock.get("Solution D", (0.5, "% en masse"))
        cE, uE = fixed_stock.get("Solution E", (1.0, "mM"))
        cF, uF = fixed_stock.get("Solution F", (1.0, "mM"))

        df.at[idx_A1, "Concentration"] = float(cA1)
        df.at[idx_A1, "Unité"] = str(uA1)

        df.at[idx_C, "Concentration"] = float(cC)
        df.at[idx_C, "Unité"] = str(uC)

        df.at[idx_D, "Concentration"] = float(cD)
        df.at[idx_D, "Unité"] = str(uD)

        df.at[idx_E, "Concentration"] = float(cE)
        df.at[idx_E, "Unité"] = str(uE)

        df.at[idx_F, "Concentration"] = float(cF)
        df.at[idx_F, "Unité"] = str(uF)

        for _, r in df_gen.iterrows():
            col = f"Tube {int(r['Tube'])}"
            df.at[idx_B, col] = float(r["V_B_titrant_uL"])
            df.at[idx_A, col] = float(r["V_A_tampon_uL"])

        # Fixes : ne remplit que si les lignes existent
        fixed_map = {
            "Echantillon": params["V_echant_uL"],
            "Contrôle": 0.0,
            "Solution A1": params["V_Solution_A1_uL"],
            "Solution C": params["V_Solution_C_uL"],
            "Solution D": params["V_Solution_D_uL"],
            "Solution E": params["V_Solution_E_uL"],
            "Solution F": params["V_Solution_F_uL"],
        }

        for reactif, val in fixed_map.items():
            mask = df["Réactif"].astype(str).str.strip().str.lower() == reactif.strip().lower()
            if not mask.any():
                continue
            ridx = int(df.index[mask][0])
            for c in tube_cols:
                df.at[ridx, c] = float(val)

        # Tous les tubes ont la même dose d'échantillon — plus de dernier tube spécial.
        # Tube contrôle optionnel : colonne supplémentaire calquée sur le tube le plus proche de Veq,
        # mais avec Contrôle=V_echant et Echantillon=0.
        if add_control:
            veq = meta.get("Veq_uL", 0.0)
            # Chercher le tube dont V_B est le plus proche de Veq
            best_col = tube_cols[0]
            best_dist = float("inf")
            ridx_B = df.index[df["Réactif"].astype(str).str.strip() == "Solution B"][0] \
                if "Solution B" in df["Réactif"].values else None
            if ridx_B is not None:
                for c in tube_cols:
                    try:
                        dist = abs(float(df.at[ridx_B, c]) - veq)
                        if dist < best_dist:
                            best_dist = dist
                            best_col = c
                    except (ValueError, TypeError):
                        pass

            ctrl_col = f"Tube {len(tube_cols) + 1}"
            df[ctrl_col] = df[best_col]  # copie du tube le plus proche de Veq

            # Remplacer Echantillon → 0, Contrôle → V_echant
            if "Echantillon" in df["Réactif"].values:
                ridx_samp = int(df.index[df["Réactif"] == "Echantillon"][0])
                df.at[ridx_samp, ctrl_col] = 0.0
            if "Contrôle" in df["Réactif"].values:
                ridx_ctrl = int(df.index[df["Réactif"] == "Contrôle"][0])
                df.at[ridx_ctrl, ctrl_col] = float(params["V_echant_uL"])

            tube_cols = mm.get_tube_columns(df)  # rafraîchir

        # Ligne Contrôle = 0 dans tous les tubes normaux (elle ne sert que dans le tube contrôle)
        if "Contrôle" in df["Réactif"].values:
            ridx_ctrl = int(df.index[df["Réactif"] == "Contrôle"][0])
            for c in (tube_cols[:-1] if add_control else tube_cols):
                df.at[ridx_ctrl, c] = 0.0

        # ── Mode compte-goutte : convertir les volumes B, C, F en nombre de gouttes ──
        # La colonne "Pas (µL)" mémorise le pas pour que compute_concentration_table
        # puisse retrouver les µL réels (vol_µL = n_gouttes × pas).
        pas = params.get("pipette_step_uL", 1.0)
        if "Pas (µL)" not in df.columns:
            df["Pas (µL)"] = 0.0
        if compte_goutte and pas > 0:
            for reactif_name in ("Solution A1", "Solution A2", "Solution B", "Solution C", "Solution F"):
                mask = df["Réactif"].astype(str).str.strip().str.lower() == reactif_name.strip().lower()
                if not mask.any():
                    continue
                ridx = int(df.index[mask][0])
                df.at[ridx, "Pas (µL)"] = pas
                for c in tube_cols:
                    vol_uL = df.at[ridx, c]
                    try:
                        n = float(vol_uL) / pas
                        df.at[ridx, c] = round(n)
                    except (ValueError, TypeError, ZeroDivisionError):
                        pass

        # Réordonner : Echantillon et Contrôle en tête, puis les solutions
        _ROW_ORDER = ["Echantillon", "Contrôle",
                      "Solution A1", "Solution A2", "Solution B", "Solution C",
                      "Solution D", "Solution E", "Solution F"]
        def _row_key(r):
            name = str(r.get("Réactif", "")).strip()
            try:
                return _ROW_ORDER.index(name)
            except ValueError:
                return len(_ROW_ORDER)
        df = df.iloc[sorted(range(len(df)), key=lambda i: _row_key(df.iloc[i]))].reset_index(drop=True)

        self.df_comp = df
        if hasattr(self, "chk_titration_done"):
            self.chk_titration_done.setChecked(True)

        vtot_used = meta.get("Vtot_real_uL", 0.0)
        vtot_msg = (f"V total utilisé = {vtot_used:.0f} µL\n\n" if compte_goutte else "")
        QMessageBox.information(
            self,
            "Volumes générés",
            "Les volumes ont été générés et appliqués au tableau des volumes.\n\n"
            f"{vtot_msg}"
            f"Veq ≈ {meta.get('Veq_uL', 0.0):.2f} µL\n\n"
            f"V_AB = {meta.get('V_AB_real_uL', 0.0):.2f} µL\n",
        )

        self._sync_df_map_with_df_comp(duplicate_last_override=add_control)

    def _get_tube_columns(self, df: pd.DataFrame) -> list[str]:
        return mm.get_tube_columns(df)

    def _infer_n_tubes_for_mapping(self) -> int:
        return mm.infer_n_tubes_for_mapping(self.df_comp)

    def _sync_df_map_with_df_comp(
        self,
        duplicate_last_override: bool | None = None,
        include_brb_override: bool | None = None,
        include_control_override: bool | None = None,
        *,
        mark_dirty: bool = True,
    ) -> None:
        """Synchronise automatiquement la correspondance spectres↔tubes avec le tableau des volumes.

        Objectifs :
        - le nombre de lignes de correspondance suit le nombre de colonnes 'Tube N'
        - le dernier tube est 'Contrôle' uniquement si un duplicat est détecté
        - pas de doublons de noms de spectres (renumérotation si nécessaire)

        Cette méthode n'ouvre pas de dialogue : elle met à jour self.df_map en place.
        Par défaut, elle marque la correspondance comme à revoir (pas de validation utilisateur).
        """
        default_cols = ["Nom du spectre", "Tube"]
        include_brb = self._map_include_brb if include_brb_override is None else bool(include_brb_override)
        include_control = (
            self._map_include_control if include_control_override is None else bool(include_control_override)
        )

        # Récupération des infos d'en-tête depuis l'UI
        try:
            manip_name = self.edit_manip.text().strip() if hasattr(self, "edit_manip") else ""
        except Exception:
            manip_name = ""

        df_header = self._header_rows_dataframe(manip_name)
        blank_row = pd.DataFrame([["", ""]], columns=default_cols)
        internal_header = pd.DataFrame([["Nom du spectre", "Tube"]], columns=default_cols)

        # Tubes attendus d'après df_comp
        n_tubes = self._infer_n_tubes_for_mapping()
        tube_labels_default = self._default_tube_labels_for_mapping(
            n_tubes,
            duplicate_last_override=duplicate_last_override,
            include_brb_override=include_brb,
            include_control_override=include_control,
        )
        start_index = int(self._spec_start_index)

        def _make_mapping_rows(labels: list[str]) -> pd.DataFrame:
            prefix = manip_name if manip_name else "Spectrum"
            names = [f"{prefix}_{i:02d}" for i in range(start_index, start_index + len(labels))]
            return pd.DataFrame({"Nom du spectre": names, "Tube": labels}, columns=default_cols)

        # Extraire la partie "mapping" existante si possible
        mapping_rows = None
        if self.df_map is not None and isinstance(self.df_map, pd.DataFrame) and not self.df_map.empty:
            existing = self.df_map.copy()
            if {"Nom du spectre", "Tube"}.issubset(existing.columns):
                try:
                    col_ns = existing["Nom du spectre"].astype(str).str.strip()
                    col_t = existing["Tube"].astype(str).str.strip()
                    hdr_mask = (col_ns == "Nom du spectre") & (col_t == "Tube")
                    if hdr_mask.any():
                        last_hdr_idx = hdr_mask[hdr_mask].index[-1]
                        mapping_rows = existing.loc[existing.index > last_hdr_idx, ["Nom du spectre", "Tube"]].copy()
                    else:
                        mapping_rows = existing[["Nom du spectre", "Tube"]].copy()
                    mapping_rows = mapping_rows.reset_index(drop=True)
                except Exception:
                    mapping_rows = None

        if mapping_rows is None or mapping_rows.empty:
            mapping_rows = _make_mapping_rows(tube_labels_default)
        else:
            def _clean_text(v) -> str:
                if v is None or (isinstance(v, float) and pd.isna(v)) or pd.isna(v):
                    return ""
                s = str(v).strip()
                return "" if s == "<NA>" else s

            existing_by_tube: dict[str, dict[str, str]] = {}
            desired_keys = {self._tube_merge_key_for_mapping(t, n_tubes) for t in tube_labels_default}

            for _, row in mapping_rows.iterrows():
                tube_raw = _clean_text(row.get("Tube", ""))
                name_raw = _clean_text(row.get("Nom du spectre", ""))
                if (not include_brb) and ("brb" in self._norm_text_key(tube_raw)):
                    continue
                tube_key = self._tube_merge_key_for_mapping(tube_raw, n_tubes)
                if not tube_key:
                    continue
                if tube_key in desired_keys:
                    if tube_key not in existing_by_tube:
                        existing_by_tube[tube_key] = {"Nom du spectre": name_raw, "Tube": tube_raw}
                    else:
                        if name_raw and not existing_by_tube[tube_key].get("Nom du spectre"):
                            existing_by_tube[tube_key]["Nom du spectre"] = name_raw

            aligned_rows: list[dict] = []
            prefix = manip_name if manip_name else "Spectrum"
            for i, tube_label in enumerate(tube_labels_default):
                tube_key = self._tube_merge_key_for_mapping(tube_label, n_tubes)
                existing = existing_by_tube.get(tube_key, {})
                name = existing.get("Nom du spectre") or f"{prefix}_{start_index + i:02d}"
                existing_tube = existing.get("Tube") or ""

                if self._is_control_label(tube_label):
                    tube_val = "Contrôle"
                else:
                    tube_val = tube_label if self._is_control_label(existing_tube) else (existing_tube or tube_label)
                aligned_rows.append({"Nom du spectre": name, "Tube": tube_val})

            # Renumérote automatiquement si doublons / mismatch (même comportement que l'ouverture du dialogue)
            names_seen: set[str] = set()
            has_dup = False
            for r in aligned_rows:
                nm = str(r.get("Nom du spectre", "")).strip()
                if not nm:
                    continue
                if nm in names_seen:
                    has_dup = True
                    break
                names_seen.add(nm)

            need_renumber = has_dup or (len(existing_by_tube) != len(desired_keys))
            if need_renumber:
                for i in range(len(aligned_rows)):
                    aligned_rows[i]["Nom du spectre"] = f"{prefix}_{start_index + i:02d}"

            mapping_rows = pd.DataFrame(aligned_rows, columns=default_cols)

        self.df_map = pd.concat([df_header, blank_row, internal_header, mapping_rows], ignore_index=True)
        self.map_dirty = bool(mark_dirty)
        self._mark_metadata_dirty(refresh=False)

        if hasattr(self, "lbl_status_map"):
            self.lbl_status_map.setText(
                f"Tableau de correspondance : {self.df_map.shape[0]} ligne(s), {self.df_map.shape[1]} colonne(s)"
            )

        self._refresh_button_states()

    def _infer_duplicate_last_for_mapping(self, n_tubes: int | None = None) -> bool:
        """Devine si le dernier tube est un duplicat (contrôle).

        Heuristique : si, dans le tableau des volumes (df_comp), la ligne "Solution B"
        a les deux derniers volumes identiques (Tube N-1 == Tube N), alors on considère
        que le dernier tube est un réplicat.

        Si l'information n'est pas disponible, on suppose True (comportement historique).
        """
        if n_tubes is None:
            n = self._infer_n_tubes_for_mapping()
        else:
            n = max(1, int(n_tubes))

        if n < 2:
            return True

        if self.df_comp is None or not isinstance(self.df_comp, pd.DataFrame) or self.df_comp.empty:
            return True

        df = self.df_comp
        if "Réactif" not in df.columns:
            return True

        # Trouver la ligne Solution B (titrant)
        reactif = df["Réactif"].astype(str).str.strip().str.lower()
        mask = reactif == "solution b"
        if not mask.any():
            mask = reactif.str.startswith("solution b")
        if not mask.any():
            return True
        ridx = int(mask[mask].index[0])

        # Colonnes des 2 derniers tubes
        col_prev = f"Tube {n - 1}"
        col_last = f"Tube {n}"
        if col_prev in df.columns and col_last in df.columns:
            cols = [col_prev, col_last]
        else:
            tube_cols = self._get_tube_columns(df)
            if len(tube_cols) < 2:
                return True
            cols = tube_cols[-2:]

        v_prev = self._to_float(df.at[ridx, cols[0]])
        v_last = self._to_float(df.at[ridx, cols[1]])
        if v_prev is None or v_last is None:
            return True

        return abs(float(v_last) - float(v_prev)) <= 1e-9

    @staticmethod
    def _norm_text_key(s: str) -> str:
        return mm.norm_text_key(s)

    def _is_control_label(self, label: str) -> bool:
        return mm.is_control_label(label)

    def _tube_merge_key_for_mapping(self, label: str, n_tubes: int | None = None) -> str:
        return mm.tube_merge_key_for_mapping(label, n_tubes, self.df_comp)

    def _default_tube_labels_for_mapping(
        self,
        n_tubes: int | None = None,
        *,
        duplicate_last_override: bool | None = None,
        include_brb_override: bool | None = None,
        include_control_override: bool | None = None,
    ) -> list[str]:
        """Liste de tubes par défaut pour la correspondance.

        Convention :
        - 1er spectre : "Contrôle BRB" (optionnel)
        - puis tubes expérimentaux :
            - si duplicat actif : Tube 1..Tube (N-1) + "Contrôle" (réplicat)
            - sinon             : Tube 1..Tube N
        """
        n = self._infer_n_tubes_for_mapping() if n_tubes is None else max(1, int(n_tubes))
        duplicate_last = (
            bool(duplicate_last_override)
            if duplicate_last_override is not None
            else self._infer_duplicate_last_for_mapping(n)
        )
        include_brb = self._map_include_brb if include_brb_override is None else bool(include_brb_override)
        include_control = (
            self._map_include_control if include_control_override is None else bool(include_control_override)
        )
        if n <= 1:
            tubes = ["Contrôle" if (duplicate_last and include_control) else "Tube 1"]
        elif duplicate_last and include_control:
            tubes = [f"Tube {i}" for i in range(1, n)] + ["Contrôle"]
        else:
            tubes = [f"Tube {i}" for i in range(1, n + 1)]
        return (["Contrôle BRB"] if include_brb else []) + tubes
    # ------------------------------------------------------------------
    # Calcul du tableau des concentrations à partir des volumes
    # ------------------------------------------------------------------
    def _compute_concentration_table(self) -> pd.DataFrame:
        return mm.compute_concentration_table(self.df_comp)
    def _on_show_conc_clicked(self) -> None:
        """Affiche le tableau des concentrations dans les cuvettes (information uniquement)."""

        if self.df_comp is None or self.df_comp.empty:
            QMessageBox.warning(self, "Erreur", "Le tableau des volumes n'est pas défini.")
            return

        conc_df = self._compute_concentration_table()

        dlg = TableEditorDialog(
            "Concentration dans les cuvettes (info)",
            list(conc_df.columns),
            conc_df,
            self,
        )

        # Lecture seule
        dlg.table.setEditTriggers(QTableWidget.NoEditTriggers)

        dlg.exec()

    # ------------------------------------------------------------------
    # Méthode utilitaire optionnelle
    # ------------------------------------------------------------------
    def build_merged_metadata(self) -> pd.DataFrame:
        df_map = self._df_map_with_current_header(self.df_map) if self.df_map is not None else self.df_map
        return mm.build_merged_metadata(self.df_comp, df_map)
    def _update_header_fields_from_df_map(self) -> None:
        """Met à jour les champs d'en-tête
        à partir des premières lignes du tableau de correspondance df_map, si possible.
        """
        if self.df_map is None or self.df_map.empty:
            return
        if "Nom du spectre" not in self.df_map.columns or "Tube" not in self.df_map.columns:
            return

        df = self.df_map

        def _clean_cell(value) -> str:
            if value is None:
                return ""
            try:
                if pd.isna(value):
                    return ""
            except Exception:
                pass
            text = str(value).strip()
            return "" if text.lower() in {"nan", "nat", "none", "<na>"} else text

        col_ns_all = df["Nom du spectre"].map(_clean_cell)
        col_tube_all = df["Tube"].map(_clean_cell)
        hdr_mask = (col_ns_all == "Nom du spectre") & (col_tube_all == "Tube")
        if hdr_mask.any():
            header_df = df.loc[df.index < hdr_mask[hdr_mask].index[0], ["Nom du spectre", "Tube"]]
        else:
            header_df = df[["Nom du spectre", "Tube"]]

        value_by_label: dict[str, str] = {}
        value_by_key: dict[str, str] = {}
        for _, row in header_df.iterrows():
            label = _clean_cell(row.get("Nom du spectre", ""))
            value = _clean_cell(row.get("Tube", ""))
            if not label:
                continue
            storage_key = self._header_storage_key(label)
            if label not in value_by_label or (not value_by_label[label] and value):
                value_by_label[label] = value
            if storage_key not in value_by_key or (not value_by_key[storage_key] and value):
                value_by_key[storage_key] = value

        def _value_for(*labels: str) -> str:
            """Cherche la valeur pour n'importe laquelle des variantes de label."""
            for label in labels:
                for key in (label, str(label or "").strip().rstrip(":").strip(), self._header_storage_key(label)):
                    if key in value_by_label:
                        return value_by_label[key]
                    if key in value_by_key:
                        return value_by_key[key]
            return ""

        # Nom du prélèvement
        name_val = _value_for("Nom du prélèvement :", "Nom du prélèvement", "Nom de la manip :", "Nom de la manip")
        if name_val and hasattr(self, "edit_manip"):
            self.edit_manip.setText(name_val)

        # Prélèvement
        sampler_val = _value_for("Préleveur·se", "Préleveur/préleveuse", "Préleveur", "Préleveuse")
        if sampler_val and hasattr(self, "edit_sampler"):
            self.edit_sampler.setText(sampler_val)
        self._clear_extra_sampler_fields()
        extra_samplers = [
            _value_for(f"Préleveur·se {idx}", f"Préleveur {idx}", f"Préleveuse {idx}", f"Preleveur {idx}")
            for idx in range(2, 7)
        ]
        if not any(extra_samplers):
            combined_samplers = _value_for("Préleveur·ses", "Préleveurs", "Preleveurs")
            if combined_samplers:
                parts = [p.strip() for p in re.split(r"[;\n]+", combined_samplers) if p.strip()]
                if parts and not sampler_val and hasattr(self, "edit_sampler"):
                    self.edit_sampler.setText(parts[0])
                extra_samplers = parts[1:]

        associations_val = _value_for(
            "Association(s) du prélèvement",
            "Associations prélèvement",
            "Associations prelevement",
            "Association prélèvement",
            "Association prelevement",
            "Associations",
            "Association",
        )
        association_parts = [p.strip() for p in re.split(r"[;\n]+", associations_val) if p.strip()]
        first_association = _value_for(
            "Association préleveur·se",
            "Association préleveur·se 1",
            "Association préleveur 1",
            "Association 1",
        )
        if not first_association and associations_val:
            first_association = association_parts[0] if association_parts else associations_val
        if first_association and hasattr(self, "edit_sampler_association"):
            self.edit_sampler_association.setText(first_association)

        extra_associations = [
            _value_for(
                f"Association préleveur·se {idx}",
                f"Association préleveur {idx}",
                f"Association {idx}",
            )
            for idx in range(2, 7)
        ]
        for idx in range(2, 7):
            part_index = idx - 1
            if not extra_associations[part_index - 1] and len(association_parts) > part_index:
                extra_associations[part_index - 1] = association_parts[part_index]

        max_extra = max(len(extra_samplers), len(extra_associations))
        for idx in range(max_extra):
            extra_sampler = extra_samplers[idx] if idx < len(extra_samplers) else ""
            extra_association = extra_associations[idx] if idx < len(extra_associations) else ""
            if extra_sampler or extra_association:
                self._add_sampler_field(extra_sampler, extra_association)

        sample_loc = _value_for("Lieu du prélèvement")
        if sample_loc and hasattr(self, "edit_sample_location"):
            self.edit_sample_location.setText(sample_loc)

        sample_lat = _value_for(
            "Latitude du prélèvement",
            "Latitude prélèvement",
            "Latitude prelevement",
            "Latitude GPS",
            "Latitude",
            "Lat",
        )
        sample_lon = _value_for(
            "Longitude du prélèvement",
            "Longitude prélèvement",
            "Longitude prelevement",
            "Longitude GPS",
            "Longitude",
            "Long",
            "Lon",
        )
        if not sample_lat or not sample_lon:
            old_gps = _value_for(
                "Coordonnées GPS du prélèvement",
                "Coordonnées GPS",
                "Coordonnees GPS",
                "GPS prélèvement",
                "GPS prelevement",
                "GPS",
            )
            old_lat, old_lon = self._split_gps_pair(old_gps)
            sample_lat = sample_lat or old_lat
            sample_lon = sample_lon or old_lon
        if sample_lat and hasattr(self, "edit_sample_lat"):
            self.edit_sample_lat.setText(self._format_coord_value(sample_lat, -90.0, 90.0))
        if sample_lon and hasattr(self, "edit_sample_lon"):
            self.edit_sample_lon.setText(self._format_coord_value(sample_lon, -180.0, 180.0))

        department = _value_for(
            "Département",
            "Departement",
            "Département du prélèvement",
            "Departement du prelevement",
        )
        if department and hasattr(self, "edit_sample_department"):
            self.edit_sample_department.setText(department)
        commune = _value_for("Commune", "Commune du prélèvement", "Commune du prelevement")
        if commune and hasattr(self, "edit_sample_commune"):
            self.edit_sample_commune.setText(commune)
        if (sample_lat and sample_lon) and (not department or not commune):
            self._try_autofill_sample_admin_from_gps()

        sample_date = _value_for(
            "Date du prélèvement",
            "Date prélèvement",
            "Date prelevement",
            "Date de prélèvement",
            "Date de prelevement",
        )
        sample_time = _value_for(
            "Heure du prélèvement",
            "Heure prélèvement",
            "Heure prelevement",
            "Heure de prélèvement",
            "Heure de prelevement",
        )
        sample_dt = _value_for("Date/heure du prélèvement")
        if sample_dt and (not sample_date or not sample_time):
            if " " in sample_dt:
                old_date, old_time = sample_dt.rsplit(" ", 1)
                sample_date = sample_date or old_date
                sample_time = sample_time or old_time
            else:
                sample_date = sample_date or sample_dt
        if sample_date and hasattr(self, "edit_sample_date"):
            self._set_date_time_text(self.edit_sample_date, sample_date)
        if sample_time and hasattr(self, "edit_sample_time"):
            self._set_time_text(self.edit_sample_time, sample_time)

        water_type = _value_for("Type d'eau", "Type d'eau du prélèvement", "Type d'eau du prelevement", "Type eau")
        if water_type and hasattr(self, "combo_water_type"):
            water_key = self._norm_text_key(water_type)
            water_aliases = {
                "choisir...": "Choisir...",
                "eau douce": "Eau douce",
                "douce": "Eau douce",
                "eau salee": "Eau de mer",
                "eau de mer": "Eau de mer",
                "salee": "Eau de mer",
                "mer": "Eau de mer",
                "eau estuarienne": "Eau estuarienne",
                "eau esturienne": "Eau estuarienne",
                "estuarienne": "Eau estuarienne",
                "esturienne": "Eau estuarienne",
            }
            self.combo_water_type.setCurrentText(water_aliases.get(water_key, water_type))

        tide_coef = _value_for("Coefficient de marée", "Coefficient de maree", "Coefficient maree", "Coef marée", "Coef maree")
        if tide_coef and hasattr(self, "spin_tide_coefficient"):
            val = self._to_float(tide_coef)
            if val is not None:
                self.spin_tide_coefficient.setValue(max(0, min(200, int(round(val)))))

        high_tide = _value_for("Heure de pleine mer", "Heure pleine mer", "Pleine mer")
        if high_tide and hasattr(self, "edit_high_tide_time"):
            self.edit_high_tide_time.setText(self._normalize_time_text(high_tide))
        self._refresh_tide_fields_enabled()
        self._refresh_water_type_style()

        water_temp = _value_for("Température de l'eau", "Temperature de l'eau", "Temp eau")
        if water_temp and hasattr(self, "spin_water_temperature"):
            val = self._to_float(water_temp)
            if val is not None:
                self.spin_water_temperature.setValue(max(-10.0, min(50.0, val)))

        self._debit_flux_values = {column: "" for column in self.DEBIT_FLUX_COLUMNS}
        for column in self.DEBIT_FLUX_COLUMNS:
            debit_flux_val = _value_for(
                f"{self.DEBIT_FLUX_PREFIX}{column}",
                f"Debit / flux - {column}",
                f"Débit/flux - {column}",
                column,
            )
            if debit_flux_val:
                self._debit_flux_values[column] = debit_flux_val

        weather = _value_for("Temps au moment de la mesure", "Meteo", "Météo")
        if weather and hasattr(self, "combo_weather"):
            idx = self.combo_weather.findText(weather, Qt.MatchFixedString | Qt.MatchCaseSensitive)
            if idx < 0:
                idx = self.combo_weather.findText(weather, Qt.MatchFixedString)
            if idx >= 0:
                self.combo_weather.setCurrentIndex(idx)

        air_temp = _value_for("Température de l'air", "Temperature de l'air", "Temp air")
        if air_temp and hasattr(self, "spin_air_temperature"):
            val = self._to_float(air_temp)
            if val is not None:
                self.spin_air_temperature.setValue(max(-30.0, min(60.0, val)))

        rainfall = _value_for("Pluie dernières 24 h (mm)", "Pluie 24h", "Pluie dernieres 24h", "Rainfall 24h")
        if rainfall and hasattr(self, "spin_rainfall_24h"):
            val = self._to_float(rainfall)
            if val is not None:
                self.spin_rainfall_24h.setValue(max(-1.0, min(500.0, val)))

        conditions_comment = _value_for("Commentaire sur les conditions", "Conditions commentaire")
        if conditions_comment and hasattr(self, "edit_conditions_comment"):
            self.edit_conditions_comment.setPlainText(conditions_comment)

        # Ammonium
        ammonium_val = _value_for("Test ammonium réalisé")
        enabled = str(ammonium_val).strip().lower() in {"oui", "yes", "true", "1", "x"}
        self._ammonium_test_enabled = enabled
        if hasattr(self, "chk_ammonium_test"):
            self.chk_ammonium_test.blockSignals(True)
            self.chk_ammonium_test.setChecked(enabled)
            self.chk_ammonium_test.blockSignals(False)
        self._on_ammonium_test_toggled(enabled)
        ammonium_operator = _value_for("Test ammonium réalisé par", "Réalisateur test ammonium", "Réalisatrice test ammonium")
        if ammonium_operator and hasattr(self, "edit_ammonium_operator"):
            self.edit_ammonium_operator.setText(ammonium_operator)
        ammonium_t1 = _value_for("Test ammonium 1 (grossier)", "Ammonium test 1", "Test 1 ammonium")
        if ammonium_t1 and hasattr(self, "edit_ammonium_test1"):
            self.edit_ammonium_test1.setText(ammonium_t1)
        ammonium_t2 = _value_for("Test ammonium 2 (précis)", "Ammonium test 2", "Test 2 ammonium")
        if ammonium_t2 and hasattr(self, "edit_ammonium_test2"):
            self.edit_ammonium_test2.setText(ammonium_t2)
        ammonium_ph = _value_for("pH (ammonium)", "pH ammonium", "pH")
        if ammonium_ph and hasattr(self, "edit_ammonium_ph"):
            self.edit_ammonium_ph.setText(ammonium_ph)

        # Paramètres terrain
        def _load_check_spin(check_attr, spin_attr, toggled_fn, value_keys, status_keys=()):
            status_val = _value_for(*status_keys) if status_keys else ""
            enabled_from_status = self._is_yes_value(status_val)
            val = _value_for(*value_keys)
            f = self._to_float(val) if val else None
            if not enabled_from_status and f is None:
                return
            chk = getattr(self, check_attr, None)
            spin = getattr(self, spin_attr, None)
            if chk is None or spin is None:
                return
            chk.blockSignals(True)
            chk.setChecked(True)
            chk.blockSignals(False)
            toggled_fn(True)
            if f is not None:
                spin.setValue(max(spin.minimum() + 0.01, min(spin.maximum(), f)))

        _load_check_spin(
            "chk_turbidity",
            "spin_turbidity",
            self._on_turbidity_toggled,
            ("Turbidité (NTU)", "Turbidite", "Turbidity"),
            ("Turbidité mesurée", "Turbidite mesuree"),
        )
        _load_check_spin(
            "chk_conductivity",
            "spin_conductivity",
            self._on_conductivity_toggled,
            ("Conductivité (µS/cm)", "Conductivite", "Conductivity"),
            ("Conductivité mesurée", "Conductivite mesuree"),
        )
        _load_check_spin(
            "chk_ph_water",
            "spin_ph_water",
            self._on_ph_water_toggled,
            ("pH de l'eau", "pH eau"),
            ("pH de l'eau mesuré", "pH eau mesure"),
        )
        _load_check_spin(
            "chk_dissolved_o2",
            "spin_dissolved_o2",
            self._on_dissolved_o2_toggled,
            ("Oxygène dissous (mg/L)", "Oxygene dissous", "O2 dissous"),
            ("Oxygène dissous mesuré", "Oxygene dissous mesure", "O2 dissous mesure"),
        )

        # Analyses bactériologiques
        bacterio_val = _value_for(
            "Analyses bactériologiques",
            "Analyse bactériologique",
            "Analyses bacteriologiques",
            "Analyse bacteriologique",
            "Bactériologie",
            "Bacteriologie",
        )
        bacterio_enabled = str(bacterio_val).strip().lower() in {"oui", "yes", "true", "1", "x"}
        if hasattr(self, "chk_bacterio_analysis"):
            self.chk_bacterio_analysis.blockSignals(True)
            self.chk_bacterio_analysis.setChecked(bacterio_enabled)
            self.chk_bacterio_analysis.blockSignals(False)
        bacterio_day = _value_for("Jour de dépôt", "Jour de depot", "Jour dépôt", "Jour depot")
        if bacterio_day and hasattr(self, "edit_bacterio_deposit_day"):
            self._set_date_time_text(self.edit_bacterio_deposit_day, bacterio_day)
        bacterio_time = _value_for("Heure du dépôt", "Heure du depot", "Heure dépôt", "Heure depot")
        if bacterio_time and hasattr(self, "edit_bacterio_deposit_time"):
            self._set_time_text(self.edit_bacterio_deposit_time, bacterio_time)
        bacterio_company = _value_for("Entreprise de mesure", "Entreprise mesure", "Entreprise de mesures")
        if bacterio_company and hasattr(self, "edit_bacterio_company"):
            self.edit_bacterio_company.setText(bacterio_company)
        bacterio_ecoli = _value_for(
            "Résultats E.Coli (npp/ml)",
            "Résultat E.Coli",
            "Résultats E.Coli",
            "Resultat E.Coli",
            "Resultats E.Coli",
            "E.Coli",
            "E. Coli",
        )
        if bacterio_ecoli and hasattr(self, "edit_bacterio_ecoli"):
            self.edit_bacterio_ecoli.setText(bacterio_ecoli)
        bacterio_enterococci = _value_for(
            "Résultats Entérocoques intestinaux (npp/100ml)",
            "Résultat Entérocoques intestinaux",
            "Résultats Entérocoques intestinaux",
            "Resultat Enterocoques intestinaux",
            "Resultats Enterocoques intestinaux",
            "Entérocoques intestinaux",
            "Enterocoques intestinaux",
        )
        if bacterio_enterococci and hasattr(self, "edit_bacterio_enterococci"):
            self.edit_bacterio_enterococci.setText(bacterio_enterococci)
        self._refresh_bacterio_fields_visible()

        titration_val = _value_for("Titration du cuivre", "Titration", "Titration réalisée", "Titration realisee", "Titration faite")
        if titration_val and hasattr(self, "chk_titration_done"):
            titration_checked = str(titration_val).strip().lower() in {"oui", "yes", "true", "1", "x"}
            self.chk_titration_done.blockSignals(True)
            self.chk_titration_done.setChecked(titration_checked)
            self.chk_titration_done.blockSignals(False)

        # Titration (fallback anciens fichiers : Date / Lieu)
        titration_date = _value_for("Date de la titration", "Date titration", "Date de titration")
        titration_time = _value_for("Heure de la titration", "Heure titration", "Heure de titration")
        titration_dt = _value_for("Date/heure de la titration", "Date :", "Date")
        if titration_dt and (not titration_date or not titration_time):
            if " " in titration_dt:
                old_date, old_time = titration_dt.rsplit(" ", 1)
                titration_date = titration_date or old_date
                titration_time = titration_time or old_time
            else:
                titration_date = titration_date or titration_dt
        if titration_date and hasattr(self, "edit_titration_date"):
            self._set_date_time_text(self.edit_titration_date, titration_date)
        if titration_time and hasattr(self, "edit_titration_time"):
            self._set_time_text(self.edit_titration_time, titration_time)

        titration_name = _value_for(
            "Nom de la manip de titration",
            "Nom de la manip",
            "Nom de la titration",
            "Nom manip titration",
        )
        titration_name_manual = _value_for(
            "Nom de la manip de titration manuel",
            "Nom de la titration manuel",
            "Nom titration manuel",
            "Nom manip titration manuel",
        )

        titration_loc = _value_for("Lieu de la titration", "Lieu :", "Lieu")
        if titration_loc and hasattr(self, "edit_titration_location"):
            self.edit_titration_location.setText(titration_loc)

        # Coordinateur
        coord_val = _value_for("Coordinateur :", "Coordinateur", "Coordinateur·ice")
        if coord_val and hasattr(self, "edit_coordinator"):
            self.edit_coordinator.setText(coord_val)

        # Opérateur
        oper_val = _value_for("Opérateur :", "Opérateur", "Opérateur·ice")
        if oper_val and hasattr(self, "edit_operator"):
            self.edit_operator.setText(oper_val)

        if hasattr(self, "edit_sampler") and self.edit_sampler.text().strip():
            self._refresh_manip_name()
        expected_titration_name = self._build_titration_manip_name()
        if titration_name:
            manual_mode = (
                self._is_yes_value(titration_name_manual)
                if titration_name_manual
                else bool(not expected_titration_name or titration_name != expected_titration_name)
            )
            self._set_titration_name_manual_mode(manual_mode)
            if manual_mode:
                self.edit_titration_manip.setText(titration_name)
            else:
                self._refresh_titration_manip_name(force=True)
        else:
            self._set_titration_name_manual_mode(False)
            self._refresh_titration_manip_name(force=True)

        self._refresh_tide_fields_enabled()
        self._refresh_water_type_style()
        self._refresh_bacterio_fields_visible()
        self._refresh_titration_fields_visible()
        self._refresh_fillable_fields_style()
        self._refresh_button_states()

    # ------------------------------------------------------------------
    # Export feuille de protocole
    # ------------------------------------------------------------------
    def _on_export_protocol_clicked(self) -> None:
        """Ouvre le dialogue de suivi interactif du protocole de paillasse."""
        if isinstance(self.df_comp, pd.DataFrame):
            self.df_comp = self._clean_loaded_volume_dataframe(self.df_comp)
        if self.df_comp is None or self.df_comp.empty:
            QMessageBox.warning(
                self,
                "Tableau des volumes manquant",
                "Veuillez d'abord définir le tableau des volumes avant d'ouvrir le protocole.",
            )
            return
        if "Réactif" not in self.df_comp.columns or not self._get_tube_columns(self.df_comp):
            QMessageBox.warning(
                self,
                "Tableau des volumes invalide",
                "Le tableau des volumes chargé ne contient pas les colonnes attendues "
                "('Réactif' et 'Tube N').",
            )
            return

        if self.df_map is not None and not self.df_map.empty:
            self.df_map = self._df_map_with_current_header(self.df_map)

        titration_manip = (
            self.edit_titration_manip.text().strip() if hasattr(self, "edit_titration_manip") else ""
        )
        if not titration_manip and hasattr(self, "edit_manip"):
            titration_manip = self.edit_manip.text().strip()

        meta = {
            "nom_manip":    titration_manip,
            "date":         self._titration_date_time_text(),
            "lieu":         self.edit_location.text().strip()                if hasattr(self, "edit_location")    else "",
            "coordinateur": self.edit_coordinator.text().strip()             if hasattr(self, "edit_coordinator") else "",
            "operateur":    self.edit_operator.text().strip()                if hasattr(self, "edit_operator")    else "",
        }

        parent_window = self.window() or self
        try:
            from protocol_dialog import ProtocolDialog
            dlg = ProtocolDialog(
                self.df_comp,
                meta,
                initial_states=self._protocol_states,
                df_map=self.df_map,
                parent=parent_window,
            )
            dlg.setWindowFlag(Qt.Window, True)
            dlg.setWindowModality(Qt.ApplicationModal)
            try:
                parent_geo = parent_window.frameGeometry()
                dlg.move(parent_geo.center() - dlg.rect().center())
            except Exception:
                pass
            dlg.show()
            dlg.raise_()
            dlg.activateWindow()
            dlg.exec()
        except Exception as e:
            QMessageBox.critical(
                parent_window,
                "Erreur d'ouverture du protocole",
                "La feuille de protocole n'a pas pu s'ouvrir.\n\n"
                f"{type(e).__name__} : {e}",
            )
            return
        # Mémoriser l'état des cases pour la prochaine ouverture / sauvegarde
        previous_states = dict(self._protocol_states)
        self._protocol_states = dlg.get_states()
        if self._protocol_states != previous_states:
            self._mark_metadata_dirty()

    # ------------------------------------------------------------------
    # Sauvegarde / chargement des métadonnées complètes
    # ------------------------------------------------------------------
    def _write_measure_summary_sheet(self, workbook) -> None:
        """Ajoute une feuille terrain lisible en première page du fichier Excel."""
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        if SHEET_MEASURES in workbook.sheetnames:
            del workbook[SHEET_MEASURES]
        ws = workbook.create_sheet(SHEET_MEASURES, 0)

        values = self._header_values()

        def _split_date_time(text: str) -> tuple[str, str]:
            txt = str(text or "").strip()
            if " " not in txt:
                return txt, ""
            date_part, time_part = txt.rsplit(" ", 1)
            return date_part, time_part

        sample_date = values.get("Date du prélèvement", "")
        sample_time = values.get("Heure du prélèvement", "")
        if not sample_date or not sample_time:
            old_sample_date, old_sample_time = _split_date_time(values.get("Date/heure du prélèvement", ""))
            sample_date = sample_date or old_sample_date
            sample_time = sample_time or old_sample_time
        weather_text = str(values.get("Temps au moment de la mesure", "") or "")
        water_temp = values.get("Température de l'eau")
        air_temp = values.get("Température de l'air")
        sampler_names = self._sampler_names()
        sampler_associations = self._sampler_associations()
        sampler_rows_count = max(len(sampler_names), len(sampler_associations), 1)

        border_side = Side(style="thin", color="000000")
        thick_side = Side(style="medium", color="000000")
        thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        thick_bottom = Border(left=border_side, right=border_side, top=border_side, bottom=thick_side)
        title_fill = PatternFill("solid", fgColor="C6E0B4")
        info_fill = PatternFill("solid", fgColor="D9E2F3")
        section_fill = PatternFill("solid", fgColor="F8CBAD")
        header_fill = PatternFill("solid", fgColor="E2F0D9")
        data_blue = PatternFill("solid", fgColor="BDD7EE")
        data_cyan = PatternFill("solid", fgColor="00CFE8")
        context_fill = PatternFill("solid", fgColor="E2F0D9")
        ammonium_fill = PatternFill("solid", fgColor="FCE4D6")
        bacterio_fill = PatternFill("solid", fgColor="DDEBF7")
        debit_fill = PatternFill("solid", fgColor="FFF2CC")
        comment_fill = PatternFill("solid", fgColor="FFE699")
        white_fill = PatternFill("solid", fgColor="FFFFFF")

        def _merge(cell_range: str, text: str = "", fill=None, font=None) -> None:
            ws.merge_cells(cell_range)
            cell = ws[cell_range.split(":")[0]]
            cell.value = text
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if fill is not None:
                cell.fill = fill
            if font is not None:
                cell.font = font

        def _style_range(cell_range: str, fill=None, font=None, border=thin_border) -> None:
            for row in ws[cell_range]:
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    if fill is not None:
                        cell.fill = fill
                    if font is not None:
                        cell.font = font

        sampler_start_row = 7
        sampler_end_row = sampler_start_row + sampler_rows_count - 1
        status_start_row = sampler_end_row + 1
        measure_title_row = status_start_row + 3
        header_top_row = measure_title_row + 1
        header_sub_row = measure_title_row + 2
        data_row_idx = measure_title_row + 3

        # Bloc compact d'identification : il reste lisible et facile à relire
        # même si des colonnes sont ajoutées au tableau des mesures plus bas.
        _merge("A1:H1", "Dispositif de mesure de la qualité de l'eau par les citoyens", title_fill, Font(bold=True))
        info_pairs = [
            (3, "A", "Nom du prélèvement", "B", values.get("Nom du prélèvement", "")),
            (3, "C", "Type d'eau", "D", values.get("Type d'eau", "")),
            (4, "A", "Département", "B", values.get("Département", "")),
            (4, "C", "Commune", "D", values.get("Commune", "")),
            (5, "A", "Lieu du prélèvement", "B", values.get("Lieu du prélèvement", "")),
            (5, "C", "Latitude", "D", values.get("Latitude du prélèvement", "")),
            (5, "E", "Longitude", "F", values.get("Longitude du prélèvement", "")),
            (6, "A", "Date du prélèvement", "B", sample_date),
            (6, "C", "Heure du prélèvement", "D", sample_time),
        ]
        _merge("A2:H2", "Informations générales", info_fill, Font(bold=True))
        for row_idx, label_col, label, value_col, value in info_pairs:
            label_cell = ws[f"{label_col}{row_idx}"]
            value_cell = ws[f"{value_col}{row_idx}"]
            label_cell.value = label
            value_cell.value = value
            for cell in (label_cell, value_cell):
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            label_cell.fill = info_fill
            value_cell.fill = white_fill
            label_cell.font = Font(bold=True, size=9)

        for row_offset in range(sampler_rows_count):
            row_idx = sampler_start_row + row_offset
            sampler_name = sampler_names[row_offset] if row_offset < len(sampler_names) else ""
            association = sampler_associations[row_offset] if row_offset < len(sampler_associations) else ""
            row_cells = [
                ("A", f"Préleveur·se {row_offset + 1}", True),
                ("B", sampler_name, False),
                ("C", "Association", True),
                ("D", association, False),
            ]
            for col_letter, text, is_label in row_cells:
                cell = ws[f"{col_letter}{row_idx}"]
                cell.value = text
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                if is_label:
                    cell.fill = info_fill
                    cell.font = Font(bold=True, size=9)
                else:
                    cell.fill = white_fill

        status_pairs = [
            (status_start_row, "A", "Titration du cuivre", "B", values.get("Titration du cuivre", "")),
            (status_start_row, "C", "Analyses bactériologiques", "D", values.get("Analyses bactériologiques", "")),
            (status_start_row, "E", "Test ammonium réalisé", "F", values.get("Test ammonium réalisé", "")),
            (status_start_row, "G", "Turbidité mesurée", "H", values.get("Turbidité mesurée", "")),
            (status_start_row + 1, "A", "Conductivité mesurée", "B", values.get("Conductivité mesurée", "")),
            (status_start_row + 1, "C", "pH eau mesuré", "D", values.get("pH de l'eau mesuré", "")),
            (status_start_row + 1, "E", "Oxygène dissous mesuré", "F", values.get("Oxygène dissous mesuré", "")),
        ]
        for row_idx, label_col, label, value_col, value in status_pairs:
            label_cell = ws[f"{label_col}{row_idx}"]
            value_cell = ws[f"{value_col}{row_idx}"]
            label_cell.value = label
            value_cell.value = value
            for cell in (label_cell, value_cell):
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            label_cell.fill = info_fill
            value_cell.fill = white_fill
            label_cell.font = Font(bold=True, size=9)

        _style_range(f"A1:H{status_start_row + 1}", border=thin_border)

        table_columns = [
            ("Météo", "context"),
            ("T° eau (°C)", "context"),
            ("T° air (°C)", "context"),
            ("Pluie 24 h (mm)", "context"),
            ("Coefficient de marée", "context"),
            ("Heure pleine mer", "context"),
            ("Test ammonium réalisé", "ammonium"),
            ("Réalisé par", "ammonium"),
            ("Test 1 grossier", "ammonium"),
            ("Test 2 précis", "ammonium"),
            ("pH ammonium", "ammonium"),
            ("Analyses bactériologiques", "bacterio"),
            ("Jour de dépôt", "bacterio"),
            ("Heure du dépôt", "bacterio"),
            ("Entreprise de mesure", "bacterio"),
            ("E. coli (npp/ml)", "bacterio_result"),
            ("Entérocoques (npp/100ml)", "bacterio_result"),
            ("Turbidité mesurée", "result"),
            ("Turbidité (NTU)", "result"),
            ("Conductivité mesurée", "result"),
            ("Conductivité (µS/cm)", "result"),
            ("pH eau mesuré", "result"),
            ("pH de l'eau", "result"),
            ("Oxygène dissous mesuré", "result"),
            ("Oxygène dissous (mg/L)", "result"),
            ("Titration du cuivre", "titration"),
            ("Nom manip titration", "titration"),
            ("Date titration", "titration"),
            ("Heure titration", "titration"),
            ("Lieu titration", "titration"),
            ("Largeur route", "debit"),
            ("Longueur 6 arches", "debit"),
            ("Hauteur eau", "debit"),
            ("Volume total (m3)", "debit"),
            ("Vitesse (s)", "debit"),
            ("Débit (m3/s)", "debit"),
            ("Commentaires", "comment"),
        ]
        last_table_col = get_column_letter(len(table_columns))
        _merge(
            f"A{measure_title_row}:{last_table_col}{measure_title_row}",
            "Préleveur·ses, mesures effectuées et résultats",
            section_fill,
            Font(bold=True),
        )

        group_specs = [
            (1, 6, "Contexte terrain", context_fill),
            (7, 11, "Test ammonium", ammonium_fill),
            (12, 17, "Analyses bactériologiques", bacterio_fill),
            (18, 25, "Autres mesures", data_cyan),
            (26, 30, "Titration du cuivre", header_fill),
            (31, 36, "Débit / flux", debit_fill),
            (37, 37, "Commentaires", comment_fill),
        ]
        for start_col, end_col, title, fill in group_specs:
            _merge(
                f"{get_column_letter(start_col)}{header_top_row}:{get_column_letter(end_col)}{header_top_row}",
                title,
                fill,
                Font(bold=True, size=9),
            )

        section_fills = {
            "people": data_blue,
            "context": context_fill,
            "status": header_fill,
            "ammonium": ammonium_fill,
            "bacterio": bacterio_fill,
            "bacterio_result": data_cyan,
            "result": data_cyan,
            "titration": white_fill,
            "debit": debit_fill,
            "comment": comment_fill,
        }
        for col_idx, (header, section) in enumerate(table_columns, start=1):
            cell = ws.cell(row=header_sub_row, column=col_idx)
            cell.value = header
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = section_fills.get(section, header_fill)
            cell.font = Font(bold=True, size=9)

        shared_values = {
            "Météo": weather_text,
            "T° eau (°C)": water_temp,
            "T° air (°C)": air_temp,
            "Pluie 24 h (mm)": values.get("Pluie dernières 24 h (mm)", ""),
            "Coefficient de marée": values.get("Coefficient de marée", ""),
            "Heure pleine mer": values.get("Heure de pleine mer", ""),
            "Test ammonium réalisé": values.get("Test ammonium réalisé", "Non") or "Non",
            "Réalisé par": values.get("Test ammonium réalisé par", ""),
            "Test 1 grossier": values.get("Test ammonium 1 (grossier)", ""),
            "Test 2 précis": values.get("Test ammonium 2 (précis)", ""),
            "pH ammonium": values.get("pH (ammonium)", ""),
            "Analyses bactériologiques": values.get("Analyses bactériologiques", "Non") or "Non",
            "Jour de dépôt": values.get("Jour de dépôt", ""),
            "Heure du dépôt": values.get("Heure du dépôt", ""),
            "Entreprise de mesure": values.get("Entreprise de mesure", ""),
            "E. coli (npp/ml)": values.get("Résultats E.Coli (npp/ml)", ""),
            "Entérocoques (npp/100ml)": values.get("Résultats Entérocoques intestinaux (npp/100ml)", ""),
            "Turbidité mesurée": values.get("Turbidité mesurée", "Non") or "Non",
            "Turbidité (NTU)": values.get("Turbidité (NTU)", ""),
            "Conductivité mesurée": values.get("Conductivité mesurée", "Non") or "Non",
            "Conductivité (µS/cm)": values.get("Conductivité (µS/cm)", ""),
            "pH eau mesuré": values.get("pH de l'eau mesuré", "Non") or "Non",
            "pH de l'eau": values.get("pH de l'eau", ""),
            "Oxygène dissous mesuré": values.get("Oxygène dissous mesuré", "Non") or "Non",
            "Oxygène dissous (mg/L)": values.get("Oxygène dissous (mg/L)", ""),
            "Titration du cuivre": values.get("Titration du cuivre", "Non") or "Non",
            "Nom manip titration": values.get("Nom de la manip de titration", ""),
            "Date titration": values.get("Date de la titration", ""),
            "Heure titration": values.get("Heure de la titration", ""),
            "Lieu titration": values.get("Lieu de la titration", ""),
            "Largeur route": values.get("Débit / flux - Largeur route", ""),
            "Longueur 6 arches": values.get("Débit / flux - Longueur 6 arches", ""),
            "Hauteur eau": values.get("Débit / flux - Hauteur eau", ""),
            "Volume total (m3)": values.get("Débit / flux - Volume total (m3)", ""),
            "Vitesse (s)": values.get("Débit / flux - Vitesse (s)", ""),
            "Débit (m3/s)": values.get("Débit / flux - Débit (m3/s)", ""),
            "Commentaires": values.get("Commentaire sur les conditions", ""),
        }
        for col_idx, (header, section) in enumerate(table_columns, start=1):
            cell = ws.cell(row=data_row_idx, column=col_idx)
            cell.value = shared_values.get(header, "")
            cell.border = thick_bottom
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = white_fill

        widths = {
            "A": 13, "B": 11, "C": 11, "D": 12, "E": 13, "F": 12, "G": 14, "H": 16,
            "I": 13, "J": 13, "K": 12, "L": 16, "M": 13, "N": 13, "O": 16, "P": 13,
            "Q": 16, "R": 13, "S": 13, "T": 14, "U": 15, "V": 13, "W": 11, "X": 16,
            "Y": 17, "Z": 14, "AA": 20, "AB": 13, "AC": 13, "AD": 16, "AE": 12,
            "AF": 13, "AG": 12, "AH": 13, "AI": 10, "AJ": 11, "AK": 28,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        last_data_row = data_row_idx
        for row_idx in range(1, last_data_row + 1):
            ws.row_dimensions[row_idx].height = 24
        ws.row_dimensions[1].height = 18
        ws.row_dimensions[header_top_row].height = 34
        ws.row_dimensions[header_sub_row].height = 34
        ws.freeze_panes = f"A{data_row_idx}"
        ws.sheet_view.zoomScale = 75

    def _write_metadata_index_sheet(self, workbook) -> None:
        """Ajoute une feuille technique clé/valeur relue en priorité."""
        from openpyxl.styles import Font, PatternFill

        if SHEET_METADATA_INDEX in workbook.sheetnames:
            del workbook[SHEET_METADATA_INDEX]
        ws = workbook.create_sheet(SHEET_METADATA_INDEX, 1)
        ws.append(["Clé", "Champ", "Valeur"])

        header_fill = PatternFill("solid", fgColor="D9E2F3")
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill

        for _, row in self._metadata_index_dataframe().iterrows():
            ws.append([row["Clé"], row["Champ"], row["Valeur"]])

        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 38
        ws.column_dimensions["C"].width = 50
        ws.freeze_panes = "A2"
        ws.sheet_state = "hidden"

    def _default_metadata_filename(self) -> str:
        base_name = self.edit_manip.text().strip() if hasattr(self, "edit_manip") else ""
        if not base_name:
            base_name = "metadonnees"

        suffixes = []
        if hasattr(self, "chk_titration_done") and self.chk_titration_done.isChecked():
            suffixes.append("T")
        if hasattr(self, "chk_bacterio_analysis") and self.chk_bacterio_analysis.isChecked():
            suffixes.append("AnBa")
        if hasattr(self, "chk_ammonium_test") and self.chk_ammonium_test.isChecked():
            suffixes.append("TAm")
        if hasattr(self, "chk_turbidity") and self.chk_turbidity.isChecked():
            suffixes.append("Tur")
        if hasattr(self, "chk_conductivity") and self.chk_conductivity.isChecked():
            suffixes.append("Cond")
        if hasattr(self, "chk_ph_water") and self.chk_ph_water.isChecked():
            suffixes.append("pH")
        if hasattr(self, "chk_dissolved_o2") and self.chk_dissolved_o2.isChecked():
            suffixes.append("Ox")

        filename = base_name + ("_" + "_".join(suffixes) if suffixes else "")
        filename = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', "_", filename)
        filename = re.sub(r"\s+", "_", filename).strip(" ._")
        filename = re.sub(r"_+", "_", filename).strip("_")
        return f"{filename or 'metadonnees'}.xlsx"

    def _on_save_metadata_clicked(self) -> None:
        """Enregistre df_comp et df_map dans un fichier Excel."""
        warnings = self._metadata_save_warnings()
        if warnings:
            QMessageBox.warning(
                self,
                "Points à vérifier",
                "L'enregistrement va continuer, mais certains points ne sont pas complets :\n\n"
                + "\n".join(f"• {warning}" for warning in warnings),
            )

        path, _ = QFileDialog.getSaveFileName(
            self,
            "Enregistrer les métadonnées",
            self._default_metadata_filename(),
            "Fichier Excel (*.xlsx)"
        )
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        try:
            if self.df_map is not None and isinstance(self.df_map, pd.DataFrame) and not self.df_map.empty:
                df_map_to_save = self._df_map_with_current_header(self.df_map)
            else:
                df_map_to_save = self._header_rows_dataframe()
            if self.df_comp is not None and isinstance(self.df_comp, pd.DataFrame):
                df_comp_to_save = self.df_comp
            else:
                df_comp_to_save = pd.DataFrame()
            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                df_comp_to_save.to_excel(writer, sheet_name=SHEET_TITRATION_VOLUMES, index=False)
                df_map_to_save.to_excel(writer, sheet_name=SHEET_TITRATION_CORRESPONDENCE, index=False)
                # État du protocole de paillasse
                if self._protocol_states:
                    rows = []
                    for key, val in self._protocol_states.items():
                        if key == "__notes__":
                            rows.append({"type": "__notes__", "r_idx": -1, "t_idx": -1, "checked": str(val)})
                        elif key[0] == "verif":
                            rows.append({"type": "verif", "r_idx": key[1], "t_idx": -1, "checked": int(val)})
                        else:
                            rows.append({"type": key[0], "r_idx": key[1], "t_idx": key[2], "checked": int(val)})
                    pd.DataFrame(rows).to_excel(writer, sheet_name=SHEET_TITRATION_PROTOCOL_STATE, index=False)
                self._write_measure_summary_sheet(writer.book)
                self._write_metadata_index_sheet(writer.book)
        except Exception as e:
            QMessageBox.critical(
                self,
                "Erreur d'enregistrement",
                f"Impossible d'enregistrer les métadonnées dans le fichier :\n{e}",
            )
            return

        self._mark_metadata_saved()
        QMessageBox.information(self, "Enregistrement réussi", f"Métadonnées enregistrées dans :\n{path}")

    def _on_load_metadata_clicked(self) -> None:
        """Charge df_comp et df_map à partir d'un fichier Excel précédemment sauvegardé."""
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Charger des métadonnées",
            "",
            "Fichier Excel (*.xlsx)"
        )
        if not path:
            return

        try:
            xls = pd.ExcelFile(path)
            loaded_comp = False
            loaded_map = False
            loaded_metadata = False
            summary_values = (
                self._header_values_from_metadata_index_sheet(xls)
                or self._header_values_from_measure_summary_sheet(xls)
            )
            # Volumes
            volumes_sheet = _first_existing_sheet(xls.sheet_names, SHEET_TITRATION_VOLUMES_ALIASES)
            if volumes_sheet:
                self.df_comp = self._clean_loaded_volume_dataframe(pd.read_excel(xls, sheet_name=volumes_sheet))
                loaded_comp = self.df_comp is not None and isinstance(self.df_comp, pd.DataFrame) and not self.df_comp.empty
                if hasattr(self, "chk_titration_done") and not self.df_comp.empty:
                    self.chk_titration_done.setChecked(True)
                try:
                    conc_df = mm.compute_concentration_table(self.df_comp)
                    tube_cols = self._get_tube_columns(self.df_comp)
                    if tube_cols and not conc_df.empty:
                        self._sum_target_uL = float(conc_df.at[0, tube_cols[0]])
                except Exception:
                    self._sum_target_uL = DEFAULT_VTOT_UL
                self.lbl_status_comp.setText(
                    f"Tableau des volumes : {self.df_comp.shape[0]} ligne(s), {self.df_comp.shape[1]} colonne(s)"
                )
            # Correspondance
            correspondence_sheet = _first_existing_sheet(
                xls.sheet_names,
                SHEET_TITRATION_CORRESPONDENCE_ALIASES,
            )
            if correspondence_sheet:
                loaded_df_map = pd.read_excel(xls, sheet_name=correspondence_sheet)
                if {"Nom du spectre", "Tube"}.issubset(loaded_df_map.columns):
                    self.df_map = loaded_df_map
                    if summary_values:
                        self.df_map = self._merge_header_values_into_df_map(self.df_map, summary_values)
                    loaded_map = True
                    loaded_metadata = True
                    self.lbl_status_map.setText(
                        f"Tableau de correspondance : {self.df_map.shape[0]} ligne(s), {self.df_map.shape[1]} colonne(s)"
                    )
                    # Mettre à jour les champs d'en-tête d'après df_map
                    self._update_header_fields_from_df_map()
                    self.map_dirty = False
                elif summary_values:
                    self.df_map = self._merge_header_values_into_df_map(pd.DataFrame(), summary_values)
                    loaded_metadata = True
                    self.lbl_status_map.setText(
                        "Métadonnées terrain restaurées depuis la feuille Mesures ; correspondance invalide"
                    )
                    self._update_header_fields_from_df_map()
                    self.map_dirty = True
                    QMessageBox.warning(
                        self,
                        "Correspondance invalide",
                        f"La feuille '{correspondence_sheet}' ne contient pas les colonnes attendues.\n"
                        "Les champs disponibles ont été restaurés depuis la feuille 'Mesures', "
                        "mais la correspondance spectres ↔ tubes devra être recréée.",
                    )
                else:
                    QMessageBox.warning(
                        self,
                        "Correspondance invalide",
                        f"La feuille '{correspondence_sheet}' ne contient pas les colonnes attendues "
                        "('Nom du spectre' et 'Tube').",
                    )
            elif summary_values:
                self.df_map = self._merge_header_values_into_df_map(pd.DataFrame(), summary_values)
                loaded_metadata = True
                self.lbl_status_map.setText(
                    "Métadonnées terrain restaurées depuis la feuille Mesures ; correspondance absente"
                )
                self._update_header_fields_from_df_map()
                self.map_dirty = True
                QMessageBox.warning(
                    self,
                    "Feuille manquante",
                    f"La feuille '{SHEET_TITRATION_CORRESPONDENCE}' est absente du fichier Excel chargé.\n"
                    "Les champs disponibles ont été restaurés depuis la feuille 'Mesures', "
                    "mais la correspondance spectres ↔ tubes devra être recréée.",
                )
            else:
                QMessageBox.warning(
                    self,
                    "Feuille manquante",
                    f"La feuille '{SHEET_TITRATION_CORRESPONDENCE}' est absente du fichier Excel chargé.\n"
                    "Le tableau de correspondance n'a pas pu être restauré.",
                )
            # État du protocole de paillasse
            self._protocol_states = {}
            protocol_sheet = _first_existing_sheet(xls.sheet_names, SHEET_TITRATION_PROTOCOL_STATE_ALIASES)
            if protocol_sheet:
                df_proto = pd.read_excel(xls, sheet_name=protocol_sheet)
                for _, r in df_proto.iterrows():
                    t = str(r.get("type", ""))
                    if t == "__notes__":
                        self._protocol_states["__notes__"] = str(r.get("checked", ""))
                        continue
                    ri  = int(r.get("r_idx", 0))
                    ti  = int(r.get("t_idx", -1))
                    chk = bool(r.get("checked", 0))
                    if t == "verif":
                        self._protocol_states[("verif", ri)] = chk
                    elif t in ("coord", "oper"):
                        self._protocol_states[(t, ri, ti)] = chk
                    elif t == "instruction":
                        self._protocol_states[("instruction", ri, ti)] = chk
            if loaded_comp and loaded_map:
                self._mark_metadata_saved()
            else:
                self._mark_metadata_dirty(refresh=False)
                self._refresh_button_states()
            if loaded_metadata:
                self._refresh_fillable_fields_style()
                self._refresh_button_states()
        except Exception as e:
            QMessageBox.critical(
                self,
                "Erreur de chargement",
                f"Impossible de charger les métadonnées depuis le fichier :\n{e}",
            )
    def _on_edit_comp_clicked(self) -> None:
        """Ouvre le dialogue d'édition du tableau des volumes.

        Règles :
        - Si un tableau existe déjà (self.df_comp), on le ré-ouvre tel quel.
        - Sinon, on propose un modèle par défaut ou pré-rempli selon le preset sélectionné.
        - À la validation, on enregistre dans self.df_comp.
        - Si l'utilisateur n'a PAS édité les concentrations (conc_user_overridden=False),
          on recalcule self.df_conc depuis les volumes.
        """

        # Colonnes de base
        base_cols = ["Réactif", "Concentration", "Unité"]

        # 1) Si un tableau existe déjà, on le ré-ouvre tel quel
        if self.df_comp is not None and not self.df_comp.empty:
            initial_df = self.df_comp.copy()

        # 2) Sinon, on charge le modèle sélectionné
        else:
            preset_name = self.combo_volume_preset.currentText()
            preset_df = self.VOLUME_PRESETS.get(preset_name)
            if preset_df is None:
                QMessageBox.warning(self, "Modèle inconnu", "Le modèle sélectionné est introuvable.")
                return
            initial_df = preset_df.copy()

        # Harmonise les alias fréquents pour les colonnes standard (Conc/Unit -> Concentration/Unité)
        def _norm_col(col: str) -> str:
            return (
                str(col)
                .strip()
                .lower()
                .replace("é", "e")
                .replace("è", "e")
                .replace("ê", "e")
            )

        # Renommer si les colonnes standard manquent
        rename_map: dict[str, str] = {}
        has_conc = "Concentration" in initial_df.columns
        has_unit = "Unité" in initial_df.columns

        for col in list(initial_df.columns):
            n = _norm_col(col)
            if n == "conc" and not has_conc:
                rename_map[col] = "Concentration"
                has_conc = True
            if n in ("unit", "units", "unite") and not has_unit:
                rename_map[col] = "Unité"
                has_unit = True

        if rename_map:
            initial_df = initial_df.rename(columns=rename_map)

        # Fusionner d'éventuelles colonnes alias supplémentaires avec les colonnes standard
        conc_aliases = [c for c in initial_df.columns if _norm_col(c) == "conc" and c != "Concentration"]
        for alias in conc_aliases:
            if "Concentration" not in initial_df.columns:
                initial_df = initial_df.rename(columns={alias: "Concentration"})
            else:
                initial_df["Concentration"] = initial_df["Concentration"].fillna(initial_df[alias])
                initial_df = initial_df.drop(columns=[alias])

        unit_aliases = [
            c for c in initial_df.columns if _norm_col(c) in ("unit", "units", "unite") and c != "Unité"
        ]
        for alias in unit_aliases:
            if "Unité" not in initial_df.columns:
                initial_df = initial_df.rename(columns={alias: "Unité"})
            else:
                initial_df["Unité"] = initial_df["Unité"].fillna(initial_df[alias])
                initial_df = initial_df.drop(columns=[alias])

        # Colonnes de tubes dynamiques : on reprend tout ce qui existe déjà pour ne rien perdre
        tube_cols = self._get_tube_columns(initial_df)
        if not tube_cols:
            tube_cols = [f"Tube {i}" for i in range(1, 11)]

        default_cols = base_cols + tube_cols

        # Sécurité : garantir les colonnes standard
        for c in default_cols:
            if c not in initial_df.columns:
                initial_df[c] = pd.NA

        # Conserver d'éventuelles colonnes supplémentaires ajoutées par l'utilisateur
        extra_cols = [c for c in initial_df.columns if c not in default_cols]
        initial_df = initial_df[default_cols + extra_cols]

        dlg = TableEditorDialog(
            title="Tableau des volumes",
            columns=list(initial_df.columns),
            initial_df=initial_df,
            parent=self,
            sum_row=True,
            sum_target=self._sum_target_uL,
        )
        # ---------------------------------------------------------
        # Synchronisation Solution A2 + Solution B = constante
        # ---------------------------------------------------------
        # On impose :
        #     Solution A2 + Solution B = constante (par tube)
        # La constante est définie à l'ouverture du tableau.
        # Si l'utilisateur modifie A → B s'ajuste, et inversement.
        table = dlg.table

        def _norm_name(txt: str) -> str:
            return (
                str(txt)
                .strip()
                .lower()
                .replace("é", "e")
                .replace("è", "e")
                .replace("ê", "e")
            )

        def _reactif_col_index() -> int:
            for j in range(table.columnCount()):
                header = table.horizontalHeaderItem(j)
                if header is None:
                    continue
                name = _norm_name(header.text())
                if name in ("reactif", "reactifs"):
                    return j
            return 0

        def _is_tube_col(col_idx: int) -> bool:
            header = table.horizontalHeaderItem(col_idx)
            if header is None:
                return False
            return header.text().strip().lower().startswith("tube")

        def _find_rows_ab() -> tuple[int | None, int | None, int]:
            col_reactif = _reactif_col_index()
            row_a = None
            row_b = None
            sum_row = getattr(dlg, "_sum_row_idx", None)
            for r in range(table.rowCount()):
                if sum_row is not None and r == sum_row:
                    continue
                item = table.item(r, col_reactif)
                if item is None:
                    continue
                name = _norm_name(item.text())
                if name == "solution a2":
                    row_a = r
                elif name == "solution b":
                    row_b = r
            return row_a, row_b, col_reactif

        def _get_pas_B(row_b: int) -> float:
            """Retourne le pas (µL/goutte) de Solution B (0 si pas de mode goutte)."""
            for j in range(table.columnCount()):
                hdr = table.horizontalHeaderItem(j)
                if hdr is not None and hdr.text().strip() == "Pas (µL)":
                    pas_item = table.item(row_b, j)
                    return dlg._parse_float(pas_item.text() if pas_item is not None else "") or 0.0
            return 0.0

        # Somme cible A+B (en µL) pour chaque tube — initialisée à l'ouverture
        target_sum: dict[int, float] = {}
        row_A, row_B, _ = _find_rows_ab()
        if row_A is not None and row_B is not None:
            pas_A_init = _get_pas_B(row_A)
            pas_B_init = _get_pas_B(row_B)
            for c in range(table.columnCount()):
                if not _is_tube_col(c):
                    continue
                itemA = table.item(row_A, c)
                itemB = table.item(row_B, c)
                vA_raw = dlg._parse_float(itemA.text() if itemA is not None else "") or 0.0
                vB_raw = dlg._parse_float(itemB.text() if itemB is not None else "") or 0.0
                # En mode goutte, la valeur brute = nombre de gouttes → convertir en µL
                vA_uL = vA_raw * pas_A_init if pas_A_init > 0 else vA_raw
                vB_uL = vB_raw * pas_B_init if pas_B_init > 0 else vB_raw
                target_sum[c] = vA_uL + vB_uL

        def sync_A_B(item: QTableWidgetItem) -> None:
            if item is None:
                return
            # Saisie manuelle activée → ne rien synchroniser
            if dlg.chk_manual_mode.isChecked():
                return
            r = item.row()
            c = item.column()
            if not _is_tube_col(c):
                return
            sum_row = getattr(dlg, "_sum_row_idx", None)
            if sum_row is not None and r == sum_row:
                return

            row_A, row_B, _ = _find_rows_ab()
            if row_A is None or row_B is None:
                return
            if r not in (row_A, row_B):
                return

            val = dlg._parse_float(item.text())
            if val is None:
                return

            pas_A = _get_pas_B(row_A)
            pas_B = _get_pas_B(row_B)

            if c not in target_sum:
                itemA = table.item(row_A, c)
                itemB = table.item(row_B, c)
                vA_raw = dlg._parse_float(itemA.text() if itemA is not None else "") or 0.0
                vB_raw = dlg._parse_float(itemB.text() if itemB is not None else "") or 0.0
                vA_uL = vA_raw * pas_A if pas_A > 0 else vA_raw
                vB_uL = vB_raw * pas_B if pas_B > 0 else vB_raw
                target_sum[c] = vA_uL + vB_uL

            total = target_sum[c]  # toujours en µL

            if r == row_B:
                # L'utilisateur a modifié Solution B
                # val = nombre de gouttes (si compte-goutte) ou µL sinon
                val_uL = val * pas_B if pas_B > 0 else val
                remaining_uL = total - val_uL
                # Solution A2 : reconvertir en gouttes si nécessaire
                new_val_A = round(remaining_uL / pas_A) if pas_A > 0 else remaining_uL
                other_row, new_val = row_A, new_val_A
            else:
                # L'utilisateur a modifié Solution A2
                val_uL = val * pas_A if pas_A > 0 else val
                remaining_uL = total - val_uL
                # Solution B : reconvertir en gouttes si nécessaire
                new_val_B = round(remaining_uL / pas_B) if pas_B > 0 else remaining_uL
                other_row, new_val = row_B, new_val_B

            table.blockSignals(True)
            try:
                other_item = table.item(other_row, c)
                if other_item is None:
                    other_item = QTableWidgetItem("")
                    table.setItem(other_row, c, other_item)
                other_item.setText(dlg._format_number(new_val))
            finally:
                table.blockSignals(False)

            # Met à jour la ligne "Total" après ajustement automatique
            try:
                dlg._refresh_sum_row()
            except Exception:
                pass

        table.itemChanged.connect(sync_A_B)
        if dlg.exec() != QDialog.Accepted:
            return

        df = dlg.to_dataframe()
        if df.empty:
            QMessageBox.warning(self, "Tableau vide", "Le tableau des volumes est vide.")
            return

        # Nettoyage léger
        if "Réactif" in df.columns:
            df["Réactif"] = df["Réactif"].astype(str).str.strip()

        self.df_comp = df
        if hasattr(self, "chk_titration_done"):
            self.chk_titration_done.setChecked(True)
        self.lbl_status_comp.setText(
            f"Tableau des volumes : {df.shape[0]} ligne(s), {df.shape[1]} colonne(s)"
        )

        self._sync_df_map_with_df_comp()
